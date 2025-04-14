import random
import re
from Execution import *
from TestCase import *
import itertools
from openpyxl import load_workbook
import copy
import sys
import time
import json
from DNA import *
from publicFun import *
import shutil
random.seed(1)


def copyFile(fileDir, tarDir):
    pathDir = os.listdir(fileDir)  # 取图片的原始路径
    filenumber = len(pathDir)
    rate = 0.1  # 自定义抽取图片的比例，比方说100张抽10张，那就是0.1
    picknumber = int(filenumber * rate)  # 按照rate比例从文件夹中取一定数量图片
    sample = random.sample(pathDir, picknumber)  # 随机选取picknumber数量的样本图片
    for i in range(len(sample)):
        shutil.copy(fileDir + sample[i], tarDir + "infile{}.txt".format(i))
    return


def getOriginalInput():
    ts = TestCase()
    for i in range(1000):
        ts.setInputOutput("infile{}.txt".format(i), "outfile{}.txt".format(i), "outtree{}.txt".format(i))
        ts.generateRandomTestcase()
    fileDir = '/Applications/work/data/MT/MFT/'+string+'/input/'  # 源文件夹路径
    tarDir = '/Applications/work/data/MT/MFT/'+string+'/RandomInput/'  # 移动到新的文件夹路径
    shutil.rmtree(tarDir)
    os.mkdir(tarDir)
    copyFile(fileDir, tarDir)


def FailureRate(mu):
    # 把SourceCases读出来
    Result = []
    for i in range(1000):
        ts.setInputOutput("infile_{}.txt".format(i), "outfile_{}_{}.txt".format(0, i), "outtree_{}_{}.txt".format(0, i))
        original_output = MR().getResults(ts)
        program_ts = ts
        program_ts.setInputOutput("infile_{}.txt".format(i), "outfile_{}_{}.txt".format(mu, i), "outtree_{}_{}.txt".format(mu, i))
        program_output = MR().getResults(program_ts)
        isViolate = MR().assertViolation(original_output, program_output)

        if isViolate:
            Result.append(1)
        else:
            Result.append(0)
    FR = round(Result.count(1) / len(Result) * 100, 2)
    return FR


def recordResult(file_name, mutants_list, mr_list):
    result = open("../results/"+file_name, "w")
    temp = [v+"\t" for v in mutants_list]
    temp.insert(0, "\t")
    temp.append("\n")
    for cmr in mr_list:
        temp.append("{}\t".format(cmr.name))
        for v in mutants_list:
            temp.append(str(cmr.table[v])+"\t")
        temp.append("\n")
    result.writelines(temp)
    result.close()


def getSource(mr_list, test_case, num_of_samples):
    # for i in range(num_of_samples):
    #     test_case.setInputOutput("infile_{}".format(i), "outfile_{}".format(i), "outtree_{}".format(i))
    #     test_case.generateRandomTestcase()
    for i in range(num_of_samples):
        for j in range(len(mr_list)):
            mr = mr_list[j]
            mr.setTestCase(test_case)
            mr.original_ts.setInputOutput("infile{}.txt".format(i), "outfile_{}.txt".format(i), "outtree_{}.txt".format(i))
            mr.getFollow(j)
            for k in range(len(mr_list)):
                mr = mr_list[k]
                mr.setTestCase(test_case)
                mr.original_ts.setInputOutput("infile{}_{}.txt".format(i, j), "outfile_{}.txt".format(i),
                                              "outtree_{}.txt".format(i))
                mr.getFollow(k)
    # for i in range(num_of_samples):
    #     for j in range(len(mr_list)):
    #         if j < 5:
    #             for k in range(5, len(mr_list)):
    #                 mr = mr_list[k]
    #                 mr.setTestCase(test_case)
    #                 mr.original_ts.setInputOutput("infile{}_{}.txt".format(i, j), "outfile_{}.txt".format(i), "outtree_{}.txt".format(i))
    #                 mr.getFollow(k)
    #         else:
    #             for k in range(len(mr_list)):
    #                 mr = mr_list[k]
    #                 mr.setTestCase(test_case)
    #                 mr.original_ts.setInputOutput("infile_{}_{}.txt".format(i, j), "outfile_{}.txt".format(i), "outtree_{}.txt".format(i))
    #                 mr.getFollow(k)


def MetamorphicTesting(executor, mutants, mr_list, test_case, num_of_samples):
    MGS = []
    for i in range(num_of_samples):
        MG = [0] * len(mr_list)
        for j in range(len(mr_list)):
            mr = mr_list[j]
            mr.setTestCase(test_case)
            mr.original_ts.setInputOutput("infile_{}".format(i), "outfile_{}".format(i), "outtree_{}".format(i))
            mr.getFollow(j)
            mr.setExecutor(executor)
            executor.setVersion(mutants)
            mr.process()
            if mr.isViolate:
                MG[j] = 1
        MGS.append(MG)
    return MGS


def getMG(mu, ts, num_of_samples, mr_list, PFS):
    MGS = []
    SMGS = []
    for i in range(num_of_samples):
        MGs = []
        MG = [0] * len(mr_list)
        ts.setInputOutput("infile{}.txt".format(i), "outfile_{}_{}.txt".format(mu, i), "outtree_{}_{}.txt".format(mu, i))
        # ts.setInputOutput("infile_{}".format(i), "outfile_{}".format(i), "outtree_{}".format(i))
        original_output = MR().getResults(ts)
        followup_ts = ts
        for j in range(len(mr_list)):
            # followup_ts.setInputOutput("infile_{}_{}_f".format(i, j), "outfile_{}_{}".format(i, j), "outtree_{}_{}".format(i, j))
            followup_ts.setInputOutput("infile{}_{}.txt".format(i, j), "outfile_{}_{}_{}.txt".format(mu, i, j),
                                        "outtree_{}_{}_{}.txt".format(mu, i, j))
            mr = mr_list[j]
            followup_output = mr.getResults(followup_ts)
            expected_output = mr.getExpectedOutput(original_output)
            isViolate = mr.assertViolation(expected_output, followup_output)
            if isViolate:
                MG[j] = 1
        MGs.append(MG)
        for m in range(len(mr_list)):
            MG = [0] * len(mr_list)
            ts.setInputOutput("infile{}_{}.txt".format(i, m), "outfile_{}_{}_{}.txt".format(mu, i, m),
                                           "outtree_{}_{}_{}.txt".format(mu, i, m))
            # ts.setInputOutput("infile_{}".format(i), "outfile_{}".format(i), "outtree_{}".format(i))
            original_output = MR().getResults(ts)
            followup_ts = ts
            for n in range(len(mr_list)):
                # followup_ts.setInputOutput("infile_{}_{}_f".format(i, j), "outfile_{}_{}".format(i, j), "outtree_{}_{}".format(i, j))
                followup_ts.setInputOutput("infile{}_{}_{}.txt".format(i, m, n), "outfile_{}_{}_{}_{}.txt".format(mu, i, m, n),
                                           "outtree_{}_{}_{}_{}.txt".format(mu, i, m, n))
                mr = mr_list[n]
                followup_output = mr.getResults(followup_ts)
                expected_output = mr.getExpectedOutput(original_output)
                isViolate = mr.assertViolation(expected_output, followup_output)
                if isViolate:
                    MG[n] = 1
            MGs.append(MG)
        # 去掉巧合满足性
        for i1 in range(len(MGs)):
            for j1 in range(len(MGs[i1])):
                if MGs[i1][j1] == 0 and (PFS[i][i1] or PFS[i][i1 * len(MGs[0]) + j1 + 1]):  # 如果satisfied
                    MGs[i1][j1] = 3

        SMGs = copy.deepcopy(MGs)
        # 随机去除一些MG
        for i2 in range(1, len(MGs)):  # 第一组不变
            t = random.randint(1, len(MGs[i2]) - 1)  # 去几个
            a = [n for n in range(len(MGs[i2]))]
            random.shuffle(a)
            b = a[:t]
            for j2 in b:
                MGs[i2][j2] = 4

        SMGS.append(SMGs)
        MGS.append(MGs)
    return MGS, SMGS


def getpf(mu, ts, num_of_samples, mr_list):
    pf = []
    Output = []
    for i in range(num_of_samples):
        output = []
        result = [0] * (len(mr_list) * (len(mr_list) + 1) + 1)
        ts.setInputOutput("infile{}.txt".format(i), "outfile_{}_{}.txt".format(0, i), "outtree_{}_{}.txt".format(0, i))
        original_output = MR().getResults(ts)
        program_ts = ts
        program_ts.setInputOutput("infile{}.txt".format(i), "outfile_{}_{}.txt".format(mu, i), "outtree_{}_{}.txt".format(mu, i))
        program_output = MR().getResults(program_ts)
        output.append([original_output.tree, original_output.total_length])
        output.append([program_output.tree, program_output.total_length])
        isViolate = MR().assertViolation(original_output, program_output)
        if isViolate:
            result[0] = 1
        for j in range(len(mr_list)):
            ts.setInputOutput("infile{}_{}.txt".format(i, j), "outfile_{}_{}_{}.txt".format(0, i, j),
                                        "outtree_{}_{}_{}.txt".format(0, i, j))
            original_output = MR().getResults(ts)
            program_ts = ts
            program_ts.setInputOutput("infile{}_{}.txt".format(i, j), "outfile_{}_{}_{}.txt".format(mu, i, j),
                                        "outtree_{}_{}_{}.txt".format(mu, i, j))
            program_output = MR().getResults(program_ts)
            output.append([program_output.tree, program_output.total_length])
            isViolate = MR().assertViolation(original_output, program_output)
            if isViolate:
                result[j + 1] = 1

        Output.append(output)

        for m in range(len(mr_list)):
            for n in range(len(mr_list)):
                ts.setInputOutput("infile{}_{}_{}.txt".format(i, m, n), "outfile_{}_{}_{}_{}.txt".format(0, i, m, n),
                                    "outtree_{}_{}_{}_{}.txt".format(0, i, m, n))
                original_output = MR().getResults(ts)
                program_ts = ts
                program_ts.setInputOutput("infile{}_{}_{}.txt".format(i, m, n), "outfile_{}_{}_{}_{}.txt".format(mu, i, m, n),
                                           "outtree_{}_{}_{}_{}.txt".format(mu, i, m, n))
                program_output = MR().getResults(program_ts)
                isViolate = MR().assertViolation(original_output, program_output)
                if isViolate:
                    result[len(mr_list) + 1 + m * len(mr_list) + n] = 1
        pf.append(result)
    return pf, Output


def riskIndex(MGS):
    index = []
    # 统计全部的v和s
    sum_s = 0
    sum_v = 0
    for i in range(len(MGS)):
        sum_s += MGS[i].count(0)
        sum_v += MGS[i].count(1)
    if not sum_s + sum_v == (len(MGS) * len(MGS[0])):
        print('异常')

    for i in range((len(MGS[0]) * (len(MGS[0]) + 1) + 1)):
        if i == 0:
            es = MGS[i].count(0)
            ev = MGS[i].count(1)
            ns = sum_s - es
            nv = sum_v - ev
        elif i <= len(MGS[0]):
            es = MGS[i].count(0)
            ev = MGS[i].count(1)
            ns = sum_s - es
            nv = sum_v - ev
            if MGS[0][i - 1] == 0:
                es += 1
                ns -= 1
            else:
                ev += 1
                nv -= 1
        else:
            t = int((i - 1) / len(MGS[0]))
            if MGS[t][(i - 1) % len(MGS[0])] == 0:
                es = 1
                ns = sum_s - es
                ev = 0
                nv = sum_v
            else:
                ev = 1
                nv = sum_v - ev
                es = 0
                ns = sum_s
        index.append([ev, es, nv, ns])
    return index


def getRisk(Index):
    Risk = []
    for i in range(len(Index)):
        risk = []
        index = Index[i]
        for j in range(len(index)):
            try:
                formula = riskformula(index[j])
                risk.append(formula)
            except:
                print(i, index)
        Risk.append(risk)
    return Risk


if __name__ == "__main__":
    string = 'DNA'
    # myenv = MyEnv()
    # myenv.CreateWorkingDirs()
    row = 1
    path = '/Applications/work/data/MT/MFT/Result/result'+sys.argv[1]+'.xlsx'  # '+sys.argv[1][:-1]+'  '+sys.argv[1]+'
    wb = load_workbook(path)
    del wb[string]
    ws = wb.create_sheet(string)
    # getOriginalInput()
    # ts = TestCase()
    # num_of_samples = 100  # 测试用例个数
    # cmr1 = CompositionMR()
    # cmr2 = CompositionMR()
    # cmr3 = CompositionMR()
    # cmr4 = CompositionMR()
    # cmr5 = CompositionMR()
    # cmr6 = CompositionMR()
    # cmr1.setMRs([MR1(), MR2()])
    # cmr2.setMRs([MR2(), MR1()])
    # cmr3.setMRs([MR1(), MR3()])
    # cmr4.setMRs([MR3(), MR1()])
    # cmr5.setMRs([MR2(), MR3()])
    # cmr6.setMRs([MR3(), MR2()])
    # mr_list = [MR1(), MR2(), MR3(), MR4(), MR6(), cmr1, cmr2, cmr3, cmr4, cmr5, cmr6]
    # getSource(mr_list, ts, num_of_samples)
    # FR = []
    MG_set = []
    PF_set = []
    SMG_set = []  # 不去除
    with open('/Applications/work/data/MT/MFT/' + string + '/FR.json', 'r') as load_f:
        data = json.load(load_f)
    data = json.loads(data)
    FR = data['FR']
    EMR = [0, 1, 2, 4, 5, 6, 7, 8, 9, 10]
    for mu in [2, 3, 6]:  # [2, 3, 6]
        # a = FailureRate(mu)
        # FR.append(a)  # original = 1-failure rate
        # data = {
        #          'FR': FR
        # }
        # # 将数据存下来
        # json_str = json.dumps(data)
        # with open('/Applications/work/data/MT/MFT/' + string + '/FR.json', 'w') as f:
        #     json.dump(json_str, f)
        # PFS, Output = getpf(mu, ts, 100, mr_list)
        # MGS, SMGS = getMG(mu, ts, 100, mr_list, PFS)
        # MG_set.append(MGS)
        # PF_set.append(PFS)
        # SMG_set.append(SMGS)
        # data = {
        #          'PF': PFS, 'MG': MGS, 'SMG': SMGS, 'Output': Output
        # }
        # # 将数据存下来
        # json_str = json.dumps(data)
        # with open('/Applications/work/data/MT/MFT/' + string + '/mutant' + str(mu) + '.json', 'w') as f:
        #     json.dump(json_str, f)
        if 0 < FR[mu-1] < 30:
            with open('/Applications/work/data/MT/MFT/' + string + '/mutant' + str(mu) + '.json', 'r') as load_f:
                data = json.load(load_f)
            data = json.loads(data)
            pf = data['PF']
            MG = data['SMG']
            Output = data['Output']
            for i in range(len(Output)):
                output = [Output[i][0], Output[i][1:][0]]
                for j in range(len(MG[0][0])):
                    if j not in EMR:
                        continue
                    output.append(Output[i][1:][j + 1])
                Output[i] = output
            MGc = []
            for i in range(len(MG)):
                mg = [MG[i][0]]
                for j in range(len(MG[i])):
                    if j in EMR:
                        mg.append(MG[i][j+1])
                MGc.append(mg)
            MG = copy.deepcopy(MGc)
            for i in range(len(MG)):
                for j in range(len(MG[i])):
                    a = []
                    for k in EMR:
                        a.append(MG[i][j][k])
                    MG[i][j] = a
            PF_set.append(pf)
            MG_set.append(MG)
            row = eval('getMetrics_v14')(row, ws, mu, MG, pf, Output, EMR)  # +sys.argv[1][-1]

    wb.save(path)

