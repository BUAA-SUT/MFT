import copy
from TSQ import *
import math
import numpy as np
import random
import json
from publicFun import *
from openpyxl import load_workbook
import sys
random.seed(1)


def getOriginalInput():
    source_case_set = []
    ty = random.randint(1, 6)  # 分别表示六个分支
    while 1:
        a = random.randint(1, 10)
        b = random.randint(1, 10)
        c = random.randint(1, 10)
        triangle = []
        if a >= b + c or b >= a + c or c >= a + b:
            continue
        else:
            Sum = a + b + c
            Max = max(a, b, c)
            Min = min(a, b, c)
            mid = Sum - Max - Min
            # if pow(Max, 2) < pow(mid, 2) + pow(Min, 2):
            #     # 锐角三角形
            #     if Max == mid:
            #         # 顶角小于或等于60度的等腰三角形
            #         if ty == 1:
            #             triangle = [a, b, c]
            #     elif Min == mid:
            #         # 顶角大于60度的等腰三角形
            #         if ty == 2:
            #             triangle = [a, b, c]
            #     else:
            #         # 不规则的锐角三角形，海伦公式计算
            #         if ty == 3:
            #             triangle = [a, b, c]
            # if pow(Max, 2) == pow(mid, 2) + pow(Min, 2):
            #     if ty == 4:
            #         triangle = [a, b, c]
            # if Min == mid:
            #     # 钝角等腰三角形
            #     if ty == 5:
            #         triangle = [a, b, c]
            # else:
            #     # 不规则钝角三角形，Max乘以高除以2
            #     if ty == 6:
            #         triangle = [a, b, c]
            if ty == 1 and pow(Max, 2) < pow(mid, 2) + pow(Min, 2) and Max == mid:
                triangle = [a, b, c]
                ty = random.randint(1, 6)  # 分别表示六个分支
            if ty == 2 and pow(Max, 2) < pow(mid, 2) + pow(Min, 2) and Min == mid:
                triangle = [a, b, c]
                ty = random.randint(1, 6)  # 分别表示六个分支
            if ty == 3 and pow(Max, 2) < pow(mid, 2) + pow(Min, 2):
                triangle = [a, b, c]
                ty = random.randint(1, 6)  # 分别表示六个分支
            if ty == 4 and pow(Max, 2) == pow(mid, 2) + pow(Min, 2):
                triangle = [a, b, c]
                ty = random.randint(1, 6)  # 分别表示六个分支
            if ty == 5 and pow(Max, 2) > pow(mid, 2) + pow(Min, 2) and Min == mid:
                triangle = [a, b, c]
                ty = random.randint(1, 6)  # 分别表示六个分支
            if ty == 6 and pow(Max, 2) > pow(mid, 2) + pow(Min, 2):
                triangle = [a, b, c]
                ty = random.randint(1, 6)  # 分别表示六个分支
        if len(triangle) == 0:
            continue
        source_case_set.append(triangle)
        if len(source_case_set) >= 1000:
            break
    # 随机取100个测试用例
    random_input = random.sample(source_case_set, 100)
    data = {
            'source_case_set': source_case_set,
            'random_input': random_input
    }
    json_str = json.dumps(data)
    with open('/Applications/work/data/MT/MFT/'+string+'/OriginalInputNew.json', 'w') as f:
        json.dump(json_str, f)

    return source_case_set, random_input


def FailureRate(dynamic):
    # 把SourceCases读出来
    Result = []
    with open('/Applications/work/data/MT/MFT/'+string+'/OriginalInputNew.json', 'r') as load_f:
        data = json.load(load_f)
    data = json.loads(data)
    source_case_set = data['source_case_set']
    for i in range(len(source_case_set)):
        result_s_a = Trisquare().trisquare(source_case_set[i])  # oracle
        result_s_m = dynamic.trisquare(source_case_set[i])
        if round(result_s_a[0], 4) == round(result_s_m[0], 4):
            Result.append(0)
        else:
            Result.append(1)
    FR = round(Result.count(1) / len(Result) * 100, 2)
    return FR


def getResult():
    Result = []
    rate = []
    # while 1:
    #     a = random.randint(1, 10)
    #     b = random.randint(1, 10)
    #     c = random.randint(1, 10)
    #     if a >= b + c or b >= a + c or c >= a + b:
    #         continue
    #     else:
    #         triangle = [a, b, c]
    #
    #         result = Trisquare().trisquare2(triangle)
    #         Result.append(result)
    #         if len(Result) >= 1000:
    #             break
    with open('/Applications/work/data/MT/MFT/'+string+'/OriginalInputNew.json', 'r') as load_f:
        data = json.load(load_f)
    data = json.loads(data)
    source_case_set = data['source_case_set']
    for i in range(len(source_case_set)):
        result = Trisquare().trisquare2(source_case_set[i])  # oracle
        Result.append(result)
    rate1 = Result.count(1) / len(Result) * 100
    rate2 = Result.count(2) / len(Result) * 100
    rate3 = Result.count(3) / len(Result) * 100
    rate4 = Result.count(4) / len(Result) * 100
    rate5 = Result.count(5) / len(Result) * 100
    rate6 = Result.count(6) / len(Result) * 100
    rate.append([rate1, rate2, rate3, rate4, rate5, rate6])

    return rate


def riskIndex(argv, dynamic):
    MGS = []
    Result = []
    testcase = []
    output = []
    source_case = argv.copy()
    result = Trisquare().trisquare(source_case)
    output.append(round(result[0], 4))
    testcase.append(source_case)
    MG, follow_case = MTG(source_case, dynamic)  # t1t2t3t4
    for i in range(len(follow_case)):
        testcase.append(follow_case[i])
    for i in range(len(testcase)):
        result = dynamic.trisquare(testcase[i])
        output.append(round(result[0], 4))
    MGS.append(MG)
    for i in range(len(follow_case)):
        MG, ffollow_case = MTG(follow_case[i], dynamic)  # t2t5...
        MGS.append(MG)
        for j in range(len(ffollow_case)):
            testcase.append(ffollow_case[j])
    # MG统计完, testcase统计完
    for i in range(len(testcase)):
        result_s_a = Trisquare().trisquare(testcase[i])  # oracle
        result_s_m = dynamic.trisquare(testcase[i])
        if round(result_s_a[0], 4) == round(result_s_m[0], 4):
            Result.append(0)
        else:
            Result.append(1)

    # 去掉巧合满足性
    for i in range(len(MGS)):
        for j in range(len(MGS[i])):
            if MGS[i][j] == 0 and (Result[i] or Result[i * len(MGS[0]) + j + 1]):  # 如果satisfied
                MGS[i][j] = 3

    SMGS = copy.deepcopy(MGS)

    # 随机去除一些MG
    for i in range(1, len(MGS)):  # 第一组不变
        t = random.randint(1, len(MGS[i])-1)  # 去几个
        a = [n for n in range(len(MGS[i]))]
        random.shuffle(a)
        b = a[:t]
        for j in b:
            MGS[i][j] = 4

    return MGS, Result, SMGS, output


if __name__ == '__main__':
    row = 1
    string = 'TSQ'
    # source_case_set, random_input = getOriginalInput()
    path = '/Applications/work/data/MT/MFT/Result/result'+sys.argv[1]+'.xlsx'  # '+sys.argv[1][:-1]+' '+sys.argv[1]+'
    wb = load_workbook(path)
    del wb[string]
    ws = wb.create_sheet(string)
    # with open('/Applications/work/data/MT/MFT/' + string + '/OriginalInputNew.json', 'r') as load_f:
    #     data = json.load(load_f)
    # data = json.loads(data)
    # random_input = data['random_input']
    # rate = getResult()
    FR = []
    MG_set = []
    PF_set = []
    SMG_set = []  # 不去除
    EMR = [0, 1, 2, 3, 4, 5, 6, 7, 8]
    for mu in range(1, 7):  # 1, 7
        dynamic = TriFactory("Mutant" + str(mu)).getTri()
        a = FailureRate(dynamic)
        FR.append(a)  # original = 1-failure rate
        # MGS = []
        # PFS = []
        # SMGS = []
        # Output = []
        # for i in range(len(random_input)):
        #     MG, PF, SMG, output = riskIndex(random_input[i], dynamic)
        #     MGS.append(MG)
        #     PFS.append(PF)
        #     Output.append(output)
        #     SMGS.append(SMG)
        # MG_set.append(MGS)
        # PF_set.append(PFS)
        # SMG_set.append(SMGS)
        # data = {
        #          'PF': PFS, 'MG': MGS, 'SMG': SMGS, 'Output': Output
        # }
        # # 将数据存下来
        # json_str = json.dumps(data)
        # with open('/Applications/work/data/MT/MFT/' + string + '/mutant' + str(mu) + '.json', 'w') as f:
        #     json.dump(json_str, f)
        if 0 < a < 30:
            with open('/Applications/work/data/MT/MFT/' + string + '/mutant' + str(mu) + '.json', 'r') as load_f:
                data = json.load(load_f)
            data = json.loads(data)
            pf = data['PF']
            MG = data['SMG']
            Output = data['Output']
            PF_set.append(pf)
            MG_set.append(MG)
            row = eval('getMetrics_v14')(row, ws, mu, MG, pf, Output, EMR)  # +sys.argv[1][-1]

    wb.save(path)








