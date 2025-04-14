from MATH import *
from openpyxl import load_workbook
from publicFun import *
import json
import sys
import os
import random
import shutil
random.seed(1)


def copyFile(fileDir, tarDir):
    pathDir = os.listdir(fileDir)  # 取图片的原始路径
    filenumber = len(pathDir)
    rate = 0.1  # 自定义抽取图片的比例，比方说100张抽10张，那就是0.1
    picknumber = int(filenumber * rate)  # 按照rate比例从文件夹中取一定数量图片
    sample = random.sample(pathDir, picknumber)  # 随机选取picknumber数量的样本图片
    for i in range(len(sample)):
        shutil.copy(fileDir + sample[i], tarDir + "input{}.txt".format(i))
    return


def getOriginalInput():
    ts = TestCase()
    for i in range(1000):
        ts.setInputOutput("input{}.txt".format(i), "output{}.txt".format(i))
        ts.generateInput()
    fileDir = "/Applications/work/data/MT/MFT/MATH/input/"  # 源文件夹路径
    tarDir = '/Applications/work/data/MT/MFT/MATH/RandomInput/'  # 移动到新的文件夹路径
    shutil.rmtree(tarDir)
    os.mkdir(tarDir)
    copyFile(fileDir, tarDir)


def FailureRate(mu):
    # 把SourceCases读出来
    Result = []
    for i in range(1000):
        file_a = open('/Applications/work/data/MT/MFT/MATH/output/' + "output{}_{}.txt".format(0, i), 'r')
        file_m = open('/Applications/work/data/MT/MFT/MATH/output/' + "output{}_{}.txt".format(mu, i), 'r')
        result_s_a = int(file_a.readlines()[0])
        result_s_m = int(file_m.readlines()[0])
        if result_s_a == result_s_m:
            Result.append(0)
        else:
            Result.append(1)
    FR = round(Result.count(1) / len(Result) * 100, 2)
    return FR


def getTestcase(mr_list, test_case, num_of_samples):
    for i in range(num_of_samples):
        # test_case.setInputOutput("input{}.txt".format(i), "output{}.txt".format(i))
        # test_case.generateInput()  # 生成原始测试用例
        for j in range(len(mr_list)):
            mr = mr_list[j]
            mr.setTestCase(test_case)
            mr.original_ts.setInputOutput("input{}.txt".format(i), "output{}.txt".format(i))
            mr.getFollowInput(j)
            for k in range(len(mr_list)):
                mr = mr_list[k]
                mr.setTestCase(test_case)
                mr.original_ts.setInputOutput("input{}_{}.txt".format(i, j), "output{}_{}.txt".format(i, j))
                mr.getFollowInput(k)


def getMG(mu, ts, num_of_samples, mr_list, PFS):
    MGS = []
    SMGS = []
    for i in range(num_of_samples):
        MGs = []
        MG = [0] * len(mr_list)
        ts.setInputOutput("input{}.txt".format(i), "output{}_{}.txt".format(mu, i))
        original_output = MR().getResults(ts)
        followup_ts = ts
        for j in range(len(mr_list)):
            followup_ts.setInputOutput("input{}_{}.txt".format(i, j), "output{}_{}_{}.txt".format(mu, i, j))
            mr = mr_list[j]
            followup_output = mr.getResults(followup_ts)
            expected_output = mr.getExpectedOutput(original_output)
            isViolate = mr.assertViolation(expected_output, followup_output)
            if isViolate:
                MG[j] = 1
        MGs.append(MG)
        for m in range(len(mr_list)):
            MG = [0] * len(mr_list)
            ts.setInputOutput("input{}_{}.txt".format(i, m), "output{}_{}_{}.txt".format(mu, i, m))
            original_output = MR().getResults(ts)
            followup_ts = ts
            for n in range(len(mr_list)):
                followup_ts.setInputOutput("input{}_{}_{}.txt".format(i, m, n), "output{}_{}_{}_{}.txt".format(mu, i, m, n))
                mr = mr_list[n]
                followup_output = mr.getResults(followup_ts)
                expected_output = mr.getExpectedOutput(original_output)
                isViolate = mr.assertViolation(expected_output, followup_output)
                if isViolate:
                    MG[n] = 1
            MGs.append(MG)
        # 去掉巧合满足性
        for i1 in range(len(MGs)):
            for j1 in range(len(MGs[i1])):
                if MGs[i1][j1] == 0 and (PFS[i][i1] or PFS[i][i1 * len(MGs[0]) + j1 + 1]):  # 如果satisfied
                    MGs[i1][j1] = 3

        SMGs = copy.deepcopy(MGs)
        # 随机去除一些MG
        for i2 in range(1, len(MGs)):  # 第一组不变
            t = random.randint(1, len(MGs[i2]) - 1)  # 去几个
            a = [n for n in range(len(MGs[i2]))]
            random.shuffle(a)
            b = a[:t]
            for j2 in b:
                MGs[i2][j2] = 4

        MGS.append(MGs)
        SMGS.append(SMGs)

    return MGS, SMGS


def getpf(mu, ts, num_of_samples, mr_list):
    pf = []
    Output = []
    for i in range(num_of_samples):
        output = []
        result = [0] * (len(mr_list) * (len(mr_list) + 1) + 1)
        ts.setInputOutput("input{}.txt".format(i), "output{}_{}.txt".format(0, i))
        original_output = MR().getResults(ts)
        program_ts = ts
        program_ts.setInputOutput("input{}.txt".format(i), "output{}_{}.txt".format(mu, i))
        program_output = MR().getResults(program_ts)
        output.append(int(original_output.value))
        output.append(int(program_output.value))
        isViolate = MR().assertViolation(original_output, program_output)
        if isViolate:
            result[0] = 1

        for j in range(len(mr_list)):
            ts.setInputOutput("input{}_{}.txt".format(i, j), "output{}_{}_{}.txt".format(0, i, j))
            original_output = MR().getResults(ts)
            program_ts = ts
            program_ts.setInputOutput("input{}_{}.txt".format(i, j), "output{}_{}_{}.txt".format(mu, i, j))
            program_output = MR().getResults(program_ts)
            if j == 2 or j == 5 or j == 6:
                output.append(int(program_output.value) / 2)
            else:
                output.append(int(program_output.value))
            isViolate = MR().assertViolation(original_output, program_output)
            if isViolate:
                result[j + 1] = 1
                # output.append(int(program_output.value) / 2)
            # else:
                # output.append(output[0])

        Output.append(output)

        for m in range(len(mr_list)):
            for n in range(len(mr_list)):
                ts.setInputOutput("input{}_{}_{}.txt".format(i, m, n),
                                           "output{}_{}_{}_{}.txt".format(0, i, m, n))
                original_output = MR().getResults(ts)
                program_ts = ts
                program_ts.setInputOutput("input{}_{}_{}.txt".format(i, m, n), "output{}_{}_{}_{}.txt".format(mu, i, m, n))
                program_output = MR().getResults(program_ts)
                isViolate = MR().assertViolation(original_output, program_output)
                if isViolate:
                    result[len(mr_list) + 1 + m * len(mr_list) + n] = 1
        pf.append(result)
    return pf, Output


def riskIndex(MGS):
    index = []
    # 统计全部的v和s
    sum_s = 0
    sum_v = 0
    for i in range(len(MGS)):
        sum_s += MGS[i].count(0)
        sum_v += MGS[i].count(1)
    if not sum_s + sum_v == (len(MGS) * len(MGS[0])):
        print('异常')

    for i in range((len(MGS[0]) * (len(MGS[0]) + 1) + 1)):
        if i == 0:
            es = MGS[i].count(0)
            ev = MGS[i].count(1)
            ns = sum_s - es
            nv = sum_v - ev
        elif i <= len(MGS[0]):
            es = MGS[i].count(0)
            ev = MGS[i].count(1)
            ns = sum_s - es
            nv = sum_v - ev
            if MGS[0][i - 1] == 0:
                es += 1
                ns -= 1
            else:
                ev += 1
                nv -= 1
        else:
            t = int((i - 1) / len(MGS[0]))
            if MGS[t][(i - 1) % len(MGS[0])] == 0:
                es = 1
                ns = sum_s - es
                ev = 0
                nv = sum_v
            else:
                ev = 1
                nv = sum_v - ev
                es = 0
                ns = sum_s
        index.append([ev, es, nv, ns])
    return index


def getRisk(Index):
    Risk = []
    for i in range(len(Index)):
        risk = []
        index = Index[i]
        for j in range(len(index)):
            try:
                formula = riskformula(index[j])
                risk.append(formula)
            except:
                print(i, index)
        Risk.append(risk)
    return Risk


if __name__ == "__main__":
    # myenv = MyEnv()
    # myenv.CreateWorkingDirs()
    # # getOriginalInput()
    # mr_list = [MR10(), MR11(), MR12()]
    # mr_list = [MR1(), MR2(), MR3(), MR4(), MR5(), MR6(), MR7(), MR8(), MR9(), MR10(), MR11(), MR12()]
    # ts = TestCase()
    # # getTestcase(mr_list, ts, 100)
    # PFS, Output = getpf(0, ts, 100, mr_list)
    # MG, SMG = getMG(0, ts, 100, mr_list, PFS)
    # for i in range(len(SMG)):
    #     for j in range(len(SMG[i])):
    #         if 1 in SMG[i][j]:
    #             print(i)
    #             break
    string = 'MATH'
    row = 1
    path = '/Applications/work/data/MT/MFT/Result/result'+sys.argv[1]+'.xlsx'  # '+sys.argv[1][:-1]+'  '+sys.argv[1]+'
    wb = load_workbook(path)
    del wb[string]
    ws = wb.create_sheet(string)
    # FR = []
    MG_set = []
    PF_set = []
    SMG_set = []  # 不去除
    with open('/Applications/work/data/MT/MFT/' + string + '/FRnew.json', 'r') as load_f:
        data = json.load(load_f)
    data = json.loads(data)
    FR = data['FR']
    EMR = [0, 2, 5, 6, 9, 10, 11]
    for mu in range(1, 14):  # 14
        # a = FailureRate(mu)
        # FR.append(a)  # original = 1-failure rate
        # data = {
        #          'FR': FR
        # }
        # # 将数据存下来
        # json_str = json.dumps(data)
        # with open('/Applications/work/data/MT/MFT/' + string + '/FRnew.json', 'w') as f:
        #     json.dump(json_str, f)
        # PFS, Output = getpf(mu, ts, 100, mr_list)
        # MGS, SMGS = getMG(mu, ts, 100, mr_list, PFS)
        # MG_set.append(MGS)
        # PF_set.append(PFS)
        # SMG_set.append(SMGS)
        # data = {
        #          'PF': PFS, 'MG': MGS, 'SMG': SMGS, 'Output': Output
        # }
        # # # 将数据存下来
        # json_str = json.dumps(data)
        # with open('/Applications/work/data/MT/MFT/' + string + '/mutant' + str(mu) + 'new.json', 'w') as f:
        #     json.dump(json_str, f)

        if 0 < FR[mu-1] < 30:
            with open('/Applications/work/data/MT/MFT/' + string + '/mutant' + str(mu) + 'new.json', 'r') as load_f:
                data = json.load(load_f)
            data = json.loads(data)
            pf = data['PF']
            MG = data['SMG']
            Output = data['Output']
            for i in range(len(Output)):
                output = [Output[i][0], Output[i][1:][0]]
                for j in range(len(MG[0][0])):
                    if j not in EMR:
                        continue
                    output.append(Output[i][1:][j + 1])
                Output[i] = output
            MGc = []
            for i in range(len(MG)):
                mg = [MG[i][0]]
                for j in range(len(MG[i])):
                    if j in EMR:
                        mg.append(MG[i][j+1])
                MGc.append(mg)
            MG = copy.deepcopy(MGc)
            for i in range(len(MG)):
                for j in range(len(MG[i])):
                    a = []
                    for k in EMR:
                        a.append(MG[i][j][k])
                    MG[i][j] = a

            Output = data['Output']
            PF_set.append(pf)
            MG_set.append(MG)
            row = eval('getMetrics_v14')(row, ws, mu, MG, pf, Output, EMR)  # +sys.argv[1][-1]

    wb.save(path)
