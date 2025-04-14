import shutil
from PT import *
from openpyxl import load_workbook
from publicFun import *
import json
import sys

random.seed(1)


def copyFile(fileDir, tarDir, picknumber):
    pathDir = os.listdir(fileDir)  # 取图片的原始路径
    sample = random.sample(pathDir, picknumber)  # 随机选取picknumber数量的样本图片
    for i in range(len(sample)):
        shutil.copy(fileDir + sample[i], tarDir + "input{}.txt".format(i))
    return


def getOriginalInput():
    fileDir1 = "/Applications/work/code/project/printtokens/inputs/"  # 源文件夹路径
    tarDir1 = '/Applications/work/data/MT/MFT/' + string + '/input/'  # 移动到新的文件夹路径
    fileDir2 = '/Applications/work/data/MT/MFT/' + string + '/input/'  # 源文件夹路径
    tarDir2 = '/Applications/work/data/MT/MFT/' + string + '/RandomInput/'  # 移动到新的文件夹路径
    shutil.rmtree(tarDir1)
    os.mkdir(tarDir1)
    copyFile(fileDir1, tarDir1, 1000)
    shutil.rmtree(tarDir2)
    os.mkdir(tarDir2)
    copyFile(fileDir2, tarDir2, 100)


def FailureRate(mu):
    # 把SourceCases读出来
    Result = []
    for i in range(1000):
        ts.setInputOutput("input{}.txt".format(i), "output{}_{}.txt".format(0, i))
        original_output = MR().getResults(ts)
        program_ts = ts
        program_ts.setInputOutput("input{}.txt".format(i), "output{}_{}.txt".format(mu, i))
        program_output = MR().getResults(program_ts)
        isViolate = MR().assertViolation(original_output, program_output)
        if isViolate:
            Result.append(1)
        else:
            Result.append(0)
    FR = round(Result.count(1) / len(Result) * 100, 2)
    return FR, Result


def getTestcase(mr_list, test_case, num_of_samples):
    # for i in range(num_of_samples):
    #     test_case.setInputOutput("infile_{}".format(i), "outfile_{}".format(i), "outtree_{}".format(i))
    #     test_case.generateRandomTestcase()
    # for i in range(num_of_samples):
    #     for j in range(len(mr_list)):
    #         mr = mr_list[j]
    #         mr.setTestCase(test_case)
    #         mr.original_ts.setInputOutput("input{}.fa".format(i), "reference{}.fa".format(i), "e{}.fa".format(i), "output{}.txt".format(i))
    #         mr.getFollowInput(j)
    for i in range(num_of_samples):
        # test_case.setInputOutput("input{}.txt".format(i), "output{}.txt".format(i))
        # test_case.generateInput()  # 生成原始测试用例
        for j in range(len(mr_list)):
            mr = mr_list[j]
            mr.setTestCase(test_case)
            mr.original_ts.setInputOutput("input{}.txt".format(i), "output{}.txt".format(i))
            mr.getFollowInput(j)
            for k in range(len(mr_list)):
                mr = mr_list[k]
                mr.setTestCase(test_case)
                mr.original_ts.setInputOutput("input{}_{}.txt".format(i, j), "output{}_{}.txt".format(i, j))
                mr.getFollowInput(k)


def getMG(mu, ts, num_of_samples, mr_list, PFS):
    MGS = []
    SMGS = []
    for i in range(num_of_samples):
        MGs = []
        MG = [0] * len(mr_list)
        ts.setInputOutput("input{}.txt".format(i), "output{}_{}.txt".format(mu, i))
        original_output = MR().getResults(ts)
        followup_ts = ts
        for j in range(len(mr_list)):
            followup_ts.setInputOutput("input{}_{}.txt".format(i, j), "output{}_{}_{}.txt".format(mu, i, j))
            mr = mr_list[j]
            followup_output = mr.getResults(followup_ts)
            expected_output = mr.getExpectedOutput(original_output)
            isViolate = mr.assertViolation(expected_output, followup_output)
            if isViolate:
                MG[j] = 1
        MGs.append(MG)
        for m in range(len(mr_list)):
            MG = [0] * len(mr_list)
            ts.setInputOutput("input{}_{}.txt".format(i, m), "output{}_{}_{}.txt".format(mu, i, m))
            original_output = MR().getResults(ts)
            followup_ts = ts
            for n in range(len(mr_list)):
                followup_ts.setInputOutput("input{}_{}_{}.txt".format(i, m, n),
                                           "output{}_{}_{}_{}.txt".format(mu, i, m, n))
                mr = mr_list[n]
                followup_output = mr.getResults(followup_ts)
                expected_output = mr.getExpectedOutput(original_output)
                isViolate = mr.assertViolation(expected_output, followup_output)
                if isViolate:
                    MG[n] = 1
            MGs.append(MG)
        # 去掉巧合满足性
        for i1 in range(len(MGs)):
            for j1 in range(len(MGs[i1])):
                if MGs[i1][j1] == 0 and (PFS[i][i1] or PFS[i][i1 * len(MGs[0]) + j1 + 1]):  # 如果satisfied
                    MGs[i1][j1] = 3

        SMGs = copy.deepcopy(MGs)
        # 随机去除一些MG
        for i2 in range(1, len(MGs)):  # 第一组不变
            t = random.randint(1, len(MGs[i2]) - 1)  # 去几个
            a = [n for n in range(len(MGs[i2]))]
            random.shuffle(a)
            b = a[:t]
            for j2 in b:
                MGs[i2][j2] = 4

        SMGS.append(SMGs)
        MGS.append(MGs)
    return MGS, SMGS


def getpf(mu, ts, num_of_samples, mr_list):
    pf = []
    Output = []
    for i in range(num_of_samples):
        output = []
        result = [0] * (len(mr_list) * (len(mr_list) + 1) + 1)
        ts.setInputOutput("input{}.txt".format(i), "output{}_{}.txt".format(0, i))
        original_output = MR().getResults(ts)
        program_ts = ts
        program_ts.setInputOutput("input{}.txt".format(i), "output{}_{}.txt".format(mu, i))
        program_output = MR().getResults(program_ts)
        output.append(original_output.count)
        output.append(program_output.count)
        isViolate = MR().assertViolation(original_output, program_output)
        if isViolate:
            result[0] = 1

        for j in range(len(mr_list)):
            ts.setInputOutput("input{}_{}.txt".format(i, j), "output{}_{}_{}.txt".format(0, i, j))
            original_output = MR().getResults(ts)
            program_ts = ts
            program_ts.setInputOutput("input{}_{}.txt".format(i, j), "output{}_{}_{}.txt".format(mu, i, j))
            program_output = MR().getResults(program_ts)
            output.append(program_output.count)
            isViolate = MR().assertViolation(original_output, program_output)
            if isViolate:
                result[j + 1] = 1

        Output.append(output)

        for m in range(len(mr_list)):
            for n in range(len(mr_list)):
                ts.setInputOutput("input{}_{}_{}.txt".format(i, m, n),
                                  "output{}_{}_{}_{}.txt".format(0, i, m, n))
                original_output = MR().getResults(ts)
                program_ts = ts
                program_ts.setInputOutput("input{}_{}_{}.txt".format(i, m, n),
                                          "output{}_{}_{}_{}.txt".format(mu, i, m, n))
                program_output = MR().getResults(program_ts)
                isViolate = MR().assertViolation(original_output, program_output)
                if isViolate:
                    result[len(mr_list) + 1 + m * len(mr_list) + n] = 1
        pf.append(result)
    return pf, Output


if __name__ == "__main__":
    string = 'PT'
    # getOriginalInput()
    # myenv = MyEnv()
    # myenv.CreateWorkingDirs()
    # ts = TestCase()
    # num_of_samples = 100  # 测试用例个数
    # mr_list = [MR1(), MR2(), MR3(), MR4(), MR5(), MR6(), MR7(), MR8(), MR9(), MR10(), MR11()]
    # getTestcase(mr_list, ts, num_of_samples)
    # PFS, Output = getpf(0, ts, 100, mr_list)
    # MG, SMG = getMG(0, ts, 100, mr_list, PFS)
    # for i in range(len(SMG)):
    #     for j in range(len(SMG[i])):
    #         if 1 in SMG[i][j]:
    #             print(i)
    #             break
    row = 1
    path = '/Applications/work/data/MT/MFT/Result/result66.xlsx'  # '+sys.argv[1][:-1]+'  '+sys.argv[1]+'
    # '+sys.argv[1][:-1]+'  '+sys.argv[1]+'
    wb = load_workbook(path)
    del wb[string]
    ws = wb.create_sheet(string)
    MG_set = []
    PF_set = []
    SMG_set = []  # 不去除
    # FR = []
    with open('/Applications/work/data/MT/MFT/' + string + '/FR.json', 'r') as load_f:
        data = json.load(load_f)
    data = json.loads(data)
    FR = data['FR']
    EMR = [0, 1, 3, 4]
    Group = []
    Result = []
    Ind = []
    for mu in range(6, 8):  # 1, 8
        # a, r = FailureRate(mu)
        # ind = [index for (index, value) in enumerate(r) if value == 1]
        # FR.append(a)  # original = 1-failure rate
        # Result.append(r)
        # Ind.append(ind)
        # data = {
        #          'FR': FR
        # }
        # # 将数据存下来
        # json_str = json.dumps(data)
        # with open('/Applications/work/data/MT/MFT/' + string + '/FR.json', 'w') as f:
        #     json.dump(json_str, f)
        # PFS, Output = getpf(mu, ts, 100, mr_list)
        # MGS, SMGS = getMG(mu, ts, 100, mr_list, PFS)
        # MG_set.append(MGS)
        # PF_set.append(PFS)
        # SMG_set.append(SMGS)
        # data = {
        #          'PF': PFS, 'MG': MGS, 'SMG': SMGS, 'Output': Output
        # }
        # # 将数据存下来
        # json_str = json.dumps(data)
        # with open('/Applications/work/data/MT/MFT/' + string + '/mutant' + str(mu) + '.json', 'w') as f:
        #     json.dump(json_str, f)

        if 0 < FR[mu - 1] < 30:
            with open('/Applications/work/data/MT/MFT/' + string + '/mutant' + str(mu) + '.json', 'r') as load_f:
                data = json.load(load_f)
            data = json.loads(data)
            pf = data['PF']
            MG = data['SMG']
            Output = data['Output']
            MGc = []
            for i in range(len(Output)):
                output = [Output[i][0], Output[i][1:][0]]
                for j in range(len(MG[0][0])):
                    if j not in EMR:
                        continue
                    output.append(Output[i][1:][j + 1])
                Output[i] = output
            for i in range(len(MG)):
                mg = [MG[i][0]]
                for j in range(len(MG[i])):
                    if j in EMR:
                        mg.append(MG[i][j + 1])
                MGc.append(mg)
            MG = copy.deepcopy(MGc)
            for i in range(len(MG)):
                for j in range(len(MG[i])):
                    a = []
                    for k in EMR:
                        a.append(MG[i][j][k])
                    MG[i][j] = a
            PF_set.append(pf)
            MG_set.append(MG)
            # output = [Output[1:][0]]
            # for j in range(len(MG[0])):
            #     if j not in EMR:
            #         continue
            #     output.append(Output[1:][j + 1])
            # count = []
            # for i in range(len(Output)):
            #     c = group2(Output[i], MG[i], EMR)
            #     count.append(c)
            # Group.append(count)
            row = eval('getMetrics_v14')(row, ws, mu, MG, pf, Output, EMR)  # +sys.argv[1][-1]

    wb.save(path)
