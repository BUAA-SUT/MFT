'''
RQ1
Fp = 0  # Case 1: t1(f) > t2(p)
FP = 0  # Case 2: t1(f) = t2(p)
Pf = 0  # Case 3: t1(f) < t2(p)
FF = 0  # Case 4: t1 and t2 both fail
'''
from Tcas import *
import math
import numpy as np
import copy
import random
random.seed(1)
import json
from publicFun import *
from openpyxl import load_workbook


def getOriginalInput():
    source_case_set = []
    path = r"/Applications/work/code/project/tcas/testplans.alt/universe.txt"
    with open(path, 'r') as f:
        list_read = f.readlines()
    for i in range(len(list_read)):
        test_case = []
        list2 = list_read[i].split()
        for j in range(len(list2)):
            test_case.append(int(list2[j]))  # 字符串转整型
        # 获取符合MR要求的源测试用例
        if test_case[6] <= 3:
            source_case_set.append(test_case)
        if len(source_case_set) >= 1000:
            break
    # 随机取100个测试用例
    random_input = random.sample(source_case_set, 100)
    data = {
            'source_case_set': source_case_set,
            'random_input': random_input
    }
    json_str = json.dumps(data)
    with open('/Applications/work/data/MT/MFT/' + string + '/OriginalInput.json', 'w') as f:
        json.dump(json_str, f)
    return source_case_set, random_input


def FailureRate(dynamic):
    # 把SourceCases读出来
    Result = []
    with open('/Applications/work/data/MT/MFT/'+string+'/OriginalInput.json', 'r') as load_f:
        data = json.load(load_f)
    data = json.loads(data)
    source_case_set = data['source_case_set']
    for i in range(len(source_case_set)):
        result_s_a = TCAS().Tcas(source_case_set[i])  # oracle
        result_s_m = dynamic.Tcas(source_case_set[i])
        if result_s_a == result_s_m:
            Result.append(0)
        else:
            Result.append(1)
    FR = round(Result.count(1) / len(Result) * 100, 2)
    return FR


def riskIndex(argv, dynamic):
    MGS = []
    Result = []
    testcase = []
    output = []
    source_case = argv.copy()
    output.append(TCAS().Tcas(source_case))
    testcase.append(source_case)
    MG, follow_case = MTG(source_case, dynamic)  # t1t2t3t4
    for i in range(len(follow_case)):
        testcase.append(follow_case[i])
    for i in range(len(testcase)):
        output.append(dynamic.Tcas(testcase[i]))
    MGS.append(MG)
    for i in range(len(follow_case)):
        MG, ffollow_case = MTG(follow_case[i], dynamic)  # t2t5...
        MGS.append(MG)
        for j in range(len(ffollow_case)):
            testcase.append(ffollow_case[j])
    # MG统计完, testcase统计完
    for i in range(len(testcase)):
        result_s_a = TCAS().Tcas(testcase[i])  # oracle
        result_s_m = dynamic.Tcas(testcase[i])
        if result_s_a == result_s_m:
            Result.append(0)
        else:
            Result.append(1)

    # 去掉巧合满足性
    for i in range(len(MGS)):
        for j in range(len(MGS[i])):
            if MGS[i][j] == 0 and (Result[i] or Result[i * len(MGS[0]) + j + 1]):  # 如果satisfied
                MGS[i][j] = 3

    SMGS = copy.deepcopy(MGS)

    # 随机去除一些MG
    for i in range(1, len(MGS)):  # 第一组不变
        t = random.randint(1, len(MGS[i])-1)  # 去几个
        a = [n for n in range(len(MGS[i]))]
        random.shuffle(a)
        b = a[:t]
        for j in b:
            MGS[i][j] = 4

    return MGS, Result, SMGS, output


if __name__ == '__main__':
    string = 'Tcas'
    # source_case_set, random_input = getOriginalInput()
    row = 1
    path = '/Applications/work/data/MT/MFT/Result/result'+sys.argv[1]+'.xlsx'  # '+sys.argv[1][:-1]+'  '+sys.argv[1]+'
    wb = load_workbook(path)
    del wb[string]
    ws = wb.create_sheet(string)
    # with open('/Applications/work/data/MT/MFT/' + string + '/OriginalInput.json', 'r') as load_f:
    #     data = json.load(load_f)
    # data = json.loads(data)
    # random_input = data['random_input']
    FR = []
    MG_set = []
    PF_set = []
    SMG_set = []  # 不去除
    Group = []
    EMR = [0, 1, 2, 3, 4, 5, 6, 7, 8]
    for mu in range(1, 12):
        dynamic = TCASFactory("Mutant" + str(mu)).getTCAS()
        a = FailureRate(dynamic)
        FR.append(a)  # original = 1-failure rate
        # MGS = []
        # PFS = []
        # SMGS = []
        # Output = []
        # for i in range(len(random_input)):
        #     MG, PF, SMG, output = riskIndex(random_input[i], dynamic)
        #     MGS.append(MG)
        #     PFS.append(PF)
        #     SMGS.append(SMG)
        #     Output.append(output)
        # MG_set.append(MGS)
        # PF_set.append(PFS)
        # SMG_set.append(SMGS)
        # data = {
        #          'PF': PFS, 'MG': MGS, 'SMG': SMGS, 'Output': Output
        # }
        # # 将数据存下来
        # json_str = json.dumps(data)
        # with open('/Applications/work/data/MT/MFT/' + string + '/mutant' + str(mu) + '.json', 'w') as f:
        #     json.dump(json_str, f)

        if 0 < a < 30:
            with open('/Applications/work/data/MT/MFT/' + string + '/mutant' + str(mu) + '.json', 'r') as load_f:
                data = json.load(load_f)
            data = json.loads(data)
            pf = data['PF']
            MG = data['SMG']
            Output = data['Output']
            PF_set.append(pf)
            MG_set.append(MG)
            count = []
            for i in range(len(Output)):
                c = group(Output[i])
                count.append(c)
            Group.append(count)
            row = eval('getMetrics_v14')(row, ws, mu, MG, pf, Output, EMR)  # +sys.argv[1][-1]

    wb.save(path)
