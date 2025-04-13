import copy
import random
import numpy as np
import operator
from openpyxl import load_workbook
from collections import Counter

random.seed(1)


def riskformula(index):  # 多个公式
    ev = index[0]
    es = index[1]
    nv = index[2]
    ns = index[3]
    F = ev + nv
    P = es + ns
    if ev < F:
        N1 = -1
        Bin = 0
    elif ev == F and F != 0:
        N1 = ns
        Bin = 1
    else:
        N1 = -1
        Bin = 0
    Jaccard = ev / (ev + es + nv)
    Aberg = ev / (ev + 2 * (nv + es))
    SDice = 2 * ev / (2 * ev + nv + es)
    Dice = 2 * ev / (ev + nv + es)
    Goodman = (2 * ev - nv - es) / (2 * ev + nv + es)
    Tar = (ev / (ev + nv)) / (ev / (ev + nv) + es / (es + ns))
    Qe = ev / (ev + es)
    CBI = (ev / (ev + es)) - ((ev + nv) / (ev + nv + es + ns))
    W2 = ev - es
    Hamann = (ev + ns - nv - es) / (ev + nv + es + ns)
    SM = (ev + ns) / (ev + nv + es + ns)
    Sokal = 2 * (ev + ns) / (2 * (ev + ns) + nv + es)
    RT = (ev + ns) / (ev + ns + 2 * (nv + es))
    Hamming = ev + ns
    Euclid = np.sqrt(ev + ns)
    Scott = (4 * ev * ns - 4 * nv * es - (nv - es) ** 2) / ((2 * ev + nv + es) * (2 * ns + nv + es))
    Rogot1 = 0.5 * (ev / (2 * ev + nv + es) + ns / (2 * ns + nv + es))
    Kul2 = 0.5 * (ev / (ev + nv) + ev / (ev + es))
    Ochiai = ev / ((ev + nv) * (ev + es)) ** 0.5
    M2 = ev / (ev + ns + 2 * (nv + es))
    AMPLE2 = ev / (ev + nv) - es / (es + ns)
    if es <= 2:
        W3 = ev - es
    elif es > 2 and es <= 10:
        W3 = ev - 2 - 0.1 * (es - 2)
    else:
        W3 = ev - 2.8 - 0.001 * (es - 10)
    AM = (2 * ev * ns - 2 * nv * es) / ((ev + es) * (ns + nv) + (ev + nv) * (es + ns))
    Cohen = (2 * ev * ns - 2 * nv * es) / ((ev + es) * (ns + es) + (ev + nv) * (nv + ns))
    Fle = (4 * ev * ns - 4 * nv * es - (nv - es) ** 2) / (2 * ev + nv + es + 2 * ns + nv + es)

    N2 = ev - es / (es + ns + 1)
    W1 = ev
    RR = ev / (ev + es + nv + ns)
    formula = [N1, N2, W1, RR, Bin, Jaccard, Aberg, SDice, Dice, Goodman, Tar, Qe, CBI, W2, Hamann,
               SM, Sokal, RT, Hamming, Euclid, Scott, Rogot1, Kul2, Ochiai, M2, AMPLE2, W3, AM, Cohen, Fle]
    return formula


def riskformula_new(index, t):  # 多个公式
    ev = index[0]
    es = index[1]
    nv = index[2]
    ns = index[3]
    ev_only = nv
    F = ev + nv
    P = es + ns
    if t == 0:
        if ev < F:
            formula = -1
        elif ev == F and F != 0:
            formula = ns
        else:
            formula = -1
    elif t == 1:
        formula = ev - es / (es + ns + 1)
    elif t == 2:
        formula = ev
    elif t == 3:
        formula = ev / (ev + es + nv + ns)
    elif t == 4:
        if ev < F:
            formula = 0
        elif ev == F and F != 0:
            formula = 1
        else:
            formula = 0
    elif t == 5:
        formula = ev / (ev + es + nv)
    elif t == 6:
        formula = ev / (ev + 2 * (nv + es))
    elif t == 7:
        formula = 2 * ev / (2 * ev + nv + es)
    elif t == 8:
        formula = 2 * ev / (ev + nv + es)
    elif t == 9:
        formula = (2 * ev - nv - es) / (2 * ev + nv + es)
    elif t == 10:
        formula = (ev / (ev + nv)) / (ev / (ev + nv) + es / (es + ns))
    elif t == 11:
        formula = ev / (ev + es)
    elif t == 12:
        formula = (ev / (ev + es)) - ((ev + nv) / (ev + nv + es + ns))
    elif t == 13:
        formula = ev - es
    elif t == 14:
        formula = (ev + ns - nv - es) / (ev + nv + es + ns)
    elif t == 15:
        formula = (ev + ns) / (ev + nv + es + ns)
    elif t == 16:
        formula = 2 * (ev + ns) / (2 * (ev + ns) + nv + es)
    elif t == 17:
        formula = (ev + ns) / (ev + ns + 2 * (nv + es))
    elif t == 18:
        formula = ev + ns
    elif t == 19:
        formula = np.sqrt(ev + ns)
    elif t == 20:
        formula = (4 * ev * ns - 4 * nv * es - (nv - es) ** 2) / ((2 * ev + nv + es) * (2 * ns + nv + es))
    elif t == 21:
        formula = 0.5 * (ev / (2 * ev + nv + es) + ns / (2 * ns + nv + es))
    elif t == 22:
        formula = 0.5 * (ev / (ev + nv) + ev / (ev + es))
    elif t == 23:
        formula = ev / ((ev + nv) * (ev + es)) ** 0.5
    elif t == 24:
        formula = ev / (ev + ns + 2 * (nv + es))
    elif t == 25:
        formula = ev / (ev + nv) - es / (es + ns)
    elif t == 26:
        if es <= 2:
            formula = ev - es
        elif 2 < es <= 10:
            formula = ev - 2 - 0.1 * (es - 2)
        else:
            formula = ev - 2.8 - 0.001 * (es - 10)
    elif t == 27:
        formula = (2 * ev * ns - 2 * nv * es) / ((ev + es) * (ns + nv) + (ev + nv) * (es + ns))
    elif t == 28:
        formula = (2 * ev * ns - 2 * nv * es) / ((ev + es) * (ns + es) + (ev + nv) * (nv + ns))
    elif t == 29:
        formula = (4 * ev * ns - 4 * nv * es - (nv - es) ** 2) / (2 * ev + nv + es + 2 * ns + nv + es)
    else:
        formula = - (es ** 3 + nv ** 2) / (ev + ns)

    return formula


# 返回一个列表中出现次数最多的元素
def showmax(lt):
    # index1 = 0  # 记录出现次数最多的元素下标
    #
    # max = 0  # 记录最大的元素出现次数
    #
    # for i in range(len(lt)):
    #
    #     flag = 0  # 记录每一个元素出现的次数
    #
    #     for j in range(i + 1, len(lt)):  # 遍历i之后的元素下标
    #
    #         if lt[j] == lt[i]:
    #             flag += 1  # 每当发现与自己相同的元素，flag+1
    #
    #     if flag > max:  # 如果此时元素出现的次数大于最大值，记录此时元素的下标
    #         max = flag
    #         index1 = i
    maxlabel = max(lt, key=lt.count)
    can = [maxlabel]
    if lt.count(maxlabel) == 1:
        # can.remove(maxlabel)
        # can.append(random.choice(lt))
        return lt
    lt2 = lt.copy()
    while 1:
        for i in range(len(lt2) - 1, -1, -1):
            if lt2[i] == maxlabel:
                lt2.remove(maxlabel)
        if len(lt2) == 0:
            break
        maxlabel2 = max(lt2, key=lt2.count)
        if lt.count(maxlabel) == lt.count(maxlabel2):
            can.append(maxlabel2)
            maxlabel = maxlabel2
        else:
            break
    return can  # 返回出现最多的元素


def showmaxsize(lt):
    can = []
    maxlen = max(map(len, lt))
    for i in lt:
        if len(i) == maxlen:
            can.append(i)
    return can  # 返回出现最多的元素


def showmin(lt):
    maxlabel = min(lt, key=lt.count)
    can = [maxlabel]
    # if lt.count(maxlabel) == 1:
    #     # can.remove(maxlabel)
    #     # can.append(random.choice(lt))
    #     return lt
    lt2 = lt.copy()
    while 1:
        for i in range(len(lt2) - 1, -1, -1):
            if lt2[i] == maxlabel:
                lt2.remove(maxlabel)
        if len(lt2) == 0:
            break
        maxlabel2 = min(lt2, key=lt2.count)
        if lt.count(maxlabel) == lt.count(maxlabel2):
            can.append(maxlabel2)
            maxlabel = maxlabel2
        else:
            break
    return can  # 返回出现最多的元素


def getSus(index, t):
    formula = 0
    try:
        formula = riskformula_new(index, t)
    except:
        if index[0] + index[1] == 0:  # e = 0
            formula = -1000
        elif index[0] + index[2] == 0:  # v = 0
            formula = -1000
        elif index[1] + index[3] == 0:  # s = 0
            formula = 1000
    return formula


# def getMetrics(row, ws, mu, MG, Risk, pf):
#     '''
#     统计所有MG
#     '''
#     Formula = ['Naish1', 'Naish2', 'Wong1', 'Russel&Rao', 'Binary', 'Jaccard', 'Anderberg', 'Sørensen-Dice', 'Dice',
#                'Goodman',
#                'Tarantula', 'qe', 'CBI Inc.', 'Wong2', 'Hamann', 'Simple Matching', 'Sokal', 'Rogers&Tanimoto',
#                'Hamming etc.',
#                'Euclid', 'Scott', 'Rogot1', 'Kulczynski2', 'Ochiai', 'M2', 'AMPLE2', 'Wong3', 'Arithmetic Mean',
#                'Cohen', 'Fleiss']
#
#     # 统计指标
#     datadist = {}
#     tablelist = {"Mutant" + str(mu): ['case1(%)', 'case2(%)', 'case3(%)', 'case4(%)', 'VMG(%)',
#                                       'failed test case(%)', 'false satisfied MG(%)']}
#     datadist.update(tablelist)
#     Index = []
#     for t in range(len(Formula)):
#         V_MGlist = []
#         Fplist = []  # Case 1: t1(f) > t2(p)
#         FPlist = []  # Case 2: t1(f) = t2(p)
#         Pflist = []  # Case 3: t1(f) < t2(p)
#         FFlist = []  # Case 4: t1 and t2 both fail
#         FFSlist = []
#         Failed = []
#         V_MGsum = []
#         index = []
#         for k in range(len(MG)):
#             V_MG = 0
#             S_MG = 0
#             Fp = 0  # Case 1: t1(f) > t2(p)
#             FP = 0  # Case 2: t1(f) = t2(p)
#             Pf = 0  # Case 3: t1(f) < t2(p)
#             FF = 0  # Case 4: t1 and t2 both fail
#             FFS = 0
#             Total_MG = len(MG[0]) * len(MG[0][0])
#             failedcase = pf[k].count(1)
#             percent_failed = failedcase / len(pf[k])
#             for i in range(len(MG[k])):
#                 for j in range(len(MG[k][i])):
#                     if MG[k][i][j] == 1:  # violated
#                         V_MG += 1
#                     elif MG[k][i][j] == 0 or MG[k][i][j] == 3:
#                         S_MG += 1
#                         if MG[k][i][j] == 0 and (pf[k][i] or pf[k][i * len(MG[0][0]) + j + 1]):
#                             FFS += 1
#
#             percent_VMG = V_MG / Total_MG
#             percent_FFS = FFS / S_MG
#
#             if V_MG == 0:
#                 continue
#
#             V_MGlist.append(percent_VMG)
#             FFSlist.append(percent_FFS)
#             Failed.append(percent_failed)
#
#             for i in range(len(MG[k])):  # len(MG[k])
#                 for j in range(len(MG[k][i])):
#                     if MG[k][i][j] == 1:  # violated
#                         if pf[k][i] and pf[k][i * len(MG[0][0]) + j + 1]:
#                             # 如果都是failed, 只要不进入此分支, 那么肯定一个failed, 一个passed (两个passed不会是violated)
#                             FF += 1
#                         elif not (Risk[k][i][t] == Risk[k][i * len(MG[0][0]) + j + 1][t]):  # 只有一个failed, 并且不相等
#                             if pf[k][i] and Risk[k][i][t] > Risk[k][i * len(MG[0][0]) + j + 1][t]:
#                                 Fp += 1
#                             elif pf[k][i * len(MG[0][0]) + j + 1] and Risk[k][i][t] < \
#                                     Risk[k][i * len(MG[0][0]) + j + 1][t]:
#                                 Fp += 1
#                             else:
#                                 Pf += 1  # p比f高
#                                 index.append((k, i, j))
#                                 # print(i)
#                                 # print(k, i)
#                                 # print(k, i, j, pf[k][i], pf[k][i * len(MG[0][0]) + j + 1])
#                         else:  # 只有一个failed, 但是相等
#                             FP += 1
#
#             percent_Fp = Fp / V_MG  # 1
#             percent_FP = FP / V_MG  # 2
#             percent_Pf = Pf / V_MG  # 3
#             percent_FF = FF / V_MG  # 4
#             Fplist.append(percent_Fp)
#             FPlist.append(percent_FP)
#             Pflist.append(percent_Pf)
#             FFlist.append(percent_FF)
#
#             # V_MGsum.append(V_MG)
#
#         value = [round(np.mean(Fplist) * 100, 2), round(np.mean(FPlist) * 100, 2), round(np.mean(Pflist) * 100, 2),
#                  round(np.mean(FFlist) * 100, 2), round(np.mean(V_MGlist) * 100, 2),
#                  round(np.mean(Failed) * 100, 2), round(np.mean(FFSlist) * 100, 2)]
#         data = {
#             Formula[t]: value
#         }
#         datadist.update(data)
#         Index.append(index)
#     for i, j in datadist.items():  # i--公式名称, j--指标值
#         ws.cell(row, 1).value = i  # 添加第 1 列的数据
#         for col in range(2, len(j) + 2):  # values列表中索引
#             ws.cell(row, col).value = j[col - 2]
#         row += 1  # 行数
#     row += 2  # 行数
#     return row
#
#
# def getMetrics_v2(row, ws, mu, MG, pf):
#     '''
#     根据TY改, 统计第一行, local
#     '''
#     Formula = ['Naish1', 'Naish2', 'Wong1', 'Russel&Rao', 'Binary', 'Jaccard', 'Anderberg', 'Sørensen-Dice', 'Dice',
#                'Goodman',
#                'Tarantula', 'qe', 'CBI Inc.', 'Wong2', 'Hamann', 'Simple Matching', 'Sokal', 'Rogers&Tanimoto',
#                'Hamming etc.',
#                'Euclid', 'Scott', 'Rogot1', 'Kulczynski2', 'Ochiai', 'M2', 'AMPLE2', 'Wong3', 'Arithmetic Mean',
#                'Cohen', 'Fleiss']
#
#     # 统计指标
#     datadist = {}
#     tablelist = {"Mutant" + str(mu): ['identify(%)', 'failed case(%)', 'VMG(%)', 'SMG(%)', 'false satisfied MG(%)']}
#     datadist.update(tablelist)
#     tt = 0
#     failedcase = 0
#     for t in range(len(Formula)):
#         FFSlist = []
#         failedcase = 0
#         VMGlist = []
#         SMGlist = []
#         true = 0
#         xx = 0
#         for k in range(len(MG)):
#             # 如果failure rate>30%或者全部正确，则不考虑
#             tt = len(MG[0][0]) + 1
#             num = pf[k][0:tt].count(1)
#             if num / tt > 0.4 or len(set(pf[k][0:tt])) == 1:
#                 xx += 1
#                 if xx == len(MG):
#                     return row
#                 continue
#             # if len(set(pf[k][0:tt])) == 1:
#             #     xx += 1
#             #     if xx == len(MG):
#             #        return row
#             #     continue
#             S_MG = 0
#             V_MG = 0
#             FFS = 0
#             for x in range(len(pf[k][0:tt])):
#                 if pf[k][x] == 1:
#                     failedcase += 1
#             for i in range(len(MG[k])):
#                 for j in range(len(MG[k][i])):
#                     if MG[k][i][j] == 0 or MG[k][i][j] == 3:
#                         S_MG += 1
#                         if MG[k][i][j] == 3:
#                             FFS += 1
#                     elif MG[k][i][j] == 1:
#                         V_MG += 1
#             if S_MG == 0:
#                 percent_FFS = 0
#             else:
#                 percent_FFS = FFS / S_MG
#
#             FFSlist.append(percent_FFS)
#             VMGlist.append(V_MG / (S_MG + V_MG))
#             SMGlist.append(S_MG / (S_MG + V_MG))
#
#             risk = []
#             dis = []
#             id = []
#             for i in range(1):  # len(MG[k])
#                 for j in range(len(MG[k][i])):
#
#                     # 根据包含的测试用例确定 MG set
#                     sum_s = MG[k][i].count(0) + MG[k][i].count(3)
#                     sum_v = MG[k][i].count(1)
#                     sum_s += (MG[k][j + 1].count(0)) + MG[k][j + 1].count(3)
#                     sum_v += MG[k][j + 1].count(1)
#                     # 求每个测试用例的ev es nv ns
#                     ev_a = MG[k][i].count(1)
#                     es_a = MG[k][i].count(0) + MG[k][i].count(3)
#                     nv_a = sum_v - ev_a
#                     ns_a = sum_s - es_a
#                     index_a = [ev_a, es_a, nv_a, ns_a]
#
#                     ev_b = MG[k][j + 1].count(1) + 1
#                     es_b = MG[k][j + 1].count(0) + MG[k][j + 1].count(3)
#                     if MG[k][i][j] == 1:
#                         ev_b += 1
#                     elif MG[k][i][j] == 0 or MG[k][i][j] == 3:
#                         es_b += 1
#                     nv_b = sum_v - ev_b
#                     ns_b = sum_s - es_b
#                     index_b = [ev_b, es_b, nv_b, ns_b]
#                     # 求测试用例的可疑度
#                     # for t in range(len(Formula)):
#                     sus_a = getSus(index_a, t)
#                     sus_b = getSus(index_b, t)
#                     if sus_b < sus_a:
#                         id.append(j + 1)
#                         risk.append(sus_b)
#                         dis.append(sus_a - sus_b)
#                         # if sus_a == 0:
#                         #     dis.append(sus_a - sus_b)
#                         # else:
#                         #     dis.append((sus_a-sus_b)/sus_a)  # 正则化
#                 if len(id) == 0:  # original input最小
#                     Id = 0
#                 else:
#                     z = zip(dis, id)
#                     z = sorted(z, reverse=True)
#                     Id = z[0][1]
#                 if pf[k][Id] == 0:
#                     true += 1
#
#         percent_identify = true / (len(MG) - xx)
#         percent_failed = failedcase / ((len(MG) - xx) * tt)
#         value = [round(percent_identify * 100, 2), round(percent_failed * 100, 2), round(np.mean(VMGlist) * 100, 2),
#                  round(np.mean(SMGlist) * 100, 2), round(np.mean(FFSlist) * 100, 2)]
#         data = {
#             Formula[t]: value
#         }
#         datadist.update(data)
#     if failedcase == 0:
#         return row
#     else:
#         for i, j in datadist.items():  # i--公式名称, j--指标值
#             ws.cell(row, 1).value = i  # 添加第 1 列的数据
#             for col in range(2, len(j) + 2):  # values列表中索引
#                 ws.cell(row, col).value = j[col - 2]
#             row += 1  # 行数
#         row += 2  # 行数
#         return row
#
#
# def getMetrics_v3(row, ws, mu, MG, pf):
#     '''
#     根据TY改, 统计第一行, global
#     '''
#     Formula = ['Naish1', 'Naish2', 'Wong1', 'Russel&Rao', 'Binary', 'Jaccard', 'Anderberg', 'Sørensen-Dice', 'Dice',
#                'Goodman',
#                'Tarantula', 'qe', 'CBI Inc.', 'Wong2', 'Hamann', 'Simple Matching', 'Sokal', 'Rogers&Tanimoto',
#                'Hamming etc.',
#                'Euclid', 'Scott', 'Rogot1', 'Kulczynski2', 'Ochiai', 'M2', 'AMPLE2', 'Wong3', 'Arithmetic Mean',
#                'Cohen', 'Fleiss']
#
#     # 统计指标
#     datadist = {}
#     tablelist = {"Mutant" + str(mu): ['identify(%)', 'failed case(%)', 'VMG(%)', 'SMG(%)', 'false satisfied MG(%)']}
#     datadist.update(tablelist)
#     tt = 0
#     failedcase = 0
#     for t in range(len(Formula)):
#         FFSlist = []
#         failedcase = 0
#         VMGlist = []
#         SMGlist = []
#         true = 0
#         xx = 0
#         for k in range(len(MG)):
#             # 如果failure rate>30%或者全部正确，则不考虑
#             tt = len(MG[0][0]) + 1
#             num = pf[k][0:tt].count(1)
#             if num / tt > 0.3 or len(set(pf[k][0:tt])) == 1:
#                 xx += 1
#                 if xx == len(MG):
#                     return row
#                 continue
#             # if len(set(pf[k][0:tt])) == 1:
#             #     xx += 1
#             #     if xx == len(MG):
#             #        return row
#             #     continue
#             S_MG = 0
#             V_MG = 0
#             FFS = 0
#             for x in range(len(pf[k][0:tt])):
#                 if pf[k][x] == 1:
#                     failedcase += 1
#             for i in range(len(MG[k])):
#                 for j in range(len(MG[k][i])):
#                     if MG[k][i][j] == 0 or MG[k][i][j] == 3:
#                         S_MG += 1
#                         if MG[k][i][j] == 3:
#                             FFS += 1
#                     elif MG[k][i][j] == 1:
#                         V_MG += 1
#             if S_MG == 0:
#                 percent_FFS = 0
#             else:
#                 percent_FFS = FFS / S_MG
#
#             FFSlist.append(percent_FFS)
#             VMGlist.append(V_MG / (S_MG + V_MG))
#             SMGlist.append(S_MG / (S_MG + V_MG))
#
#             risk = []
#             dis = []
#             id = []
#             sum_s = 0
#             sum_v = 0
#             for i in range(len(MG[k])):
#                 sum_s += MG[k][i].count(0) + MG[k][i].count(3)
#                 sum_v += MG[k][i].count(1)
#             for i in range(1):  # len(MG[k])
#                 for j in range(len(MG[k][i])):
#                     # 求每个测试用例的ev es nv ns
#                     ev_a = MG[k][i].count(1)
#                     es_a = MG[k][i].count(0) + MG[k][i].count(3)
#                     nv_a = sum_v - ev_a
#                     ns_a = sum_s - es_a
#                     index_a = [ev_a, es_a, nv_a, ns_a]
#
#                     ev_b = MG[k][j + 1].count(1) + 1
#                     es_b = MG[k][j + 1].count(0) + MG[k][j + 1].count(3)
#                     if MG[k][i][j] == 1:
#                         ev_b += 1
#                     elif MG[k][i][j] == 0 or MG[k][i][j] == 3:
#                         es_b += 1
#                     nv_b = sum_v - ev_b
#                     ns_b = sum_s - es_b
#                     index_b = [ev_b, es_b, nv_b, ns_b]
#                     # 求测试用例的可疑度
#                     # for t in range(len(Formula)):
#                     sus_a = getSus(index_a, t)
#                     sus_b = getSus(index_b, t)
#                     if sus_b < sus_a:
#                         id.append(j + 1)
#                         risk.append(sus_b)
#                         # dis.append(sus_a - sus_b)
#                         if sus_a == 0:
#                             dis.append(sus_a - sus_b)
#                         else:
#                             dis.append((sus_a - sus_b) / sus_a)  # 正则化
#                 if len(id) == 0:  # original input最小
#                     Id = 0
#                 else:
#                     z = zip(dis, id)
#                     z = sorted(z, reverse=True)
#                     Id = z[0][1]
#                 if pf[k][Id] == 0:
#                     true += 1
#
#         percent_identify = true / (len(MG) - xx)
#         percent_failed = failedcase / ((len(MG) - xx) * tt)
#         value = [round(percent_identify * 100, 2), round(percent_failed * 100, 2), round(np.mean(VMGlist) * 100, 2),
#                  round(np.mean(SMGlist) * 100, 2), round(np.mean(FFSlist) * 100, 2)]
#         data = {
#             Formula[t]: value
#         }
#         datadist.update(data)
#     if failedcase == 0:
#         return row
#     else:
#         for i, j in datadist.items():  # i--公式名称, j--指标值
#             ws.cell(row, 1).value = i  # 添加第 1 列的数据
#             for col in range(2, len(j) + 2):  # values列表中索引
#                 ws.cell(row, col).value = j[col - 2]
#             row += 1  # 行数
#         row += 2  # 行数
#         return row


# def FaultTolerance(MG, pf, Output):
#     Formula = ['Naish1', 'Naish2', 'Wong1', 'Russel&Rao', 'Binary', 'Jaccard', 'Anderberg', 'Sørensen-Dice', 'Dice',
#                'Goodman', 'Tarantula', 'qe', 'CBI Inc.', 'Wong2', 'Hamann', 'Simple Matching', 'Sokal',
#                'Rogers&Tanimoto', 'Hamming etc.', 'Euclid', 'Scott', 'Rogot1', 'Kulczynski2', 'Ochiai',
#                'M2', 'AMPLE2', 'Wong3', 'Arithmetic Mean', 'Cohen', 'Fleiss']
#     flag = []
#     true = 0
#
#     # a = random.choice(MG[0])
#     # if a == 0 or a == 3:
#     #     if pf[0] == 0:
#     #         true = 1
#     #     for _ in range(len(Formula)):
#     #         flag.append(true)
#     #     return flag
#     # a = random.sample(MG[0], 3)
#     if 1 not in MG[0]:
#         if pf[0] == 0:
#             true = 1
#         for _ in range(len(Formula)):
#             flag.append(true)
#         return flag
#     else:
#         for t in range(len(Formula)):
#             true = 0
#             id = []
#             for j in range(len(MG[0])):
#                 # 根据包含的测试用例确定 MG set
#                 sum_s = MG[0].count(0) + MG[0].count(3)
#                 sum_v = MG[0].count(1)
#                 sum_s += (MG[j + 1].count(0)) + MG[j + 1].count(3)
#                 sum_v += MG[j + 1].count(1)
#                 # 求每个测试用例的ev es nv ns
#                 ev_a = MG[0].count(1)
#                 es_a = MG[0].count(0) + MG[0].count(3)
#                 nv_a = sum_v - ev_a
#                 ns_a = sum_s - es_a
#                 index_a = [ev_a, es_a, nv_a, ns_a]
#
#                 ev_b = MG[j + 1].count(1) + 1
#                 es_b = MG[j + 1].count(0) + MG[j + 1].count(3)
#                 if MG[0][j] == 1:
#                     ev_b += 1
#                 elif MG[0][j] == 0 or MG[0][j] == 3:
#                     es_b += 1
#                 nv_b = sum_v - ev_b
#                 ns_b = sum_s - es_b
#                 index_b = [ev_b, es_b, nv_b, ns_b]
#                 # 求测试用例的可疑度
#                 # for t in range(len(Formula)):
#                 sus_a = getSus(index_a, t)
#                 sus_b = getSus(index_b, t)
#                 if sus_b < sus_a:
#                     id.append(j + 1)
#                 elif sus_b == sus_a:
#                     id.append(random.choice([0, j + 1]))
#                 else:
#                     id.append(0)
#             # if MG[0].count(1) < len(MG[0]) * 0.3:
#             #     # 策略一，投票选最多的，不去重x
#             #     pass
#             # elif MG[0].count(1) > len(MG[0]) * 0.6:
#             #     # 策略三，删掉x
#             #     id = list(set(id))
#             #     if 0 not in id:
#             #         pass
#             #     else:
#             #         id.remove(0)
#             #         if len(id) == 0:
#             #             id.append(0)
#             # else:
#             #     # 策略二，投票选最多的，不去重x
#             id = list(set(id))
#             # if 0 in id:
#             #     id.remove(0)
#             # if len(id) == 0:
#             #     id.append(0)
#             output = []
#             for ii in id:
#                 output.append(Output[1:][ii])
#             m = showmax(output)
#             m = random.choice(m)
#             if output.count(m) > 1:
#                 if m == Output[0]:
#                     true = 1
#             else:
#                 iid = id.copy()
#                 for m in range(len(id) - 1):
#                     for n in range(m + 1, len(id)):
#                         sum_s = MG[id[m]].count(0) + MG[id[m]].count(3) + MG[id[n]].count(0) + MG[id[n]].count(
#                             3)
#                         sum_v = MG[id[m]].count(1) + MG[id[n]].count(1)
#                         # 求每个测试用例的ev es nv ns
#                         ev_a = MG[id[m]].count(1)
#                         es_a = MG[id[m]].count(0) + MG[id[m]].count(3)
#                         ev_b = MG[id[n]].count(1)
#                         es_b = MG[id[n]].count(0) + MG[id[n]].count(3)
#                         if MG[0][id[n] - 1] == 1:
#                             ev_b += 1
#                         if MG[0][id[n] - 1] == 0 or MG[0][id[n] - 1] == 3:
#                             es_b += 1
#                         if id[m] == 0:
#                             pass
#                         else:
#                             if MG[0][id[m] - 1] == 1:
#                                 ev_a += 1
#                                 sum_v += 1
#                             if MG[0][id[m] - 1] == 0 or MG[0][id[m] - 1] == 3:
#                                 es_a += 1
#                                 sum_s += 1
#                             if MG[0][id[n] - 1] == 1:
#                                 sum_v += 1
#                             if MG[0][id[n] - 1] == 0 or MG[0][id[n] - 1] == 3:
#                                 sum_s += 1
#                         nv_a = sum_v - ev_a
#                         ns_a = sum_s - es_a
#                         index_a = [ev_a, es_a, nv_a, ns_a]
#                         nv_b = sum_v - ev_b
#                         ns_b = sum_s - es_b
#                         index_b = [ev_b, es_b, nv_b, ns_b]
#                         # 求测试用例的可疑度
#                         # for t in range(len(Formula)):
#                         sus_a = getSus(index_a, t)
#                         sus_b = getSus(index_b, t)
#                         if sus_a == sus_b:
#                             pass
#                         elif sus_a > sus_b:
#                             if id[m] in iid:
#                                 iid.remove(id[m])
#                         else:
#                             if id[n] in iid:
#                                 iid.remove(id[n])
#
#                 if len(iid) == 0:
#                     index = random.choice(id)
#                     if Output[1:][index] == Output[0]:
#                         true = 1
#                 else:
#                     if Output[1:][iid[0]] == Output[0]:
#                         true = 1
#             flag.append(true)
#
#     return flag
#
#
# def FaultTolerance2(MG, pf, Output, EMR):
#     Formula = ['Naish1', 'Naish2', 'Wong1', 'Russel&Rao', 'Binary', 'Jaccard', 'Anderberg', 'Sørensen-Dice', 'Dice',
#                'Goodman', 'Tarantula', 'qe', 'CBI Inc.', 'Wong2', 'Hamann', 'Simple Matching', 'Sokal',
#                'Rogers&Tanimoto', 'Hamming etc.', 'Euclid', 'Scott', 'Rogot1', 'Kulczynski2', 'Ochiai',
#                'M2', 'AMPLE2', 'Wong3', 'Arithmetic Mean', 'Cohen', 'Fleiss']
#     flag = []
#     true = 0
#
#     # a = random.choice(MG[0])
#     # if a == 0 or a == 3:
#     #     if pf[0] == 0:
#     #         true = 1
#     #     for _ in range(len(Formula)):
#     #         flag.append(true)
#     #     return flag
#     # a = random.sample(MG[0], 7)
#     if 1 not in MG[0]:
#         if pf[0] == 0:
#             true = 1
#         for _ in range(len(Formula)):
#             flag.append(true)
#         return flag
#     else:
#         for t in range(len(Formula)):
#             true = 0
#             id = []
#             for j in range(len(MG[0])):
#                 if j not in EMR:
#                     continue
#                 # 根据包含的测试用例确定 MG set
#                 sum_s = MG[0].count(0) + MG[0].count(3)
#                 sum_v = MG[0].count(1)
#                 sum_s += (MG[j + 1].count(0)) + MG[j + 1].count(3)
#                 sum_v += MG[j + 1].count(1)
#                 # 求每个测试用例的ev es nv ns
#                 ev_a = MG[0].count(1)
#                 es_a = MG[0].count(0) + MG[0].count(3)
#                 nv_a = sum_v - ev_a
#                 ns_a = sum_s - es_a
#                 index_a = [ev_a, es_a, nv_a, ns_a]
#
#                 ev_b = MG[j + 1].count(1) + 1
#                 es_b = MG[j + 1].count(0) + MG[j + 1].count(3)
#                 if MG[0][j] == 1:
#                     ev_b += 1
#                 elif MG[0][j] == 0 or MG[0][j] == 3:
#                     es_b += 1
#                 nv_b = sum_v - ev_b
#                 ns_b = sum_s - es_b
#                 index_b = [ev_b, es_b, nv_b, ns_b]
#                 # 求测试用例的可疑度
#                 # for t in range(len(Formula)):
#                 sus_a = getSus(index_a, t)
#                 sus_b = getSus(index_b, t)
#                 if sus_b < sus_a:
#                     id.append(j + 1)
#                 elif sus_b == sus_a:
#                     id.append(random.choice([0, j + 1]))
#                 else:
#                     id.append(0)
#
#             # if MG[0].count(1) < len(MG[0]) * 0.3:
#             #     # 策略一，投票选最多的，不去重x
#             #     pass
#             # elif MG[0].count(1) > len(MG[0]) * 0.6:
#             #     # 策略三，删掉x
#             #     id = list(set(id))
#             #     if 0 not in id:
#             #         pass
#             #     else:
#             #         id.remove(0)
#             #         if len(id) == 0:
#             #             id.append(0)
#             # else:
#             #     # 策略二，投票选最多的，不去重x
#             id = list(set(id))
#             if 0 in id:
#                 id.remove(0)
#             if len(id) == 0:
#                 id.append(0)
#             output = []
#             for ii in id:
#                 output.append(Output[1:][ii])
#             m = showmax(output)
#             m = random.choice(m)
#             if output.count(m) > 1:
#                 if m == Output[0]:
#                     true = 1
#             else:
#                 iid = id.copy()
#                 for m in range(len(id) - 1):
#                     for n in range(m + 1, len(id)):
#                         sum_s = MG[id[m]].count(0) + MG[id[m]].count(3) + MG[id[n]].count(0) + MG[id[n]].count(
#                             3)
#                         sum_v = MG[id[m]].count(1) + MG[id[n]].count(1)
#                         # 求每个测试用例的ev es nv ns
#                         ev_a = MG[id[m]].count(1)
#                         es_a = MG[id[m]].count(0) + MG[id[m]].count(3)
#                         ev_b = MG[id[n]].count(1)
#                         es_b = MG[id[n]].count(0) + MG[id[n]].count(3)
#                         if MG[0][id[n] - 1] == 1:
#                             ev_b += 1
#                         if MG[0][id[n] - 1] == 0 or MG[0][id[n] - 1] == 3:
#                             es_b += 1
#                         if id[m] == 0:
#                             pass
#                         else:
#                             if MG[0][id[m] - 1] == 1:
#                                 ev_a += 1
#                                 sum_v += 1
#                             if MG[0][id[m] - 1] == 0 or MG[0][id[m] - 1] == 3:
#                                 es_a += 1
#                                 sum_s += 1
#                             if MG[0][id[n] - 1] == 1:
#                                 sum_v += 1
#                             if MG[0][id[n] - 1] == 0 or MG[0][id[n] - 1] == 3:
#                                 sum_s += 1
#                         nv_a = sum_v - ev_a
#                         ns_a = sum_s - es_a
#                         index_a = [ev_a, es_a, nv_a, ns_a]
#                         nv_b = sum_v - ev_b
#                         ns_b = sum_s - es_b
#                         index_b = [ev_b, es_b, nv_b, ns_b]
#                         # 求测试用例的可疑度
#                         # for t in range(len(Formula)):
#                         sus_a = getSus(index_a, t)
#                         sus_b = getSus(index_b, t)
#                         if sus_a == sus_b:
#                             pass
#                         elif sus_a > sus_b:
#                             if id[m] in iid:
#                                 iid.remove(id[m])
#                         else:
#                             if id[n] in iid:
#                                 iid.remove(id[n])
#
#                 if len(iid) == 0:
#                     index = random.choice(id)
#                     if Output[1:][index] == Output[0]:
#                         true = 1
#                 else:
#                     if Output[1:][iid[0]] == Output[0]:
#                         true = 1
#             flag.append(true)
#
#     return flag


def FaultTolerance3(MG, pf, Output):
    """
    voting
    """
    Formula = ['Naish1', 'Naish2', 'Wong1', 'Russel&Rao', 'Binary', 'Jaccard', 'Anderberg', 'Sørensen-Dice', 'Dice',
               'Goodman', 'Tarantula', 'qe', 'CBI Inc.', 'Wong2', 'Hamann', 'Simple Matching', 'Sokal',
               'Rogers&Tanimoto', 'Hamming etc.', 'Euclid', 'Scott', 'Rogot1', 'Kulczynski2', 'Ochiai',
               'M2', 'AMPLE2', 'Wong3', 'Arithmetic Mean', 'Cohen', 'Fleiss']
    flag = []
    true = 0
    A = []
    for _ in range(len(Formula)):
        A.append(0)
    if 1 not in MG[0]:
        if pf[0] == 0:
            true = 1
        for _ in range(len(Formula)):
            flag.append(true)
        return flag, A
    else:
        A = []
        output = Output[1:]
        m = showmax(output)
        if len(m) == 1:
            a = 0
            if m[0] == Output[0]:
                true = 1
        else:
            a = 1
            m = random.choice(m)
            if m == Output[0]:
                true = 1
        for _ in range(len(Formula)):
            flag.append(true)
            A.append(a)

    return flag, A


def FaultTolerance4(MG, pf, Output):
    """
    voting
    去掉 FS
    """
    Formula = ['Naish1', 'Naish2', 'Wong1', 'Russel&Rao', 'Binary', 'Jaccard', 'Anderberg', 'Sørensen-Dice', 'Dice',
               'Goodman', 'Tarantula', 'qe', 'CBI Inc.', 'Wong2', 'Hamann', 'Simple Matching', 'Sokal',
               'Rogers&Tanimoto', 'Hamming etc.', 'Euclid', 'Scott', 'Rogot1', 'Kulczynski2', 'Ochiai',
               'M2', 'AMPLE2', 'Wong3', 'Arithmetic Mean', 'Cohen', 'Fleiss']
    flag = []
    true = 0
    A = []
    for _ in range(len(Formula)):
        A.append(0)
    if 1 not in MG[0]:
        if pf[0] == 0:
            true = 1
        for _ in range(len(Formula)):
            flag.append(true)
        return flag, A
    else:
        A = []
        output = Output[1:]
        output1 = output[1:]  # 只包含follow
        index_list = [a for a, b in enumerate(MG[0]) if b == 3]
        output1 = [n for i, n in enumerate(output1) if i not in index_list]
        output1.insert(0, output[0])
        output = output1
        m = showmax(output)
        if len(m) == 1:
            a = 0
            if m[0] == Output[0]:
                true = 1
        else:
            a = 1
            m = random.choice(m)
            if m == Output[0]:
                true = 1
        for _ in range(len(Formula)):
            flag.append(true)
            A.append(a)

    return flag, A


# def FaultTolerance4(MG, pf, Output, EMR):
#     Formula = ['Naish1', 'Naish2', 'Wong1', 'Russel&Rao', 'Binary', 'Jaccard', 'Anderberg', 'Sørensen-Dice', 'Dice',
#                'Goodman', 'Tarantula', 'qe', 'CBI Inc.', 'Wong2', 'Hamann', 'Simple Matching', 'Sokal',
#                'Rogers&Tanimoto', 'Hamming etc.', 'Euclid', 'Scott', 'Rogot1', 'Kulczynski2', 'Ochiai',
#                'M2', 'AMPLE2', 'Wong3', 'Arithmetic Mean', 'Cohen', 'Fleiss']
#     flag = []
#     true = 0
#     A = []
#     for _ in range(len(Formula)):
#         A.append(0)
#     if 1 not in MG[0]:
#         if pf[0] == 0:
#             true = 1
#         for _ in range(len(Formula)):
#             flag.append(true)
#         return flag, A
#     else:
#         A = []
#         output = [Output[1:][0]]
#         for j in range(len(MG[0])):
#             if j not in EMR:
#                 continue
#             output.append(Output[1:][j + 1])
#         m = showmax(output)
#         if len(m) == 1:
#             a = 0
#             if m[0] == Output[0]:
#                 true = 1
#         else:
#             a = 1
#             m = random.choice(m)
#             if m == Output[0]:
#                 true = 1
#         for _ in range(len(Formula)):
#             flag.append(true)
#             A.append(a)
#
#     return flag, A


# def FaultTolerance5(MG, pf, Output):
#     """
#     对v11
#     """
#     Formula = ['Naish1', 'Naish2', 'Wong1', 'Russel&Rao', 'Binary', 'Jaccard', 'Anderberg', 'Sørensen-Dice', 'Dice',
#                'Goodman', 'Tarantula', 'qe', 'CBI Inc.', 'Wong2', 'Hamann', 'Simple Matching', 'Sokal',
#                'Rogers&Tanimoto', 'Hamming etc.', 'Euclid', 'Scott', 'Rogot1', 'Kulczynski2', 'Ochiai',
#                'M2', 'AMPLE2', 'Wong3', 'Arithmetic Mean', 'Cohen', 'Fleiss']
#     flag = []
#     true = 0
#
#     # a = random.choice(MG[0])
#     # if a == 0 or a == 3:
#     #     if pf[0] == 0:
#     #         true = 1
#     #     for _ in range(len(Formula)):
#     #         flag.append(true)
#     #     return flag
#     # a = random.sample(MG[0], 3)
#     if 1 not in MG[0]:
#         if pf[0] == 0:
#             true = 1
#         for _ in range(len(Formula)):
#             flag.append(true)
#         return flag
#     else:
#         for t in range(len(Formula)):
#             true = 0
#             id = []
#             for j in range(len(MG[0])):
#                 # 根据包含的测试用例确定 MG set
#                 sum_s = MG[0].count(0) + MG[0].count(3)
#                 sum_v = MG[0].count(1)
#                 sum_s += (MG[j + 1].count(0)) + MG[j + 1].count(3)
#                 sum_v += MG[j + 1].count(1)
#                 # 求每个测试用例的ev es nv ns
#                 ev_a = MG[0].count(1)
#                 es_a = MG[0].count(0) + MG[0].count(3)
#                 nv_a = sum_v - ev_a
#                 ns_a = sum_s - es_a
#                 index_a = [ev_a, es_a, nv_a, ns_a]
#
#                 ev_b = MG[j + 1].count(1) + 1
#                 es_b = MG[j + 1].count(0) + MG[j + 1].count(3)
#                 if MG[0][j] == 1:
#                     ev_b += 1
#                 elif MG[0][j] == 0 or MG[0][j] == 3:
#                     es_b += 1
#                 nv_b = sum_v - ev_b
#                 ns_b = sum_s - es_b
#                 index_b = [ev_b, es_b, nv_b, ns_b]
#                 # 求测试用例的可疑度
#                 # for t in range(len(Formula)):
#                 sus_a = getSus(index_a, t)
#                 sus_b = getSus(index_b, t)
#                 if sus_b < sus_a:
#                     id.append(j + 1)
#                 elif sus_b == sus_a:
#                     id.append(random.choice([0, j + 1]))
#                 else:
#                     id.append(0)
#             # if MG[0].count(1) < len(MG[0]) * 0.3:
#             #     # 策略一，投票选最多的，不去重x
#             #     pass
#             # elif MG[0].count(1) > len(MG[0]) * 0.6:
#             #     # 策略三，删掉x
#             #     id = list(set(id))
#             #     if 0 not in id:
#             #         pass
#             #     else:
#             #         id.remove(0)
#             #         if len(id) == 0:
#             #             id.append(0)
#             # else:
#             #     # 策略二，投票选最多的，不去重x
#             id = list(set(id))
#             # # 策略四
#             if 0 in id:
#                 id.remove(0)
#             if len(id) == 0:
#                 id.append(0)
#             output = []
#             for ii in id:
#                 output.append(Output[1][ii])
#             m = showmax(output)
#             m = random.choice(m)
#             ind = Output[1].index(m)
#             if output.count(m) > 1:
#                 if m == Output[0][ind]:
#                     true = 1
#             else:
#                 iid = id.copy()
#                 for m in range(len(id) - 1):
#                     for n in range(m + 1, len(id)):
#                         sum_s = MG[id[m]].count(0) + MG[id[m]].count(3) + MG[id[n]].count(0) + MG[id[n]].count(
#                             3)
#                         sum_v = MG[id[m]].count(1) + MG[id[n]].count(1)
#                         # 求每个测试用例的ev es nv ns
#                         ev_a = MG[id[m]].count(1)
#                         es_a = MG[id[m]].count(0) + MG[id[m]].count(3)
#                         ev_b = MG[id[n]].count(1)
#                         es_b = MG[id[n]].count(0) + MG[id[n]].count(3)
#                         if MG[0][id[n] - 1] == 1:
#                             ev_b += 1
#                         if MG[0][id[n] - 1] == 0 or MG[0][id[n] - 1] == 3:
#                             es_b += 1
#                         if id[m] == 0:
#                             pass
#                         else:
#                             if MG[0][id[m] - 1] == 1:
#                                 ev_a += 1
#                                 sum_v += 1
#                             if MG[0][id[m] - 1] == 0 or MG[0][id[m] - 1] == 3:
#                                 es_a += 1
#                                 sum_s += 1
#                             if MG[0][id[n] - 1] == 1:
#                                 sum_v += 1
#                             if MG[0][id[n] - 1] == 0 or MG[0][id[n] - 1] == 3:
#                                 sum_s += 1
#                         nv_a = sum_v - ev_a
#                         ns_a = sum_s - es_a
#                         index_a = [ev_a, es_a, nv_a, ns_a]
#                         nv_b = sum_v - ev_b
#                         ns_b = sum_s - es_b
#                         index_b = [ev_b, es_b, nv_b, ns_b]
#                         # 求测试用例的可疑度
#                         # for t in range(len(Formula)):
#                         sus_a = getSus(index_a, t)
#                         sus_b = getSus(index_b, t)
#                         if sus_a == sus_b:
#                             pass
#                         elif sus_a > sus_b:
#                             if id[m] in iid:
#                                 iid.remove(id[m])
#                         else:
#                             if id[n] in iid:
#                                 iid.remove(id[n])
#
#                 if len(iid) == 0:
#                     index = random.choice(id)
#                     if Output[1][index] == Output[0][index]:
#                         true = 1
#                 else:
#                     if Output[1][iid[0]] == Output[0][iid[0]]:
#                         true = 1
#             flag.append(true)
#
#     return flag
#
#
# def FaultTolerance6(MG, pf, Output):
#     """
#     数据多样性
#     """
#     Formula = ['Naish1', 'Naish2', 'Wong1', 'Russel&Rao', 'Binary', 'Jaccard', 'Anderberg', 'Sørensen-Dice', 'Dice',
#                'Goodman', 'Tarantula', 'qe', 'CBI Inc.', 'Wong2', 'Hamann', 'Simple Matching', 'Sokal',
#                'Rogers&Tanimoto', 'Hamming etc.', 'Euclid', 'Scott', 'Rogot1', 'Kulczynski2', 'Ochiai',
#                'M2', 'AMPLE2', 'Wong3', 'Arithmetic Mean', 'Cohen', 'Fleiss']
#     flag = []
#     true = 0
#     if 1 not in MG[0]:
#         if pf[0] == 0:
#             true = 1
#         for _ in range(len(Formula)):
#             flag.append(true)
#         return flag
#     else:
#         t1 = 0
#         t2 = 0
#         for i in range(len(Output[1])):
#             if Output[1][i] == Output[0][i]:
#                 t1 += 1
#             else:
#                 t2 += 1
#         if t1 > t2:
#             true = 1
#         elif t1 == t2:
#             true = random.choice([0, 1])
#         else:
#             true = 0
#         for _ in range(len(Formula)):
#             flag.append(true)
#     return flag
#
#
# def FaultTolerance7(MG, pf, Output):
#     Formula = ['Naish1', 'Naish2', 'Wong1', 'Russel&Rao', 'Binary', 'Jaccard', 'Anderberg', 'Sørensen-Dice', 'Dice',
#                'Goodman', 'Tarantula', 'qe', 'CBI Inc.', 'Wong2', 'Hamann', 'Simple Matching', 'Sokal',
#                'Rogers&Tanimoto', 'Hamming etc.', 'Euclid', 'Scott', 'Rogot1', 'Kulczynski2', 'Ochiai',
#                'M2', 'AMPLE2', 'Wong3', 'Arithmetic Mean', 'Cohen', 'Fleiss']
#     flag = []
#     true = 0
#     a = 0
#     b = 0
#     aa = []
#     bb = []
#     # a = random.choice(MG[0])
#     # if a == 0 or a == 3:
#     #     if pf[0] == 0:
#     #         true = 1
#     #     for _ in range(len(Formula)):
#     #         flag.append(true)
#     #     return flag
#     # a = random.sample(MG[0], 3)
#     if 1 not in MG[0]:
#         a = 1
#         if pf[0] == 0:
#             true = 1
#             b = 11
#         else:
#             b = 12
#         for _ in range(len(Formula)):
#             flag.append(true)
#             aa.append(a)
#             bb.append(b)
#         return flag, aa, bb
#     else:
#         out = Output[1:]
#         result = out.count(out[0]) == len(out)
#         if result:
#             a = 2
#             if out[0] == Output[0]:
#                 true = 1
#                 b = 21
#             else:
#                 b = 22
#             for _ in range(len(Formula)):
#                 flag.append(true)
#                 aa.append(a)
#                 bb.append(b)
#             return flag, aa, bb
#         else:
#             for t in range(len(Formula)):
#                 true = 0
#                 a = 0
#                 b = 0
#                 id = []
#                 for j in range(len(MG[0])):
#                     # 根据包含的测试用例确定 MG set
#                     sum_s = MG[0].count(0) + MG[0].count(3)
#                     sum_v = MG[0].count(1)
#                     sum_s += (MG[j + 1].count(0)) + MG[j + 1].count(3)
#                     sum_v += MG[j + 1].count(1)
#                     # 求每个测试用例的ev es nv ns
#                     ev_a = MG[0].count(1)
#                     es_a = MG[0].count(0) + MG[0].count(3)
#                     nv_a = sum_v - ev_a
#                     ns_a = sum_s - es_a
#                     index_a = [ev_a, es_a, nv_a, ns_a]
#
#                     ev_b = MG[j + 1].count(1) + 1
#                     es_b = MG[j + 1].count(0) + MG[j + 1].count(3)
#                     if MG[0][j] == 1:
#                         ev_b += 1
#                     elif MG[0][j] == 0 or MG[0][j] == 3:
#                         es_b += 1
#                     nv_b = sum_v - ev_b
#                     ns_b = sum_s - es_b
#                     index_b = [ev_b, es_b, nv_b, ns_b]
#                     # 求测试用例的可疑度
#                     sus_a = getSus(index_a, t)
#                     sus_b = getSus(index_b, t)
#                     if sus_b < sus_a:
#                         id.append(j + 1)
#                     elif sus_b == sus_a:
#                         id.append(random.choice([0, j + 1]))
#                     else:
#                         id.append(0)
#                 id = list(set(id))
#                 if 0 in id:
#                     id.remove(0)
#                 if len(id) == 0:
#                     id.append(0)
#                 output = []
#                 for ii in id:
#                     output.append(Output[1:][ii])
#                 result = output.count(output[0]) == len(output)
#                 if result:
#                     a = 3
#                     if output[0] == Output[0]:
#                         true = 1
#                         b = 31
#                     else:
#                         b = 32
#                 else:
#                     iid = id.copy()
#                     for m in range(len(id) - 1):
#                         for n in range(m + 1, len(id)):
#                             sum_s = MG[id[m]].count(0) + MG[id[m]].count(3) + MG[id[n]].count(0) + MG[id[n]].count(
#                                 3)
#                             sum_v = MG[id[m]].count(1) + MG[id[n]].count(1)
#                             # 求每个测试用例的ev es nv ns
#                             ev_a = MG[id[m]].count(1)
#                             es_a = MG[id[m]].count(0) + MG[id[m]].count(3)
#                             ev_b = MG[id[n]].count(1)
#                             es_b = MG[id[n]].count(0) + MG[id[n]].count(3)
#                             if MG[0][id[n] - 1] == 1:
#                                 ev_b += 1
#                             if MG[0][id[n] - 1] == 0 or MG[0][id[n] - 1] == 3:
#                                 es_b += 1
#                             if id[n] == 0:
#                                 pass
#                             else:
#                                 if MG[0][id[n] - 1] == 1:
#                                     ev_b += 1
#                                     sum_v += 1
#                                 if MG[0][id[n] - 1] == 0 or MG[0][id[n] - 1] == 3:
#                                     es_b += 1
#                                     sum_s += 1
#                             if id[m] == 0:
#                                 pass
#                             else:
#                                 if MG[0][id[m] - 1] == 1:
#                                     ev_a += 1
#                                     sum_v += 1
#                                 if MG[0][id[m] - 1] == 0 or MG[0][id[m] - 1] == 3:
#                                     es_a += 1
#                                     sum_s += 1
#                             nv_a = sum_v - ev_a
#                             ns_a = sum_s - es_a
#                             index_a = [ev_a, es_a, nv_a, ns_a]
#                             nv_b = sum_v - ev_b
#                             ns_b = sum_s - es_b
#                             index_b = [ev_b, es_b, nv_b, ns_b]
#                             # 求测试用例的可疑度
#                             # for t in range(len(Formula)):
#                             sus_a = getSus(index_a, t)
#                             sus_b = getSus(index_b, t)
#                             if sus_a == sus_b:
#                                 pass
#                             elif sus_a > sus_b:
#                                 if id[m] in iid:
#                                     iid.remove(id[m])
#                             else:
#                                 if id[n] in iid:
#                                     iid.remove(id[n])
#
#                     if len(iid) == 0:
#                         a = 4
#                         index = random.choice(id)
#                         if Output[1:][index] == Output[0]:
#                             true = 1
#                             b = 41
#                         else:
#                             b = 42
#                     elif len(iid) == 1:
#                         a = 5
#                         if Output[1:][iid[0]] == Output[0]:
#                             true = 1
#                             b = 51
#                         else:
#                             b = 52
#                     else:
#                         a = 6
#                         index = random.choice(iid)
#                         if Output[1:][index] == Output[0]:
#                             true = 1
#                             b = 61
#                         else:
#                             b = 62
#                 flag.append(true)
#                 aa.append(a)
#                 bb.append(b)
#
#     return flag, aa, bb


# def FaultTolerance8(MG, pf, Output, EMR):
#     Formula = ['Naish1', 'Naish2', 'Wong1', 'Russel&Rao', 'Binary', 'Jaccard', 'Anderberg', 'Sørensen-Dice', 'Dice',
#                'Goodman', 'Tarantula', 'qe', 'CBI Inc.', 'Wong2', 'Hamann', 'Simple Matching', 'Sokal',
#                'Rogers&Tanimoto', 'Hamming etc.', 'Euclid', 'Scott', 'Rogot1', 'Kulczynski2', 'Ochiai',
#                'M2', 'AMPLE2', 'Wong3', 'Arithmetic Mean', 'Cohen', 'Fleiss']
#     flag = []
#     true = 0
#     a = 0
#     b = 0
#     aa = []
#     bb = []
#     # a = random.choice(MG[0])
#     # if a == 0 or a == 3:
#     #     if pf[0] == 0:
#     #         true = 1
#     #     for _ in range(len(Formula)):
#     #         flag.append(true)
#     #     return flag
#     # a = random.sample(MG[0], 7)
#     if 1 not in MG[0]:
#         a = 1
#         if pf[0] == 0:
#             true = 1
#             b = 11
#         else:
#             b = 12
#         for _ in range(len(Formula)):
#             flag.append(true)
#             aa.append(a)
#             bb.append(b)
#         return flag, aa, bb
#     else:
#         out = [Output[1:][0]]
#         for j in range(len(MG[0])):
#             if j not in EMR:
#                 continue
#             out.append(Output[1:][j + 1])
#         result = out.count(out[0]) == len(out)
#         if result:
#             a = 2
#             if out[0] == Output[0]:
#                 true = 1
#                 b = 21
#             else:
#                 b = 22
#             for _ in range(len(Formula)):
#                 flag.append(true)
#                 aa.append(a)
#                 bb.append(bb)
#             return flag, aa, bb
#         else:
#             for t in range(len(Formula)):
#                 true = 0
#                 a = 0
#                 b = 0
#                 id = []
#                 for j in range(len(MG[0])):
#                     if j not in EMR:
#                         continue
#                     # 根据包含的测试用例确定 MG set
#                     sum_s = MG[0].count(0) + MG[0].count(3)
#                     sum_v = MG[0].count(1)
#                     sum_s += (MG[j + 1].count(0)) + MG[j + 1].count(3)
#                     sum_v += MG[j + 1].count(1)
#                     # 求每个测试用例的ev es nv ns
#                     ev_a = MG[0].count(1)
#                     es_a = MG[0].count(0) + MG[0].count(3)
#                     nv_a = sum_v - ev_a
#                     ns_a = sum_s - es_a
#                     index_a = [ev_a, es_a, nv_a, ns_a]
#
#                     ev_b = MG[j + 1].count(1) + 1
#                     es_b = MG[j + 1].count(0) + MG[j + 1].count(3)
#                     if MG[0][j] == 1:
#                         ev_b += 1
#                     elif MG[0][j] == 0 or MG[0][j] == 3:
#                         es_b += 1
#                     nv_b = sum_v - ev_b
#                     ns_b = sum_s - es_b
#                     index_b = [ev_b, es_b, nv_b, ns_b]
#                     # 求测试用例的可疑度
#                     sus_a = getSus(index_a, t)
#                     sus_b = getSus(index_b, t)
#                     if sus_b < sus_a:
#                         id.append(j + 1)
#                     elif sus_b == sus_a:
#                         id.append(random.choice([0, j + 1]))
#                     else:
#                         id.append(0)
#
#                 id = list(set(id))
#                 if 0 in id:
#                     id.remove(0)
#                 if len(id) == 0:
#                     id.append(0)
#                 output = []
#                 for ii in id:
#                     output.append(Output[1:][ii])
#                 result = output.count(output[0]) == len(output)
#                 if result:
#                     a = 3
#                     if output[0] == Output[0]:
#                         true = 1
#                         b = 31
#                     else:
#                         b = 32
#                 else:
#                     iid = id.copy()
#                     for m in range(len(id) - 1):
#                         for n in range(m + 1, len(id)):
#                             sum_s = MG[id[m]].count(0) + MG[id[m]].count(3) + MG[id[n]].count(0) + MG[id[n]].count(
#                                 3)
#                             sum_v = MG[id[m]].count(1) + MG[id[n]].count(1)
#                             # 求每个测试用例的ev es nv ns
#                             ev_a = MG[id[m]].count(1)
#                             es_a = MG[id[m]].count(0) + MG[id[m]].count(3)
#                             ev_b = MG[id[n]].count(1)
#                             es_b = MG[id[n]].count(0) + MG[id[n]].count(3)
#                             if MG[0][id[n] - 1] == 1:
#                                 ev_b += 1
#                             if MG[0][id[n] - 1] == 0 or MG[0][id[n] - 1] == 3:
#                                 es_b += 1
#                             if id[n] == 0:
#                                 pass
#                             else:
#                                 if MG[0][id[n] - 1] == 1:
#                                     ev_b += 1
#                                     sum_v += 1
#                                 if MG[0][id[n] - 1] == 0 or MG[0][id[n] - 1] == 3:
#                                     es_b += 1
#                                     sum_s += 1
#                             if id[m] == 0:
#                                 pass
#                             else:
#                                 if MG[0][id[m] - 1] == 1:
#                                     ev_a += 1
#                                     sum_v += 1
#                                 if MG[0][id[m] - 1] == 0 or MG[0][id[m] - 1] == 3:
#                                     es_a += 1
#                                     sum_s += 1
#                             nv_a = sum_v - ev_a
#                             ns_a = sum_s - es_a
#                             index_a = [ev_a, es_a, nv_a, ns_a]
#                             nv_b = sum_v - ev_b
#                             ns_b = sum_s - es_b
#                             index_b = [ev_b, es_b, nv_b, ns_b]
#                             # 求测试用例的可疑度
#                             # for t in range(len(Formula)):
#                             sus_a = getSus(index_a, t)
#                             sus_b = getSus(index_b, t)
#                             if sus_a == sus_b:
#                                 pass
#                             elif sus_a > sus_b:
#                                 if id[m] in iid:
#                                     iid.remove(id[m])
#                             else:
#                                 if id[n] in iid:
#                                     iid.remove(id[n])
#
#                     if len(iid) == 0:
#                         a = 4
#                         index = random.choice(id)
#                         if Output[1:][index] == Output[0]:
#                             true = 1
#                             b = 41
#                         else:
#                             b = 42
#                     elif len(iid) == 1:
#                         a = 5
#                         if Output[1:][iid[0]] == Output[0]:
#                             true = 1
#                             b = 51
#                         else:
#                             b = 52
#                     else:
#                         a = 6
#                         index = random.choice(iid)
#                         if Output[1:][index] == Output[0]:
#                             true = 1
#                             b = 61
#                         else:
#                             b = 62
#                 flag.append(true)
#                 aa.append(a)
#                 bb.append(b)
#     return flag, aa, bb


def FaultTolerance9(MG, pf, Output):
    """
    对v11
    """
    Formula = ['Naish1', 'Naish2', 'Wong1', 'Russel&Rao', 'Binary', 'Jaccard', 'Anderberg', 'Sørensen-Dice', 'Dice',
               'Goodman', 'Tarantula', 'qe', 'CBI Inc.', 'Wong2', 'Hamann', 'Simple Matching', 'Sokal',
               'Rogers&Tanimoto', 'Hamming etc.', 'Euclid', 'Scott', 'Rogot1', 'Kulczynski2', 'Ochiai',
               'M2', 'AMPLE2', 'Wong3', 'Arithmetic Mean', 'Cohen', 'Fleiss']
    flag = []
    true = 0
    a = 0
    b = 0
    aa = []
    bb = []
    if 1 not in MG[0]:
        a = 1
        if pf[0] == 0:
            true = 1
            b = 11
        else:
            b = 12
        for _ in range(len(Formula)):
            flag.append(true)
            aa.append(a)
            bb.append(b)
        return flag, aa, bb
    else:
        result = 1
        for i in range(len(Output[0])):
            if Output[1][i] != Output[0][i]:
                result = 0
                break
        if result:
            true = 1
            b = 21
            for _ in range(len(Formula)):
                flag.append(true)
                a = 2
                aa.append(a)
                bb.append(b)
            return flag, aa, bb
        else:
            for t in range(len(Formula)):
                true = 0
                a = 0
                b = 0
                id = []
                for j in range(len(MG[0])):
                    # 根据包含的测试用例确定 MG set
                    sum_s = MG[0].count(0) + MG[0].count(3)
                    sum_v = MG[0].count(1)
                    sum_s += (MG[j + 1].count(0)) + MG[j + 1].count(3)
                    sum_v += MG[j + 1].count(1)
                    # 求每个测试用例的ev es nv ns
                    ev_a = MG[0].count(1)
                    es_a = MG[0].count(0) + MG[0].count(3)
                    nv_a = sum_v - ev_a
                    ns_a = sum_s - es_a
                    index_a = [ev_a, es_a, nv_a, ns_a]
                    ev_b = MG[j + 1].count(1) + 1
                    es_b = MG[j + 1].count(0) + MG[j + 1].count(3)
                    if MG[0][j] == 1:
                        ev_b += 1
                    elif MG[0][j] == 0 or MG[0][j] == 3:
                        es_b += 1
                    nv_b = sum_v - ev_b
                    ns_b = sum_s - es_b
                    index_b = [ev_b, es_b, nv_b, ns_b]
                    # 求测试用例的可疑度
                    # for t in range(len(Formula)):
                    sus_a = getSus(index_a, t)
                    sus_b = getSus(index_b, t)
                    if sus_b < sus_a:
                        id.append(j + 1)
                    elif sus_b == sus_a:
                        id.append(random.choice([0, j + 1]))
                    else:
                        id.append(0)
                id = list(set(id))
                if 0 in id:
                    id.remove(0)
                if len(id) == 0:
                    id.append(0)
                output = []
                result = 1
                for ii in id:
                    output.append(Output[1][ii])
                    if Output[1][ii] != Output[0][ii]:
                        result = 0
                if result:
                    true = 1
                    b = 31
                    a = 3
                else:
                    iid = id.copy()
                    for m in range(len(id) - 1):
                        for n in range(m + 1, len(id)):
                            sum_s = MG[id[m]].count(0) + MG[id[m]].count(3) + MG[id[n]].count(0) + MG[id[n]].count(3)
                            sum_v = MG[id[m]].count(1) + MG[id[n]].count(1)
                            # 求每个测试用例的ev es nv ns
                            ev_a = MG[id[m]].count(1)
                            es_a = MG[id[m]].count(0) + MG[id[m]].count(3)
                            ev_b = MG[id[n]].count(1)
                            es_b = MG[id[n]].count(0) + MG[id[n]].count(3)
                            if id[n] == 0:
                                pass
                            else:
                                if MG[0][id[n] - 1] == 1:
                                    ev_b += 1
                                    sum_v += 1
                                if MG[0][id[n] - 1] == 0 or MG[0][id[n] - 1] == 3:
                                    es_b += 1
                                    sum_s += 1
                            if id[m] == 0:
                                pass
                            else:
                                if MG[0][id[m] - 1] == 1:
                                    ev_a += 1
                                    sum_v += 1
                                if MG[0][id[m] - 1] == 0 or MG[0][id[m] - 1] == 3:
                                    es_a += 1
                                    sum_s += 1
                            nv_a = sum_v - ev_a
                            ns_a = sum_s - es_a
                            index_a = [ev_a, es_a, nv_a, ns_a]
                            nv_b = sum_v - ev_b
                            ns_b = sum_s - es_b
                            index_b = [ev_b, es_b, nv_b, ns_b]
                            # 求测试用例的可疑度
                            # for t in range(len(Formula)):
                            sus_a = getSus(index_a, t)
                            sus_b = getSus(index_b, t)
                            if sus_a == sus_b:
                                pass
                            elif sus_a > sus_b:
                                if id[m] in iid:
                                    iid.remove(id[m])
                            else:
                                if id[n] in iid:
                                    iid.remove(id[n])

                    if len(iid) == 0:
                        a = 4
                        index = random.choice(id)
                        if Output[1][index] == Output[0][index]:
                            true = 1
                            b = 41
                        else:
                            b = 42
                    elif len(iid) == 1:
                        a = 5
                        if Output[1][iid[0]] == Output[0][iid[0]]:
                            true = 1
                            b = 51
                        else:
                            b = 52
                    else:
                        a = 6
                        index = random.choice(iid)
                        if Output[1][index] == Output[0][index]:
                            true = 1
                            b = 61
                        else:
                            b = 62
                flag.append(true)
                aa.append(a)
                bb.append(b)
    return flag, aa, bb


def FaultTolerance10(MG, pf, Output):
    Formula = ['Naish1', 'Naish2', 'Wong1', 'Russel&Rao', 'Binary', 'Jaccard', 'Anderberg', 'Sørensen-Dice', 'Dice',
               'Goodman', 'Tarantula', 'qe', 'CBI Inc.', 'Wong2', 'Hamann', 'Simple Matching', 'Sokal',
               'Rogers&Tanimoto', 'Hamming etc.', 'Euclid', 'Scott', 'Rogot1', 'Kulczynski2', 'Ochiai',
               'M2', 'AMPLE2', 'Wong3', 'Arithmetic Mean', 'Cohen', 'Fleiss']
    flag = []
    true = 0
    if 1 not in MG[0]:
        if pf[0] == 0:
            true = 1
        for _ in range(len(Formula)):
            flag.append(true)
        return flag
    else:
        out = Output[1:]
        result = out.count(out[0]) == len(out)
        if result:
            if out[0] == Output[0]:
                true = 1
            for _ in range(len(Formula)):
                flag.append(true)
            return flag
        else:
            for t in range(len(Formula)):
                true = 0
                sum_s = 0
                sum_v = 0
                sus = []
                for k in range(len(MG)):
                    sum_s += MG[k].count(0) + MG[k].count(3)
                    sum_v += MG[k].count(1)
                for j in range(len(MG[0])):
                    # 求每个测试用例的ev es nv ns
                    ev_a = MG[0].count(1)
                    es_a = MG[0].count(0) + MG[0].count(3)
                    nv_a = sum_v - ev_a
                    ns_a = sum_s - es_a
                    index_a = [ev_a, es_a, nv_a, ns_a]

                    ev_b = MG[j + 1].count(1) + 1
                    es_b = MG[j + 1].count(0) + MG[j + 1].count(3)
                    if MG[0][j] == 1:
                        ev_b += 1
                    elif MG[0][j] == 0 or MG[0][j] == 3:
                        es_b += 1
                    nv_b = sum_v - ev_b
                    ns_b = sum_s - es_b
                    index_b = [ev_b, es_b, nv_b, ns_b]
                    # 求测试用例的可疑度
                    sus_a = getSus(index_a, t)
                    sus_b = getSus(index_b, t)
                    if j == 0:
                        sus.append(sus_a)
                        sus.append(sus_b)
                    else:
                        sus.append(sus_b)
                minsus = showmin(sus)
                index = []
                mins = minsus.copy()
                for i in minsus:
                    index.append(sus.index(i))
                    mins.remove(i)
                    if len(mins) == 0:
                        break
                index = random.choice(index)
                if Output[1:][index] == Output[0]:
                    true = 1
                flag.append(true)

    return flag


# def FaultTolerance11(MG, pf, Output, EMR):
#     Formula = ['Naish1', 'Naish2', 'Wong1', 'Russel&Rao', 'Binary', 'Jaccard', 'Anderberg', 'Sørensen-Dice', 'Dice',
#                'Goodman', 'Tarantula', 'qe', 'CBI Inc.', 'Wong2', 'Hamann', 'Simple Matching', 'Sokal',
#                'Rogers&Tanimoto', 'Hamming etc.', 'Euclid', 'Scott', 'Rogot1', 'Kulczynski2', 'Ochiai',
#                'M2', 'AMPLE2', 'Wong3', 'Arithmetic Mean', 'Cohen', 'Fleiss']
#     flag = []
#     true = 0
#     if 1 not in MG[0]:
#         if pf[0] == 0:
#             true = 1
#         for _ in range(len(Formula)):
#             flag.append(true)
#         return flag
#     else:
#         out = [Output[1:][0]]
#         for j in range(len(MG[0])):
#             if j not in EMR:
#                 continue
#             out.append(Output[1:][j + 1])
#         result = out.count(out[0]) == len(out)
#         if result:
#             if out[0] == Output[0]:
#                 true = 1
#             for _ in range(len(Formula)):
#                 flag.append(true)
#             return flag
#         else:
#             for t in range(len(Formula)):
#                 true = 0
#                 sum_s = 0
#                 sum_v = 0
#                 sus = []
#                 for k in range(len(MG)):
#                     sum_s += MG[k].count(0) + MG[k].count(3)
#                     sum_v += MG[k].count(1)
#                 for j in range(len(MG[0])):
#                     if j not in EMR:
#                         continue
#                     # 根据包含的测试用例确定 MG set
#                     sum_s = MG[0].count(0) + MG[0].count(3)
#                     sum_v = MG[0].count(1)
#                     sum_s += (MG[j + 1].count(0)) + MG[j + 1].count(3)
#                     sum_v += MG[j + 1].count(1)
#                     # 求每个测试用例的ev es nv ns
#                     ev_a = MG[0].count(1)
#                     es_a = MG[0].count(0) + MG[0].count(3)
#                     nv_a = sum_v - ev_a
#                     ns_a = sum_s - es_a
#                     index_a = [ev_a, es_a, nv_a, ns_a]
#
#                     ev_b = MG[j + 1].count(1) + 1
#                     es_b = MG[j + 1].count(0) + MG[j + 1].count(3)
#                     if MG[0][j] == 1:
#                         ev_b += 1
#                     elif MG[0][j] == 0 or MG[0][j] == 3:
#                         es_b += 1
#                     nv_b = sum_v - ev_b
#                     ns_b = sum_s - es_b
#                     index_b = [ev_b, es_b, nv_b, ns_b]
#                     # 求测试用例的可疑度
#                     # for t in range(len(Formula)):
#                     sus_a = getSus(index_a, t)
#                     sus_b = getSus(index_b, t)
#                     if j == 0:
#                         sus.append(sus_a)
#                         sus.append(sus_b)
#                     else:
#                         sus.append(sus_b)
#                 minsus = showmin(sus)
#                 index = []
#                 mins = minsus.copy()
#                 for i in minsus:
#                     index.append(sus.index(i))
#                     mins.remove(i)
#                     if len(mins) == 0:
#                         break
#                 index = random.choice(index)
#                 if Output[1:][index] == Output[0]:
#                     true = 1
#                 flag.append(true)
#
#     return flag


# def FaultTolerance12(MG, pf, Output):
#     """
#     对v11
#     """
#     Formula = ['Naish1', 'Naish2', 'Wong1', 'Russel&Rao', 'Binary', 'Jaccard', 'Anderberg', 'Sørensen-Dice', 'Dice',
#                'Goodman', 'Tarantula', 'qe', 'CBI Inc.', 'Wong2', 'Hamann', 'Simple Matching', 'Sokal',
#                'Rogers&Tanimoto', 'Hamming etc.', 'Euclid', 'Scott', 'Rogot1', 'Kulczynski2', 'Ochiai',
#                'M2', 'AMPLE2', 'Wong3', 'Arithmetic Mean', 'Cohen', 'Fleiss']
#     flag = []
#     true = 0
#     if 1 not in MG[0]:
#         if pf[0] == 0:
#             true = 1
#         for _ in range(len(Formula)):
#             flag.append(true)
#         return flag
#     else:
#         result = 1
#         for i in range(len(Output[0])):
#             if Output[1][i] != Output[0][i]:
#                 result = 0
#                 break
#         if result:
#             true = 1
#             for _ in range(len(Formula)):
#                 flag.append(true)
#             return flag
#         else:
#             for t in range(len(Formula)):
#                 true = 0
#                 sum_s = 0
#                 sum_v = 0
#                 sus = []
#                 for k in range(len(MG)):
#                     sum_s += MG[k].count(0) + MG[k].count(3)
#                     sum_v += MG[k].count(1)
#                 for j in range(len(MG[0])):
#                     # 根据包含的测试用例确定 MG set
#                     sum_s = MG[0].count(0) + MG[0].count(3)
#                     sum_v = MG[0].count(1)
#                     sum_s += (MG[j + 1].count(0)) + MG[j + 1].count(3)
#                     sum_v += MG[j + 1].count(1)
#                     # 求每个测试用例的ev es nv ns
#                     ev_a = MG[0].count(1)
#                     es_a = MG[0].count(0) + MG[0].count(3)
#                     nv_a = sum_v - ev_a
#                     ns_a = sum_s - es_a
#                     index_a = [ev_a, es_a, nv_a, ns_a]
#
#                     ev_b = MG[j + 1].count(1) + 1
#                     es_b = MG[j + 1].count(0) + MG[j + 1].count(3)
#                     if MG[0][j] == 1:
#                         ev_b += 1
#                     elif MG[0][j] == 0 or MG[0][j] == 3:
#                         es_b += 1
#                     nv_b = sum_v - ev_b
#                     ns_b = sum_s - es_b
#                     index_b = [ev_b, es_b, nv_b, ns_b]
#                     # 求测试用例的可疑度
#                     # for t in range(len(Formula)):
#                     sus_a = getSus(index_a, t)
#                     sus_b = getSus(index_b, t)
#                     if j == 0:
#                         sus.append(sus_a)
#                         sus.append(sus_b)
#                     else:
#                         sus.append(sus_b)
#                 minsus = showmin(sus)
#                 index = []
#                 mins = minsus.copy()
#                 for i in minsus:
#                     index.append(sus.index(i))
#                     mins.remove(i)
#                     if len(mins) == 0:
#                         break
#                 index = random.choice(index)
#                 if Output[1][index] == Output[0][index]:
#                     true = 1
#                 flag.append(true)
#
#     return flag
#
#
# def FaultTolerance13(MG, pf, Output):
#     """
#     group分组，failtim
#     """
#     Formula = ['Naish1', 'Naish2', 'Wong1', 'Russel&Rao', 'Binary', 'Jaccard', 'Anderberg', 'Sørensen-Dice', 'Dice',
#                'Goodman', 'Tarantula', 'qe', 'CBI Inc.', 'Wong2', 'Hamann', 'Simple Matching', 'Sokal',
#                'Rogers&Tanimoto', 'Hamming etc.', 'Euclid', 'Scott', 'Rogot1', 'Kulczynski2', 'Ochiai',
#                'M2', 'AMPLE2', 'Wong3', 'Arithmetic Mean', 'Cohen', 'Fleiss']
#     flag = []
#     true = 0
#     A = []
#     for _ in range(len(Formula)):
#         A.append(0)
#     if 1 not in MG[0]:
#         if pf[0] == 0:
#             true = 1
#         for _ in range(len(Formula)):
#             flag.append(true)
#         return flag, A
#     else:
#         out = Output[1:]
#         result = out.count(out[0]) == len(out)
#         if result:
#             if out[0] == Output[0]:
#                 true = 1
#             for _ in range(len(Formula)):
#                 flag.append(true)
#             return flag, A
#         else:
#             A = []
#             for t in range(len(Formula)):
#                 a = 0
#                 true = 0
#                 index = groupindex(Output)
#                 iid = index.copy()
#                 for i in range(len(index) - 1):
#                     for j in range(i + 1, len(index)):
#                         v = []
#                         for m in range(len(index[i])):
#                             for n in range(len(index[j])):
#                                 sum_s = MG[index[i][m]].count(0) + MG[index[i][m]].count(3) + \
#                                         MG[index[j][n]].count(0) + MG[index[j][n]].count(3)
#                                 sum_v = MG[index[i][m]].count(1) + MG[index[j][n]].count(1)
#                                 # 求每个测试用例的ev es nv ns
#                                 ev_a = MG[index[i][m]].count(1)
#                                 es_a = MG[index[i][m]].count(0) + MG[index[i][m]].count(3)
#                                 ev_b = MG[index[j][n]].count(1)
#                                 es_b = MG[index[j][n]].count(0) + MG[index[j][n]].count(3)
#                                 if index[i][m] == 0:
#                                     if MG[0][index[j][n] - 1] == 1:
#                                         ev_b += 1
#                                     if MG[0][index[j][n] - 1] == 0 or MG[0][index[j][n] - 1] == 3:
#                                         es_b += 1
#                                 elif index[j][n] == 0:
#                                     if MG[0][index[i][m] - 1] == 1:
#                                         ev_a += 1
#                                     if MG[0][index[i][m] - 1] == 0 or MG[0][index[i][m] - 1] == 3:
#                                         es_a += 1
#                                 else:
#                                     if MG[0][index[j][n] - 1] == 1:
#                                         ev_b += 1
#                                         sum_v += 1
#                                     if MG[0][index[j][n] - 1] == 0 or MG[0][index[j][n] - 1] == 3:
#                                         es_b += 1
#                                         sum_s += 1
#                                     if MG[0][index[i][m] - 1] == 1:
#                                         ev_a += 1
#                                         sum_v += 1
#                                     if MG[0][index[i][m] - 1] == 0 or MG[0][index[i][m] - 1] == 3:
#                                         es_a += 1
#                                         sum_s += 1
#                                 nv_a = sum_v - ev_a
#                                 ns_a = sum_s - es_a
#                                 index_a = [ev_a, es_a, nv_a, ns_a]
#                                 nv_b = sum_v - ev_b
#                                 ns_b = sum_s - es_b
#                                 index_b = [ev_b, es_b, nv_b, ns_b]
#                                 sus_a = getSus(index_a, t)
#                                 sus_b = getSus(index_b, t)
#                                 if sus_a == sus_b:
#                                     v.append(random.choice([index[i], index[j]]))
#                                 elif sus_a > sus_b:
#                                     v.append(index[j])
#                                 else:
#                                     v.append(index[i])
#                         m = showmax(v)
#                         m = random.choice(m)
#                         if m == index[i]:
#                             if index[j] in iid:
#                                 iid.remove(index[j])
#                         else:
#                             if index[i] in iid:
#                                 iid.remove(index[i])
#                 if len(iid) == 0:
#                     a = 1
#                     ind = random.choice(index)
#                     if Output[1:][ind[0]] == Output[0]:
#                         true = 1
#                 else:
#                     if Output[1:][iid[0][0]] == Output[0]:
#                         true = 1
#                 flag.append(true)
#                 A.append(a)
#     return flag, A


def FaultTolerance14(MG, pf, Output):
    """
    group分组，failtim with random
    """
    Formula = ['Naish1', 'Naish2', 'Wong1', 'Russel&Rao', 'Binary', 'Jaccard', 'Anderberg', 'Sørensen-Dice', 'Dice',
               'Goodman', 'Tarantula', 'qe', 'CBI Inc.', 'Wong2', 'Hamann', 'Simple Matching', 'Sokal',
               'Rogers&Tanimoto', 'Hamming etc.', 'Euclid', 'Scott', 'Rogot1', 'Kulczynski2', 'Ochiai',
               'M2', 'AMPLE2', 'Wong3', 'Arithmetic Mean', 'Cohen', 'Fleiss']
    flag = []
    true = 0
    A = []
    for _ in range(len(Formula)):
        A.append(0)
    if 1 not in MG[0]:
        if pf[0] == 0:
            true = 1
        for _ in range(len(Formula)):
            flag.append(true)
        return flag, A
    else:
        out = Output[1:]
        result = out.count(out[0]) == len(out)
        if result:
            if out[0] == Output[0]:
                true = 1
            for _ in range(len(Formula)):
                flag.append(true)
            return flag, A
        else:
            A = []
            for t in range(len(Formula)):
                a = 0
                true = 0
                index = groupindex(Output)
                vote = copy.deepcopy(index)
                for i in range(len(index) - 1):
                    for j in range(i + 1, len(index)):
                        v = []
                        for m in range(len(index[i])):
                            for n in range(len(index[j])):
                                sum_s = MG[index[i][m]].count(0) + MG[index[i][m]].count(3) + \
                                        MG[index[j][n]].count(0) + MG[index[j][n]].count(3)
                                sum_v = MG[index[i][m]].count(1) + MG[index[j][n]].count(1)
                                # 求每个测试用例的ev es nv ns
                                ev_a = MG[index[i][m]].count(1)
                                es_a = MG[index[i][m]].count(0) + MG[index[i][m]].count(3)
                                ev_b = MG[index[j][n]].count(1)
                                es_b = MG[index[j][n]].count(0) + MG[index[j][n]].count(3)
                                if index[i][m] == 0:
                                    if MG[0][index[j][n] - 1] == 1:
                                        ev_b += 1
                                    if MG[0][index[j][n] - 1] == 0 or MG[0][index[j][n] - 1] == 3:
                                        es_b += 1
                                elif index[j][n] == 0:
                                    if MG[0][index[i][m] - 1] == 1:
                                        ev_a += 1
                                    if MG[0][index[i][m] - 1] == 0 or MG[0][index[i][m] - 1] == 3:
                                        es_a += 1
                                else:
                                    if MG[0][index[j][n] - 1] == 1:
                                        ev_b += 1
                                        sum_v += 1
                                    if MG[0][index[j][n] - 1] == 0 or MG[0][index[j][n] - 1] == 3:
                                        es_b += 1
                                        sum_s += 1
                                    if MG[0][index[i][m] - 1] == 1:
                                        ev_a += 1
                                        sum_v += 1
                                    if MG[0][index[i][m] - 1] == 0 or MG[0][index[i][m] - 1] == 3:
                                        es_a += 1
                                        sum_s += 1
                                nv_a = sum_v - ev_a
                                ns_a = sum_s - es_a
                                index_a = [ev_a, es_a, nv_a, ns_a]
                                nv_b = sum_v - ev_b
                                ns_b = sum_s - es_b
                                index_b = [ev_b, es_b, nv_b, ns_b]
                                sus_a = getSus(index_a, t)
                                sus_b = getSus(index_b, t)
                                if sus_a == sus_b:
                                    pass
                                elif sus_a > sus_b:
                                    v.append(index[j])
                                else:
                                    v.append(index[i])
                        if len(v) == 0:
                            # 两两case之间全部相同
                            v.append(index[j])
                            v.append(index[i])
                        m = showmax(v)
                        if len(m) == 1:
                            # 箭头
                            if m[0] == index[i]:
                                vote.append(index[j])
                            else:
                                vote.append(index[i])
                        else:
                            # 同票，都有箭头
                            vote.append(index[j])
                            vote.append(index[i])
                mi = showmin(vote)
                if len(mi) == 1:
                    if Output[1:][mi[0][0]] == Output[0]:
                        true = 1
                else:
                    a = 1
                    ind = random.choice(mi)
                    if Output[1:][ind[0]] == Output[0]:
                        true = 1
                flag.append(true)
                A.append(a)
    return flag, A


# def FaultTolerance15(MG, pf, Output):
#     """
#     group分组，failtim，根据TY改
#     相比14，取前两个
#     """
#     Formula = ['Naish1', 'Naish2', 'Wong1', 'Russel&Rao', 'Binary', 'Jaccard', 'Anderberg', 'Sørensen-Dice', 'Dice',
#                'Goodman', 'Tarantula', 'qe', 'CBI Inc.', 'Wong2', 'Hamann', 'Simple Matching', 'Sokal',
#                'Rogers&Tanimoto', 'Hamming etc.', 'Euclid', 'Scott', 'Rogot1', 'Kulczynski2', 'Ochiai',
#                'M2', 'AMPLE2', 'Wong3', 'Arithmetic Mean', 'Cohen', 'Fleiss']
#     flag = []
#     true = 0
#     A = []
#     for _ in range(len(Formula)):
#         A.append(0)
#     if 1 not in MG[0]:
#         if pf[0] == 0:
#             true = 1
#         for _ in range(len(Formula)):
#             flag.append(true)
#         return flag, A
#     else:
#         out = Output[1:]
#         result = out.count(out[0]) == len(out)
#         if result:
#             if out[0] == Output[0]:
#                 true = 1
#             for _ in range(len(Formula)):
#                 flag.append(true)
#             return flag, A
#         else:
#             A = []
#             for t in range(len(Formula)):
#                 a = 0
#                 true = 0
#                 index = groupindex(Output)
#                 vote = copy.deepcopy(index)
#                 for i in range(len(index) - 1):
#                     for j in range(i + 1, len(index)):
#                         v = []
#                         for m in range(len(index[i])):
#                             for n in range(len(index[j])):
#                                 sum_s = MG[index[i][m]].count(0) + MG[index[i][m]].count(3) + \
#                                         MG[index[j][n]].count(0) + MG[index[j][n]].count(3)
#                                 sum_v = MG[index[i][m]].count(1) + MG[index[j][n]].count(1)
#                                 # 求每个测试用例的ev es nv ns
#                                 ev_a = MG[index[i][m]].count(1)
#                                 es_a = MG[index[i][m]].count(0) + MG[index[i][m]].count(3)
#                                 ev_b = MG[index[j][n]].count(1)
#                                 es_b = MG[index[j][n]].count(0) + MG[index[j][n]].count(3)
#                                 if index[i][m] == 0:
#                                     if MG[0][index[j][n] - 1] == 1:
#                                         ev_b += 1
#                                     if MG[0][index[j][n] - 1] == 0 or MG[0][index[j][n] - 1] == 3:
#                                         es_b += 1
#                                 elif index[j][n] == 0:
#                                     if MG[0][index[i][m] - 1] == 1:
#                                         ev_a += 1
#                                     if MG[0][index[i][m] - 1] == 0 or MG[0][index[i][m] - 1] == 3:
#                                         es_a += 1
#                                 else:
#                                     if MG[0][index[j][n] - 1] == 1:
#                                         ev_b += 1
#                                         sum_v += 1
#                                     if MG[0][index[j][n] - 1] == 0 or MG[0][index[j][n] - 1] == 3:
#                                         es_b += 1
#                                         sum_s += 1
#                                     if MG[0][index[i][m] - 1] == 1:
#                                         ev_a += 1
#                                         sum_v += 1
#                                     if MG[0][index[i][m] - 1] == 0 or MG[0][index[i][m] - 1] == 3:
#                                         es_a += 1
#                                         sum_s += 1
#                                 nv_a = sum_v - ev_a
#                                 ns_a = sum_s - es_a
#                                 index_a = [ev_a, es_a, nv_a, ns_a]
#                                 nv_b = sum_v - ev_b
#                                 ns_b = sum_s - es_b
#                                 index_b = [ev_b, es_b, nv_b, ns_b]
#                                 sus_a = getSus(index_a, t)
#                                 sus_b = getSus(index_b, t)
#                                 if sus_a == sus_b:
#                                     pass
#                                 elif sus_a > sus_b:
#                                     v.append(index[j])
#                                 else:
#                                     v.append(index[i])
#                         if len(v) == 0:
#                             # 两两case之间全部相同
#                             v.append(index[j])
#                             v.append(index[i])
#                         m = showmax(v)
#                         if len(m) == 1:
#                             # 箭头
#                             if m[0] == index[i]:
#                                 vote.append(index[j])
#                             else:
#                                 vote.append(index[i])
#                         else:
#                             # 同票，都有箭头
#                             vote.append(index[j])
#                             vote.append(index[i])
#                 mi = showmin(vote)
#                 if len(mi) >= 2:
#                     a = 1
#                     ind = random.sample(mi, 2)
#                     if Output[1:][ind[0][0]] == Output[0] or Output[1:][ind[1][0]] == Output[0]:
#                         true = 1
#                 else:
#                     can = mi
#                     vote.remove(can[0])
#                     mi = showmin(vote)
#                     if len(mi) == 1:
#                         can.append(mi[0])
#                     else:
#                         a = 1
#                         ind = random.choice(mi)
#                         can.append(ind)
#                     if Output[1:][can[0][0]] == Output[0] or Output[1:][can[1][0]] == Output[0]:
#                         true = 1
#                 flag.append(true)
#                 A.append(a)
#     return flag, A
#
#
# def FaultTolerance16(MG, pf, Output):
#     """
#     group分组，failtim，根据TY改
#     相比14，取前两个，之后voting
#     结论：还不如failtim
#     """
#     Formula = ['Naish1', 'Naish2', 'Wong1', 'Russel&Rao', 'Binary', 'Jaccard', 'Anderberg', 'Sørensen-Dice', 'Dice',
#                'Goodman', 'Tarantula', 'qe', 'CBI Inc.', 'Wong2', 'Hamann', 'Simple Matching', 'Sokal',
#                'Rogers&Tanimoto', 'Hamming etc.', 'Euclid', 'Scott', 'Rogot1', 'Kulczynski2', 'Ochiai',
#                'M2', 'AMPLE2', 'Wong3', 'Arithmetic Mean', 'Cohen', 'Fleiss']
#     flag = []
#     true = 0
#     A = []
#     for _ in range(len(Formula)):
#         A.append(0)
#     if 1 not in MG[0]:
#         if pf[0] == 0:
#             true = 1
#         for _ in range(len(Formula)):
#             flag.append(true)
#         return flag, A
#     else:
#         out = Output[1:]
#         result = out.count(out[0]) == len(out)
#         if result:
#             if out[0] == Output[0]:
#                 true = 1
#             for _ in range(len(Formula)):
#                 flag.append(true)
#             return flag, A
#         else:
#             A = []
#             for t in range(len(Formula)):
#                 a = 0
#                 true = 0
#                 index = groupindex(Output)
#                 vote = copy.deepcopy(index)
#                 for i in range(len(index) - 1):
#                     for j in range(i + 1, len(index)):
#                         v = []
#                         for m in range(len(index[i])):
#                             for n in range(len(index[j])):
#                                 sum_s = MG[index[i][m]].count(0) + MG[index[i][m]].count(3) + \
#                                         MG[index[j][n]].count(0) + MG[index[j][n]].count(3)
#                                 sum_v = MG[index[i][m]].count(1) + MG[index[j][n]].count(1)
#                                 # 求每个测试用例的ev es nv ns
#                                 ev_a = MG[index[i][m]].count(1)
#                                 es_a = MG[index[i][m]].count(0) + MG[index[i][m]].count(3)
#                                 ev_b = MG[index[j][n]].count(1)
#                                 es_b = MG[index[j][n]].count(0) + MG[index[j][n]].count(3)
#                                 if index[i][m] == 0:
#                                     if MG[0][index[j][n] - 1] == 1:
#                                         ev_b += 1
#                                     if MG[0][index[j][n] - 1] == 0 or MG[0][index[j][n] - 1] == 3:
#                                         es_b += 1
#                                 elif index[j][n] == 0:
#                                     if MG[0][index[i][m] - 1] == 1:
#                                         ev_a += 1
#                                     if MG[0][index[i][m] - 1] == 0 or MG[0][index[i][m] - 1] == 3:
#                                         es_a += 1
#                                 else:
#                                     if MG[0][index[j][n] - 1] == 1:
#                                         ev_b += 1
#                                         sum_v += 1
#                                     if MG[0][index[j][n] - 1] == 0 or MG[0][index[j][n] - 1] == 3:
#                                         es_b += 1
#                                         sum_s += 1
#                                     if MG[0][index[i][m] - 1] == 1:
#                                         ev_a += 1
#                                         sum_v += 1
#                                     if MG[0][index[i][m] - 1] == 0 or MG[0][index[i][m] - 1] == 3:
#                                         es_a += 1
#                                         sum_s += 1
#                                 nv_a = sum_v - ev_a
#                                 ns_a = sum_s - es_a
#                                 index_a = [ev_a, es_a, nv_a, ns_a]
#                                 nv_b = sum_v - ev_b
#                                 ns_b = sum_s - es_b
#                                 index_b = [ev_b, es_b, nv_b, ns_b]
#                                 sus_a = getSus(index_a, t)
#                                 sus_b = getSus(index_b, t)
#                                 if sus_a == sus_b:
#                                     pass
#                                 elif sus_a > sus_b:
#                                     v.append(index[j])
#                                 else:
#                                     v.append(index[i])
#                         if len(v) == 0:
#                             # 两两case之间全部相同
#                             v.append(index[j])
#                             v.append(index[i])
#                         m = showmax(v)
#                         if len(m) == 1:
#                             # 箭头
#                             if m[0] == index[i]:
#                                 vote.append(index[j])
#                             else:
#                                 vote.append(index[i])
#                         else:
#                             # 同票，都有箭头
#                             vote.append(index[j])
#                             vote.append(index[i])
#                 mi = showmin(vote)
#                 if len(mi) >= 2:
#                     a = 1
#                     ind = random.sample(mi, 2)
#                     if len(ind[0]) > len(ind[1]):
#                         v = ind[0]
#                     elif len(ind[0]) < len(ind[1]):
#                         v = ind[1]
#                     else:
#                         v = random.choice(ind)
#                     if Output[1:][v[0]] == Output[0]:
#                         true = 1
#                 else:
#                     can = mi
#                     vote.remove(can[0])
#                     mi = showmin(vote)
#                     if len(mi) == 1:
#                         can.append(mi[0])
#                     else:
#                         a = 1
#                         ind = random.choice(mi)
#                         can.append(ind)
#                     if len(can[0]) > len(can[1]):
#                         v = can[0]
#                     elif len(can[0]) < len(can[1]):
#                         v = can[1]
#                     else:
#                         v = random.choice(can)
#                     if Output[1:][v[0]] == Output[0]:
#                         true = 1
#                 flag.append(true)
#                 A.append(a)
#     return flag, A


def FaultTolerance17(MG, pf, Output):
    """
    group分组，failtim with voting
    """
    Formula = ['Naish1', 'Naish2', 'Wong1', 'Russel&Rao', 'Binary', 'Jaccard', 'Anderberg', 'Sørensen-Dice', 'Dice',
               'Goodman', 'Tarantula', 'qe', 'CBI Inc.', 'Wong2', 'Hamann', 'Simple Matching', 'Sokal',
               'Rogers&Tanimoto', 'Hamming etc.', 'Euclid', 'Scott', 'Rogot1', 'Kulczynski2', 'Ochiai',
               'M2', 'AMPLE2', 'Wong3', 'Arithmetic Mean', 'Cohen', 'Fleiss']
    flag = []
    true = 0
    A = []
    for _ in range(len(Formula)):
        A.append(0)
    if 1 not in MG[0]:
        if pf[0] == 0:
            true = 1
        for _ in range(len(Formula)):
            flag.append(true)
        return flag, A
    else:
        out = Output[1:]
        result = out.count(out[0]) == len(out)
        if result:
            if out[0] == Output[0]:
                true = 1
            for _ in range(len(Formula)):
                flag.append(true)
            return flag, A
        else:
            A = []
            for t in range(len(Formula)):
                a = 0
                true = 0
                index = groupindex(Output)
                vote = copy.deepcopy(index)
                for i in range(len(index) - 1):
                    for j in range(i + 1, len(index)):
                        v = []
                        for m in range(len(index[i])):
                            for n in range(len(index[j])):
                                sum_s = MG[index[i][m]].count(0) + MG[index[i][m]].count(3) + \
                                        MG[index[j][n]].count(0) + MG[index[j][n]].count(3)
                                sum_v = MG[index[i][m]].count(1) + MG[index[j][n]].count(1)
                                # 求每个测试用例的ev es nv ns
                                ev_a = MG[index[i][m]].count(1)
                                es_a = MG[index[i][m]].count(0) + MG[index[i][m]].count(3)
                                ev_b = MG[index[j][n]].count(1)
                                es_b = MG[index[j][n]].count(0) + MG[index[j][n]].count(3)
                                if index[i][m] == 0:
                                    if MG[0][index[j][n] - 1] == 1:
                                        ev_b += 1
                                    if MG[0][index[j][n] - 1] == 0 or MG[0][index[j][n] - 1] == 3:
                                        es_b += 1
                                elif index[j][n] == 0:
                                    if MG[0][index[i][m] - 1] == 1:
                                        ev_a += 1
                                    if MG[0][index[i][m] - 1] == 0 or MG[0][index[i][m] - 1] == 3:
                                        es_a += 1
                                else:
                                    if MG[0][index[j][n] - 1] == 1:
                                        ev_b += 1
                                        sum_v += 1
                                    if MG[0][index[j][n] - 1] == 0 or MG[0][index[j][n] - 1] == 3:
                                        es_b += 1
                                        sum_s += 1
                                    if MG[0][index[i][m] - 1] == 1:
                                        ev_a += 1
                                        sum_v += 1
                                    if MG[0][index[i][m] - 1] == 0 or MG[0][index[i][m] - 1] == 3:
                                        es_a += 1
                                        sum_s += 1
                                nv_a = sum_v - ev_a
                                ns_a = sum_s - es_a
                                index_a = [ev_a, es_a, nv_a, ns_a]
                                nv_b = sum_v - ev_b
                                ns_b = sum_s - es_b
                                index_b = [ev_b, es_b, nv_b, ns_b]
                                sus_a = getSus(index_a, t)
                                sus_b = getSus(index_b, t)
                                if sus_a == sus_b:
                                    pass
                                elif sus_a > sus_b:
                                    v.append(index[j])
                                else:
                                    v.append(index[i])
                        if len(v) == 0:
                            # 两两case之间全部相同
                            v.append(index[j])
                            v.append(index[i])
                        m = showmax(v)
                        if len(m) == 1:
                            # 箭头
                            if m[0] == index[i]:
                                vote.append(index[j])
                            else:
                                vote.append(index[i])
                        else:
                            # 同票，都有箭头
                            vote.append(index[j])
                            vote.append(index[i])
                mi = showmin(vote)
                if len(mi) == 1:
                    if Output[1:][mi[0][0]] == Output[0]:
                        true = 1
                else:
                    can = showmaxsize(mi)
                    if len(can) == 1:
                        if Output[1:][can[0][0]] == Output[0]:
                            true = 1
                    else:
                        a = 1
                        ind = random.choice(can)
                        if Output[1:][ind[0]] == Output[0]:
                            true = 1
                flag.append(true)
                A.append(a)
    return flag, A


# def FaultTolerance18(MG, pf, Output):
#     """
#     group分组，failtim，根据TY改
#     相比14，取前两个
#     """
#     Formula = ['Naish1', 'Naish2', 'Wong1', 'Russel&Rao', 'Binary', 'Jaccard', 'Anderberg', 'Sørensen-Dice', 'Dice',
#                'Goodman', 'Tarantula', 'qe', 'CBI Inc.', 'Wong2', 'Hamann', 'Simple Matching', 'Sokal',
#                'Rogers&Tanimoto', 'Hamming etc.', 'Euclid', 'Scott', 'Rogot1', 'Kulczynski2', 'Ochiai',
#                'M2', 'AMPLE2', 'Wong3', 'Arithmetic Mean', 'Cohen', 'Fleiss']
#     flag = []
#     true = 0
#     A = []
#     for _ in range(len(Formula)):
#         A.append(0)
#     if 1 not in MG[0]:
#         if pf[0] == 0:
#             true = 1
#         for _ in range(len(Formula)):
#             flag.append(true)
#         return flag, A
#     else:
#         out = Output[1:]
#         result = out.count(out[0]) == len(out)
#         if result:
#             if out[0] == Output[0]:
#                 true = 1
#             for _ in range(len(Formula)):
#                 flag.append(true)
#             return flag, A
#         else:
#             A = []
#             for t in range(len(Formula)):
#                 a = 0
#                 true = 0
#                 index = groupindex(Output)
#                 vote = copy.deepcopy(index)
#                 for i in range(len(index) - 1):
#                     for j in range(i + 1, len(index)):
#                         v = []
#                         for m in range(len(index[i])):
#                             for n in range(len(index[j])):
#                                 sum_s = MG[index[i][m]].count(0) + MG[index[i][m]].count(3) + \
#                                         MG[index[j][n]].count(0) + MG[index[j][n]].count(3)
#                                 sum_v = MG[index[i][m]].count(1) + MG[index[j][n]].count(1)
#                                 # 求每个测试用例的ev es nv ns
#                                 ev_a = MG[index[i][m]].count(1)
#                                 es_a = MG[index[i][m]].count(0) + MG[index[i][m]].count(3)
#                                 ev_b = MG[index[j][n]].count(1)
#                                 es_b = MG[index[j][n]].count(0) + MG[index[j][n]].count(3)
#                                 if index[i][m] == 0:
#                                     if MG[0][index[j][n] - 1] == 1:
#                                         ev_b += 1
#                                     if MG[0][index[j][n] - 1] == 0 or MG[0][index[j][n] - 1] == 3:
#                                         es_b += 1
#                                 elif index[j][n] == 0:
#                                     if MG[0][index[i][m] - 1] == 1:
#                                         ev_a += 1
#                                     if MG[0][index[i][m] - 1] == 0 or MG[0][index[i][m] - 1] == 3:
#                                         es_a += 1
#                                 else:
#                                     if MG[0][index[j][n] - 1] == 1:
#                                         ev_b += 1
#                                         sum_v += 1
#                                     if MG[0][index[j][n] - 1] == 0 or MG[0][index[j][n] - 1] == 3:
#                                         es_b += 1
#                                         sum_s += 1
#                                     if MG[0][index[i][m] - 1] == 1:
#                                         ev_a += 1
#                                         sum_v += 1
#                                     if MG[0][index[i][m] - 1] == 0 or MG[0][index[i][m] - 1] == 3:
#                                         es_a += 1
#                                         sum_s += 1
#                                 nv_a = sum_v - ev_a
#                                 ns_a = sum_s - es_a
#                                 index_a = [ev_a, es_a, nv_a, ns_a]
#                                 nv_b = sum_v - ev_b
#                                 ns_b = sum_s - es_b
#                                 index_b = [ev_b, es_b, nv_b, ns_b]
#                                 sus_a = getSus(index_a, t)
#                                 sus_b = getSus(index_b, t)
#                                 if sus_a == sus_b:
#                                     pass
#                                 elif sus_a > sus_b:
#                                     v.append(index[j])
#                                 else:
#                                     v.append(index[i])
#                         if len(v) == 0:
#                             # 两两case之间全部相同
#                             v.append(index[j])
#                             v.append(index[i])
#                         m = showmax(v)
#                         if len(m) == 1:
#                             # 箭头
#                             if m[0] == index[i]:
#                                 vote.append(index[j])
#                             else:
#                                 vote.append(index[i])
#                         else:
#                             # 同票，都有箭头
#                             vote.append(index[j])
#                             vote.append(index[i])
#                 mi = showmin(vote)
#                 if len(mi) >= 2:
#                     a = 1
#                     ind = random.sample(mi, 2)
#                     if Output[1:][ind[0][0]] == Output[0] or Output[1:][ind[1][0]] == Output[0]:
#                         true = 1
#                 else:
#                     can = mi
#                     vote.remove(can[0])
#                     mi = showmin(vote)
#                     if len(mi) == 1:
#                         can.append(mi[0])
#                     else:
#                         a = 1
#                         ind = random.choice(mi)
#                         can.append(ind)
#                     if Output[1:][can[0][0]] == Output[0] or Output[1:][can[1][0]] == Output[0]:
#                         true = 1
#                 flag.append(true)
#                 A.append(a)
#     return flag, A


def FaultTolerance19(MG, pf, Output):
    """
    group分组，failtim with voting
    去掉fs
    """
    Formula = ['Naish1', 'Naish2', 'Wong1', 'Russel&Rao', 'Binary', 'Jaccard', 'Anderberg', 'Sørensen-Dice', 'Dice',
               'Goodman', 'Tarantula', 'qe', 'CBI Inc.', 'Wong2', 'Hamann', 'Simple Matching', 'Sokal',
               'Rogers&Tanimoto', 'Hamming etc.', 'Euclid', 'Scott', 'Rogot1', 'Kulczynski2', 'Ochiai',
               'M2', 'AMPLE2', 'Wong3', 'Arithmetic Mean', 'Cohen', 'Fleiss', 'New']
    flag = []
    true = 0
    A = []
    for _ in range(len(Formula)):
        A.append(0)
    if 1 not in MG[0]:
        if pf[0] == 0:
            true = 1
        for _ in range(len(Formula)):
            flag.append(true)
        return flag, A
    else:
        out = Output[1:]
        result = out.count(out[0]) == len(out)
        if result:
            if out[0] == Output[0]:
                true = 1
            for _ in range(len(Formula)):
                flag.append(true)
            return flag, A
        else:
            A = []
            for t in range(len(Formula)):
                a = 0
                true = 0
                index = groupindex(Output)
                vote = copy.deepcopy(index)
                for i in range(len(index) - 1):
                    for j in range(i + 1, len(index)):
                        v = []
                        for m in range(len(index[i])):
                            for n in range(len(index[j])):
                                sum_s = MG[index[i][m]].count(0) + \
                                        MG[index[j][n]].count(0)
                                sum_v = MG[index[i][m]].count(1) + MG[index[j][n]].count(1)
                                # 求每个测试用例的ev es nv ns
                                ev_a = MG[index[i][m]].count(1)
                                es_a = MG[index[i][m]].count(0)
                                ev_b = MG[index[j][n]].count(1)
                                es_b = MG[index[j][n]].count(0)
                                if index[i][m] == 0:
                                    if MG[0][index[j][n] - 1] == 1:
                                        ev_b += 1
                                    if MG[0][index[j][n] - 1] == 0:
                                        es_b += 1
                                elif index[j][n] == 0:
                                    if MG[0][index[i][m] - 1] == 1:
                                        ev_a += 1
                                    if MG[0][index[i][m] - 1] == 0:
                                        es_a += 1
                                else:
                                    if MG[0][index[j][n] - 1] == 1:
                                        ev_b += 1
                                        sum_v += 1
                                    if MG[0][index[j][n] - 1] == 0:
                                        es_b += 1
                                        sum_s += 1
                                    if MG[0][index[i][m] - 1] == 1:
                                        ev_a += 1
                                        sum_v += 1
                                    if MG[0][index[i][m] - 1] == 0:
                                        es_a += 1
                                        sum_s += 1
                                nv_a = sum_v - ev_a
                                ns_a = sum_s - es_a
                                index_a = [ev_a, es_a, nv_a, ns_a]
                                nv_b = sum_v - ev_b
                                ns_b = sum_s - es_b
                                index_b = [ev_b, es_b, nv_b, ns_b]
                                sus_a = getSus(index_a, t)
                                sus_b = getSus(index_b, t)
                                if sus_a == sus_b:
                                    pass
                                elif sus_a > sus_b:
                                    v.append(index[j])
                                else:
                                    v.append(index[i])
                        if len(v) == 0:
                            # 两两case之间全部相同
                            v.append(index[j])
                            v.append(index[i])
                        m = showmax(v)
                        if len(m) == 1:
                            # 箭头
                            if m[0] == index[i]:
                                vote.append(index[j])
                            else:
                                vote.append(index[i])
                        else:
                            # 同票，都有箭头
                            vote.append(index[j])
                            vote.append(index[i])
                mi = showmin(vote)
                if len(mi) == 1:
                    if Output[1:][mi[0][0]] == Output[0]:
                        true = 1
                else:
                    can = showmaxsize(mi)
                    if len(can) == 1:
                        if Output[1:][can[0][0]] == Output[0]:
                            true = 1
                    else:
                        a = 1
                        ind = random.choice(can)
                        if Output[1:][ind[0]] == Output[0]:
                            true = 1
                flag.append(true)
                A.append(a)
    return flag, A


def FaultTolerance20(MG, pf, Output):
    """
    group分组，failtim with random
    去掉fs
    """
    Formula = ['Naish1', 'Naish2', 'Wong1', 'Russel&Rao', 'Binary', 'Jaccard', 'Anderberg', 'Sørensen-Dice', 'Dice',
               'Goodman', 'Tarantula', 'qe', 'CBI Inc.', 'Wong2', 'Hamann', 'Simple Matching', 'Sokal',
               'Rogers&Tanimoto', 'Hamming etc.', 'Euclid', 'Scott', 'Rogot1', 'Kulczynski2', 'Ochiai',
               'M2', 'AMPLE2', 'Wong3', 'Arithmetic Mean', 'Cohen', 'Fleiss']
    flag = []
    true = 0
    A = []
    for _ in range(len(Formula)):
        A.append(0)
    if 1 not in MG[0]:
        if pf[0] == 0:
            true = 1
        for _ in range(len(Formula)):
            flag.append(true)
        return flag, A
    else:
        out = Output[1:]
        result = out.count(out[0]) == len(out)
        if result:
            if out[0] == Output[0]:
                true = 1
            for _ in range(len(Formula)):
                flag.append(true)
            return flag, A
        else:
            A = []
            for t in range(len(Formula)):
                a = 0
                true = 0
                index = groupindex(Output)
                vote = copy.deepcopy(index)
                for i in range(len(index) - 1):
                    for j in range(i + 1, len(index)):
                        v = []
                        for m in range(len(index[i])):
                            for n in range(len(index[j])):
                                sum_s = MG[index[i][m]].count(0) + \
                                        MG[index[j][n]].count(0)
                                sum_v = MG[index[i][m]].count(1) + MG[index[j][n]].count(1)
                                # 求每个测试用例的ev es nv ns
                                ev_a = MG[index[i][m]].count(1)
                                es_a = MG[index[i][m]].count(0)
                                ev_b = MG[index[j][n]].count(1)
                                es_b = MG[index[j][n]].count(0)
                                if index[i][m] == 0:
                                    if MG[0][index[j][n] - 1] == 1:
                                        ev_b += 1
                                    if MG[0][index[j][n] - 1] == 0:
                                        es_b += 1
                                elif index[j][n] == 0:
                                    if MG[0][index[i][m] - 1] == 1:
                                        ev_a += 1
                                    if MG[0][index[i][m] - 1] == 0:
                                        es_a += 1
                                else:
                                    if MG[0][index[j][n] - 1] == 1:
                                        ev_b += 1
                                        sum_v += 1
                                    if MG[0][index[j][n] - 1] == 0:
                                        es_b += 1
                                        sum_s += 1
                                    if MG[0][index[i][m] - 1] == 1:
                                        ev_a += 1
                                        sum_v += 1
                                    if MG[0][index[i][m] - 1] == 0:
                                        es_a += 1
                                        sum_s += 1
                                nv_a = sum_v - ev_a
                                ns_a = sum_s - es_a
                                index_a = [ev_a, es_a, nv_a, ns_a]
                                nv_b = sum_v - ev_b
                                ns_b = sum_s - es_b
                                index_b = [ev_b, es_b, nv_b, ns_b]
                                sus_a = getSus(index_a, t)
                                sus_b = getSus(index_b, t)
                                if sus_a == sus_b:
                                    pass
                                elif sus_a > sus_b:
                                    v.append(index[j])
                                else:
                                    v.append(index[i])
                        if len(v) == 0:
                            # 两两case之间全部相同
                            v.append(index[j])
                            v.append(index[i])
                        m = showmax(v)
                        if len(m) == 1:
                            # 箭头
                            if m[0] == index[i]:
                                vote.append(index[j])
                            else:
                                vote.append(index[i])
                        else:
                            # 同票，都有箭头
                            vote.append(index[j])
                            vote.append(index[i])
                mi = showmin(vote)
                if len(mi) == 1:
                    if Output[1:][mi[0][0]] == Output[0]:
                        true = 1
                else:
                    a = 1
                    ind = random.choice(mi)
                    if Output[1:][ind[0]] == Output[0]:
                        true = 1
                flag.append(true)
                A.append(a)
    return flag, A


def FaultTolerance21(MG, pf, Output):
    """
    FAILTIM取最小可疑度
    """
    Formula = ['Naish1', 'Naish2', 'Wong1', 'Russel&Rao', 'Binary', 'Jaccard', 'Anderberg', 'Sørensen-Dice', 'Dice',
               'Goodman', 'Tarantula', 'qe', 'CBI Inc.', 'Wong2', 'Hamann', 'Simple Matching', 'Sokal',
               'Rogers&Tanimoto', 'Hamming etc.', 'Euclid', 'Scott', 'Rogot1', 'Kulczynski2', 'Ochiai',
               'M2', 'AMPLE2', 'Wong3', 'Arithmetic Mean', 'Cohen', 'Fleiss', 'New']
    flag = []
    A = []
    for _ in range(len(Formula)):
        A.append(0)
    for t in range(len(Formula)):
        true = 0
        output = []
        for i in range(1):  # 1
            for j in range(len(MG[i])):
                if MG[i][j] == 1:  # violated
                    # 根据包含的测试用例确定 MG set
                    sum_s = MG[i].count(0)  # + MG[i].count(3)
                    sum_v = MG[i].count(1)  # 注意
                    sum_s += MG[j + 1].count(0)  # + MG[j + 1].count(3)
                    sum_v += MG[j + 1].count(1)
                    # 求每个测试用例的ev es nv ns
                    ev_a = MG[i].count(1)
                    es_a = MG[i].count(0)  # + MG[i].count(3)
                    nv_a = sum_v - ev_a
                    ns_a = sum_s - es_a
                    index_a = [ev_a, es_a, nv_a, ns_a]
                    ev_b = MG[j + 1].count(1) + 1
                    es_b = MG[j + 1].count(0)  # + MG[j + 1].count(3)
                    nv_b = sum_v - ev_b
                    ns_b = sum_s - es_b
                    index_b = [ev_b, es_b, nv_b, ns_b]
                    # 求测试用例的可疑度
                    sus_a = getSus(index_a, t)
                    sus_b = getSus(index_b, t)
                    if sus_a > sus_b:
                        output.append(Output[j+2])
                    elif sus_a == sus_b:
                        index = random.choice([1, j+2])
                        output.append(Output[index])
                    else:
                        output.append(Output[1])
        if len(output) == 0:
            out = Output[1]
        else:
            out = random.choice(output)
        if out == Output[0]:
            true = 1
        flag.append(true)
    return flag, A


def group(Output):
    """返回group的size"""
    output = Output[1:]
    count = []
    s = []
    [s.append(i) for i in output if not i in s]
    for i in s:
        count.append(output.count(i))
    return count


def groupindex(Output):
    """返回group的index"""
    output = Output[1:]
    ind = []
    s = []
    [s.append(i) for i in output if not i in s]
    for i in s:
        index_list = []
        index = -1
        count = output.count(i)
        # 通过list.index()方法的__start参数，指定起始索引
        for _ in range(0, count):
            index = output.index(i, index + 1)
            index_list.append(index)
        ind.append(index_list)
    return ind


# def group2(Output, MG, EMR):
#     """返回group的size"""
#     output = [Output[1:][0]]
#     for j in range(len(MG[0])):
#         if j not in EMR:
#             continue
#         output.append(Output[1:][j + 1])
#     count = []
#     s = []
#     [s.append(i) for i in output if not i in s]
#     for i in s:
#         count.append(output.count(i))
#     return count


# def getMetrics_v1(row, ws, mu, MG, pf, Output):
#     """
#     根据TY改, 全部MR
#     """
#     Formula = ['Naish1', 'Naish2', 'Wong1', 'Russel&Rao', 'Binary', 'Jaccard', 'Anderberg', 'Sørensen-Dice', 'Dice',
#                'Goodman', 'Tarantula', 'qe', 'CBI Inc.', 'Wong2', 'Hamann', 'Simple Matching', 'Sokal',
#                'Rogers&Tanimoto', 'Hamming etc.', 'Euclid', 'Scott', 'Rogot1', 'Kulczynski2', 'Ochiai',
#                'M2', 'AMPLE2', 'Wong3', 'Arithmetic Mean', 'Cohen', 'Fleiss']
#
#     # 统计指标
#     datadist = {}
#     tablelist = {"Mutant" + str(mu): ['identify(%)', 'failed case(%)', 'VMG(%)', 'SMG(%)',
#                                       'false satisfied MG(%)', 'FFSF', 'FFST', 'All Failed(%)',
#                                       'failed case2(%)', 'num', 'var']}
#     datadist.update(tablelist)
#
#     FFSlist = []
#     FFSF = []
#     FFST = []
#     failedcase = 0
#     failedcase2 = 0
#     VMGlist = []
#     SMGlist = []
#     allF = 0
#     tt = len(MG[0][0]) + 1
#     Flag = []
#     count = []
#     var = []
#     for k in range(len(MG)):
#         S_MG = 0
#         V_MG = 0
#         FFS = 0
#         for i in range(len(MG[k])):
#             S_MG += (MG[k][i].count(0) + MG[k][i].count(3))
#             FFS += MG[k][i].count(3)
#             V_MG += MG[k][i].count(1)
#         if S_MG == 0:
#             percent_FFS = 0
#         else:
#             percent_FFS = FFS / S_MG
#
#         # if percent_FFS > 0.3:
#         #     continue
#
#         # 如果全部failed，则标记
#         num = pf[k][0:tt].count(1)
#         if tt == num:
#             allF += 1
#         if pf[k][0] == 1:
#             failedcase += 1
#         failedcase2 += pf[k][0:tt].count(1)
#
#         FFSlist.append(percent_FFS)
#         VMGlist.append(V_MG / (S_MG + V_MG))
#         SMGlist.append(S_MG / (S_MG + V_MG))
#         flag = FaultTolerance3(MG[k], pf[k], Output[k])
#         Flag.append(flag)
#         c = group(Output[k])
#         if len(c) == 1:
#             c.append(0)
#             var.append(np.var(c))
#         else:
#             var.append(np.var(c))
#         if 0 in c:
#             c.remove(0)
#         count.append(len(c))
#     for t in range(len(Flag[0])):
#         true = 0
#         ffsf = []
#         ffst = []
#         a = []
#         for i in range(len(Flag)):
#             true += Flag[i][t]
#             if Flag[i][t] == 0:
#                 ffsf.append(FFSlist[i])
#             else:
#                 ffst.append(FFSlist[i])
#         if len(ffsf) == 0:
#             ffsf = 0
#         if len(ffst) == 0:
#             ffst = 0
#         FFSF.append(ffsf)
#         FFST.append(ffst)
#         percent_identify = true / len(Flag)
#         percent_allF = allF / len(Flag)
#         percent_failed = failedcase / len(Flag)
#         percent_failed2 = failedcase2 / (len(Flag) * tt)
#         value = [round(percent_identify * 100, 2), round(percent_failed * 100, 2), round(np.mean(VMGlist) * 100, 2),
#                  round(np.mean(SMGlist) * 100, 2), round(np.mean(FFSlist) * 100, 2), round(np.mean(ffsf) * 100, 2),
#                  round(np.mean(ffst) * 100, 2), round(percent_allF * 100, 2), round(percent_failed2 * 100, 2),
#                  round(float(np.mean(count)), 2), round(float(np.mean(var)), 2)]
#         data = {
#             Formula[t]: value
#         }
#         datadist.update(data)
#         if failedcase == 0:
#             return row
#
#     for i, j in datadist.items():  # i--公式名称, j--指标值
#         ws.cell(row, 1).value = i  # 添加第 1 列的数据
#         for col in range(2, len(j) + 2):  # values列表中索引
#             ws.cell(row, col).value = j[col - 2]
#         row += 1  # 行数
#     row += 2  # 行数
#     return row
#
#
# def getMetrics_v2(row, ws, mu, MG, pf, Output, EMR):
#     """
#     去掉不等的MR
#     """
#     Formula = ['Naish1', 'Naish2', 'Wong1', 'Russel&Rao', 'Binary', 'Jaccard', 'Anderberg', 'Sørensen-Dice', 'Dice',
#                'Goodman', 'Tarantula', 'qe', 'CBI Inc.', 'Wong2', 'Hamann', 'Simple Matching', 'Sokal',
#                'Rogers&Tanimoto', 'Hamming etc.', 'Euclid', 'Scott', 'Rogot1', 'Kulczynski2', 'Ochiai',
#                'M2', 'AMPLE2', 'Wong3', 'Arithmetic Mean', 'Cohen', 'Fleiss']
#
#     # 统计指标
#     datadist = {}
#     tablelist = {"Mutant" + str(mu): ['identify(%)', 'failed case(%)', 'VMG(%)', 'SMG(%)',
#                                       'false satisfied MG(%)', 'FFSF', 'FFST', 'All Failed(%)',
#                                       'failed case2(%)', 'num', 'var']}
#     datadist.update(tablelist)
#     FFSlist = []
#     FFSF = []
#     FFST = []
#     failedcase = 0
#     failedcase2 = 0
#     VMGlist = []
#     SMGlist = []
#     allF = 0
#     tt1 = len(EMR) + 1
#     Flag = []
#     var = []
#     count = []
#     for k in range(len(MG)):
#         S_MG = 0
#         V_MG = 0
#         FFS = 0
#         for i in range(len(MG[k])):
#             S_MG += (MG[k][i].count(0) + MG[k][i].count(3))
#             FFS += MG[k][i].count(3)
#             V_MG += MG[k][i].count(1)
#         if S_MG == 0:
#             percent_FFS = 0
#         else:
#             percent_FFS = FFS / S_MG
#
#         # if percent_FFS > 0.3:
#         #     continue
#
#         # 如果全部failed，则标记
#         if pf[k][0] == 1:
#             num = 1
#             failedcase += 1
#             failedcase2 += 1
#         else:
#             num = 0
#         for i in EMR:
#             if pf[k][i + 1] == 1:
#                 num += 1
#                 failedcase2 += 1
#         if tt1 == num:
#             allF += 1
#
#         FFSlist.append(percent_FFS)
#         VMGlist.append(V_MG / (S_MG + V_MG))
#         SMGlist.append(S_MG / (S_MG + V_MG))
#         flag = FaultTolerance4(MG[k], pf[k], Output[k], EMR)
#         Flag.append(flag)
#         c = group2(Output[k], MG[k], EMR)
#         if len(c) == 1:
#             c.append(0)
#             var.append(np.var(c))
#         else:
#             var.append(np.var(c))
#         if 0 in c:
#             c.remove(0)
#         count.append(len(c))
#     for t in range(len(Flag[0])):
#         true = 0
#         ffsf = []
#         ffst = []
#         a = []
#         for i in range(len(Flag)):
#             true += Flag[i][t]
#             if Flag[i][t] == 0:
#                 ffsf.append(FFSlist[i])
#             else:
#                 ffst.append(FFSlist[i])
#         if len(ffsf) == 0:
#             ffsf = 0
#         if len(ffst) == 0:
#             ffst = 0
#         FFSF.append(ffsf)
#         FFST.append(ffst)
#         percent_identify = true / len(Flag)
#         percent_allF = allF / len(Flag)
#         percent_failed = failedcase / len(Flag)
#         percent_failed2 = failedcase2 / (len(Flag) * tt1)
#         value = [round(percent_identify * 100, 2), round(percent_failed * 100, 2), round(np.mean(VMGlist) * 100, 2),
#                  round(np.mean(SMGlist) * 100, 2), round(np.mean(FFSlist) * 100, 2), round(np.mean(ffsf) * 100, 2),
#                  round(np.mean(ffst) * 100, 2), round(percent_allF * 100, 2), round(percent_failed2 * 100, 2),
#                  round(float(np.mean(count)), 2), round(float(np.mean(var)), 2)]
#         data = {
#             Formula[t]: value
#         }
#         datadist.update(data)
#         if failedcase == 0:
#             return row
#
#     for i, j in datadist.items():  # i--公式名称, j--指标值
#         ws.cell(row, 1).value = i  # 添加第 1 列的数据
#         for col in range(2, len(j) + 2):  # values列表中索引
#             ws.cell(row, col).value = j[col - 2]
#         row += 1  # 行数
#     row += 2  # 行数
#     return row
#
#
# def getMetrics_v3(row, ws, mu, MG, pf, Output):
#     """
#     根据TY改, 全部MR
#     """
#     Formula = ['Naish1', 'Naish2', 'Wong1', 'Russel&Rao', 'Binary', 'Jaccard', 'Anderberg', 'Sørensen-Dice', 'Dice',
#                'Goodman', 'Tarantula', 'qe', 'CBI Inc.', 'Wong2', 'Hamann', 'Simple Matching', 'Sokal',
#                'Rogers&Tanimoto', 'Hamming etc.', 'Euclid', 'Scott', 'Rogot1', 'Kulczynski2', 'Ochiai',
#                'M2', 'AMPLE2', 'Wong3', 'Arithmetic Mean', 'Cohen', 'Fleiss']
#
#     # 统计指标
#     datadist = {}
#     tablelist = {"Mutant" + str(mu): ['identify(%)', 'failed case(%)', 'VMG(%)', 'SMG(%)',
#                                       'false satisfied MG(%)', 'FFSF', 'FFST', 'All Failed(%)',
#                                       'failed case2(%)']}
#     datadist.update(tablelist)
#
#     FFSlist = []
#     FFSF = []
#     FFST = []
#     failedcase = 0
#     failedcase2 = 0
#     VMGlist = []
#     SMGlist = []
#     allF = 0
#     tt = len(MG[0][0]) + 1
#     Flag = []
#     for k in range(len(MG)):
#         S_MG = 0
#         V_MG = 0
#         FFS = 0
#         for i in range(len(MG[k])):
#             S_MG += (MG[k][i].count(0) + MG[k][i].count(3))
#             FFS += MG[k][i].count(3)
#             V_MG += MG[k][i].count(1)
#         if S_MG == 0:
#             percent_FFS = 0
#         else:
#             percent_FFS = FFS / S_MG
#
#         # if percent_FFS > 0.3:
#         #     continue
#
#         # 如果全部failed，则标记
#         num = pf[k][0:tt].count(1)
#         if tt == num:
#             allF += 1
#
#         if pf[k][0] == 1:
#             failedcase += 1
#         failedcase2 += pf[k][0:tt].count(1)
#
#         FFSlist.append(percent_FFS)
#         VMGlist.append(V_MG / (S_MG + V_MG))
#         SMGlist.append(S_MG / (S_MG + V_MG))
#         flag = FaultTolerance6(MG[k], pf[k], Output[k])
#         Flag.append(flag)
#     for t in range(len(Flag[0])):
#         true = 0
#         ffsf = []
#         ffst = []
#         a = []
#         for i in range(len(Flag)):
#             true += Flag[i][t]
#             if Flag[i][t] == 0:
#                 ffsf.append(FFSlist[i])
#             else:
#                 ffst.append(FFSlist[i])
#         if len(ffsf) == 0:
#             ffsf = 0
#         if len(ffst) == 0:
#             ffst = 0
#         FFSF.append(ffsf)
#         FFST.append(ffst)
#         percent_identify = true / len(Flag)
#         percent_allF = allF / len(Flag)
#         percent_failed = failedcase / len(Flag)
#         percent_failed2 = failedcase2 / (len(Flag) * tt)
#         value = [round(percent_identify * 100, 2), round(percent_failed * 100, 2), round(np.mean(VMGlist) * 100, 2),
#                  round(np.mean(SMGlist) * 100, 2), round(np.mean(FFSlist) * 100, 2), round(np.mean(ffsf) * 100, 2),
#                  round(np.mean(ffst) * 100, 2), round(percent_allF * 100, 2), round(percent_failed2 * 100, 2)]
#         data = {
#             Formula[t]: value
#         }
#         datadist.update(data)
#         if failedcase == 0:
#             return row
#
#     for i, j in datadist.items():  # i--公式名称, j--指标值
#         ws.cell(row, 1).value = i  # 添加第 1 列的数据
#         for col in range(2, len(j) + 2):  # values列表中索引
#             ws.cell(row, col).value = j[col - 2]
#         row += 1  # 行数
#     row += 2  # 行数
#     return row
#
#
# def getMetrics_v4(row, ws, mu, MG, pf, Output, EMR):
#     """
#     fv在每个test case的结果
#     """
#     Formula = ['Naish1', 'Naish2', 'Wong1', 'Russel&Rao', 'Binary', 'Jaccard', 'Anderberg', 'Sørensen-Dice', 'Dice',
#                'Goodman', 'Tarantula', 'qe', 'CBI Inc.', 'Wong2', 'Hamann', 'Simple Matching', 'Sokal',
#                'Rogers&Tanimoto', 'Hamming etc.', 'Euclid', 'Scott', 'Rogot1', 'Kulczynski2', 'Ochiai',
#                'M2', 'AMPLE2', 'Wong3', 'Arithmetic Mean', 'Cohen', 'Fleiss']
#     # 统计指标
#     datadist = {}
#     tablelist = {"Mutant" + str(mu): Formula}
#     datadist.update(tablelist)
#     failedcase = 0
#     Flag = []
#     count = []
#     var = []
#     for k in range(len(MG)):
#         # 如果全部failed，则标记
#         if pf[k][0] == 1:
#             failedcase += 1
#         flag, A = FaultTolerance17(MG[k], pf[k], Output[k])
#         Flag.append(flag)
#         c = group(Output[k])
#         if len(c) == 1:
#             c.append(0)
#             var.append(np.var(c))
#         else:
#             var.append(np.var(c))
#         # if 0 in c:
#         #     c.remove(0)
#         count.append(len(c))
#
#         data = {
#             'test case' + str(k): flag
#         }
#         datadist.update(data)
#     if failedcase == 0:
#         return row
#     for i, j in datadist.items():  # i--公式名称, j--指标值
#         ws.cell(row, 1).value = i  # 添加第 1 列的数据
#         for col in range(2, len(j) + 2):  # values列表中索引
#             ws.cell(row, col).value = j[col - 2]
#         row += 1  # 行数
#     row += 2  # 行数
#     return row
#
#
# def getMetrics_v5(row, ws, mu, MG, pf, Output, EMR):
#     """
#     去掉不等的MR
#     """
#     Formula = ['Naish1', 'Naish2', 'Wong1', 'Russel&Rao', 'Binary', 'Jaccard', 'Anderberg', 'Sørensen-Dice', 'Dice',
#                'Goodman', 'Tarantula', 'qe', 'CBI Inc.', 'Wong2', 'Hamann', 'Simple Matching', 'Sokal',
#                'Rogers&Tanimoto', 'Hamming etc.', 'Euclid', 'Scott', 'Rogot1', 'Kulczynski2', 'Ochiai',
#                'M2', 'AMPLE2', 'Wong3', 'Arithmetic Mean', 'Cohen', 'Fleiss']
#     # 统计指标
#     datadist = {}
#     # tablelist = {"Mutant" + str(mu): ['result', 'group size', 'var', 'random']}
#     tablelist = {"Mutant" + str(mu): Formula}
#     datadist.update(tablelist)
#     failedcase = 0
#     Flag = []
#     var = []
#     count = []
#     for k in range(len(MG)):
#         # 如果全部failed，则标记
#         if pf[k][0] == 1:
#             failedcase += 1
#         flag, A = FaultTolerance17(MG[k], pf[k], Output[k])
#         Flag.append(flag)
#         c = group(Output[k])
#         if len(c) == 1:
#             c.append(0)
#             var.append(np.var(c))
#         else:
#             var.append(np.var(c))
#         # if 0 in c:
#         #     c.remove(0)
#         count.append(len(c))
#         value = [flag[0], str(c), round(float(np.var(c)), 2), A[0]]
#         data = {
#             'test case' + str(k + 1): A
#         }
#         datadist.update(data)
#     if failedcase == 0:
#         return row
#     for i, j in datadist.items():  # i--公式名称, j--指标值
#         ws.cell(row, 1).value = i  # 添加第 1 列的数据
#         for col in range(2, len(j) + 2):  # values列表中索引
#             ws.cell(row, col).value = j[col - 2]
#         row += 1  # 行数
#     row += 2  # 行数
#     return row
#
#
# def getMetrics_v6(row, ws, mu, MG, pf, Output):
#     """
#     根据TY改, 全部MR
#     """
#     Formula = ['Naish1', 'Naish2', 'Wong1', 'Russel&Rao', 'Binary', 'Jaccard', 'Anderberg', 'Sørensen-Dice', 'Dice',
#                'Goodman', 'Tarantula', 'qe', 'CBI Inc.', 'Wong2', 'Hamann', 'Simple Matching', 'Sokal',
#                'Rogers&Tanimoto', 'Hamming etc.', 'Euclid', 'Scott', 'Rogot1', 'Kulczynski2', 'Ochiai',
#                'M2', 'AMPLE2', 'Wong3', 'Arithmetic Mean', 'Cohen', 'Fleiss']
#
#     # 统计指标
#     datadist = {}
#     tablelist = {"Mutant" + str(mu): ['identify(%)', 'failed case(%)', 'VMG(%)', 'SMG(%)',
#                                       'false satisfied MG(%)', 'FFSF', 'FFST', 'All Failed(%)',
#                                       'failed case2(%)', 'num', 'var', 'random']}
#     datadist.update(tablelist)
#
#     FFSlist = []
#     FFSF = []
#     FFST = []
#     failedcase = 0
#     failedcase2 = 0
#     VMGlist = []
#     SMGlist = []
#     allF = 0
#     tt = len(MG[0][0]) + 1
#     Flag = []
#     count = []
#     var = []
#     aa = []
#     nn = 0
#     for k in range(len(MG)):
#         A = []
#         for _ in range(len(Formula)):
#             A.append(0)
#         S_MG = 0
#         V_MG = 0
#         FFS = 0
#         for i in range(len(MG[k])):
#             S_MG += (MG[k][i].count(0) + MG[k][i].count(3))
#             FFS += MG[k][i].count(3)
#             V_MG += MG[k][i].count(1)
#         if S_MG == 0:
#             percent_FFS = 0
#         else:
#             percent_FFS = FFS / S_MG
#
#         # if percent_FFS > 0.3:
#         #     continue
#
#         # 如果全部failed，则标记
#         num = pf[k][0:tt].count(1)
#         if tt == num:
#             allF += 1
#         if pf[k][0] == 1:
#             failedcase += 1
#         failedcase2 += pf[k][0:tt].count(1)
#
#         FFSlist.append(percent_FFS)
#         VMGlist.append(V_MG / (S_MG + V_MG))
#         SMGlist.append(S_MG / (S_MG + V_MG))
#         c = group(Output[k])
#         if len(c) == 1:
#             c.append(0)
#         var.append(np.var(c))
#         # 策略一
#         # if np.var(c) >= len(MG[0][0]) / 2:
#         #     flag = FaultTolerance3(MG[k], pf[k], Output[k])
#         # else:
#         #     nn += 1
#         #     flag, A = FaultTolerance13(MG[k], pf[k], Output[k])
#         # 策略二
#         # m = showmax(Output[k][1:])
#         # if np.var(c) < 0.1 * (np.var([len(MG[0][0])+1, 0])) or len(m) > 1:
#         nn += 1
#         flag, A = FaultTolerance18(MG[k], pf[k], Output[k])
#         # else:
#         #     flag, A = FaultTolerance3(MG[k], pf[k], Output[k])
#         aa.append(A)
#         Flag.append(flag)
#         if 0 in c:
#             c.remove(0)
#         count.append(len(c))
#     for t in range(len(Flag[0])):
#         true = 0
#         nnn = 0
#         ffsf = []
#         ffst = []
#         a = []
#         for i in range(len(Flag)):
#             true += Flag[i][t]
#             nnn += aa[i][t]
#             if Flag[i][t] == 0:
#                 ffsf.append(FFSlist[i])
#             else:
#                 ffst.append(FFSlist[i])
#         if len(ffsf) == 0:
#             ffsf = 0
#         if len(ffst) == 0:
#             ffst = 0
#         FFSF.append(ffsf)
#         FFST.append(ffst)
#         percent_identify = true / len(Flag)
#         percent_allF = allF / len(Flag)
#         if nn == 0:
#             pp = 0
#         else:
#             pp = nnn / nn
#         percent_failed = failedcase / len(Flag)
#         percent_failed2 = failedcase2 / (len(Flag) * tt)
#         value = [round(percent_identify * 100, 2), round(percent_failed * 100, 2), round(np.mean(VMGlist) * 100, 2),
#                  round(np.mean(SMGlist) * 100, 2), round(np.mean(FFSlist) * 100, 2), round(np.mean(ffsf) * 100, 2),
#                  round(np.mean(ffst) * 100, 2), round(percent_allF * 100, 2), round(percent_failed2 * 100, 2),
#                  round(float(np.mean(count)), 2), round(float(np.mean(var)), 2), round(pp * 100, 2)]
#         data = {
#             Formula[t]: value
#         }
#         datadist.update(data)
#         if failedcase == 0:
#             return row
#
#     for i, j in datadist.items():  # i--公式名称, j--指标值
#         ws.cell(row, 1).value = i  # 添加第 1 列的数据
#         for col in range(2, len(j) + 2):  # values列表中索引
#             ws.cell(row, col).value = j[col - 2]
#         row += 1  # 行数
#     row += 2  # 行数
#     return row


def getMetrics_v7(row, ws, mu, MG, pf, Output, EMR):
    """
    final fr and fv
    voting random
    """
    Formula = ['Naish1', 'Naish2', 'Wong1', 'Russel&Rao', 'Binary', 'Jaccard', 'Anderberg', 'Sørensen-Dice', 'Dice',
               'Goodman', 'Tarantula', 'qe', 'CBI Inc.', 'Wong2', 'Hamann', 'Simple Matching', 'Sokal',
               'Rogers&Tanimoto', 'Hamming etc.', 'Euclid', 'Scott', 'Rogot1', 'Kulczynski2', 'Ochiai',
               'M2', 'AMPLE2', 'Wong3', 'Arithmetic Mean', 'Cohen', 'Fleiss']

    # 统计指标
    datadist = {}
    tablelist = {"Mutant" + str(mu): ['identify(%)', 'failed case(%)', 'VMG(%)', 'allFoV(%)', 'SMG(%)',
                                      'false satisfied MG(%)', 'FFSF', 'FFST', 'All Failed(%)',
                                      'failed case2(%)', 'num', 'var', 'random']}
    datadist.update(tablelist)
    FFSlist = []
    FFSF = []
    FFST = []
    failedcase = 0
    failedcase2 = 0
    VMGlist = []
    SMGlist = []
    allFlist = []
    allF = 0
    tt = len(MG[0][0]) + 1
    Flag = []
    var = []
    count = []
    aa = []
    nn = 0
    for k in range(len(MG)):
        A = []
        for _ in range(len(Formula)):
            A.append(0)
        S_MG = 0
        V_MG = 0
        FFS = 0
        allFoV = 0
        for i in range(len(MG[k])):
            S_MG += (MG[k][i].count(0) + MG[k][i].count(3))
            FFS += MG[k][i].count(3)
            V_MG += MG[k][i].count(1)
            for j in range(len(MG[k][i])):
                if MG[k][i][j] == 1 and pf[k][i] and pf[k][i * len(MG[0][0]) + j + 1]:
                    allFoV += 1

        if S_MG == 0:
            percent_FFS = 0
        else:
            percent_FFS = FFS / S_MG
        if V_MG == 0:
            percent_allFoV = 0
        else:
            percent_allFoV = allFoV / V_MG
            allFlist.append(percent_allFoV)
        # if percent_FFS > 0.3:
        #     continue

        # 如果全部failed，则标记
        if pf[k][0] == 1:
            num = 1
            failedcase += 1
            failedcase2 += 1
        else:
            num = 0
        for i in EMR:
            if pf[k][i + 1] == 1:
                num += 1
                failedcase2 += 1
        if tt == num:
            allF += 1

        FFSlist.append(percent_FFS)
        VMGlist.append(V_MG / (S_MG + V_MG))
        SMGlist.append(S_MG / (S_MG + V_MG))

        c = group(Output[k])
        if len(c) == 1:
            c.append(0)
        var.append(np.var(c))
        # 策略一
        # if np.var(c) >= len(EMR) / 2:
        #     flag = FaultTolerance4(MG[k], pf[k], Output[k], EMR)
        # else:
        #     nn += 1
        #     output = [Output[k][0], Output[k][1]]
        #     for j in range(len(MG[0][0])):
        #         if j not in EMR:
        #             continue
        #         output.append(Output[k][1:][j + 1])
        #     flag, A = FaultTolerance13(MG[k], pf[k], output)
        #
        # 策略二
        # output = [Output[k][0], Output[k][1]]
        # for j in range(len(MG[0][0])):
        #     if j not in EMR:
        #         continue
        #     output.append(Output[k][1:][j + 1])
        # m = showmax(output[1:])
        # if np.var(c) < 0.1 * (np.var([len(EMR)+1, 0])) or len(m) > 1:
        nn += 1
        flag, A = FaultTolerance4(MG[k], pf[k], Output[k])
        # else:
        #     flag, A = FaultTolerance3(MG[k], pf[k], output)
        if 0 in c:
            c.remove(0)
        count.append(len(c))
        Flag.append(flag)
        aa.append(A)

    low = []
    for n in range(100):  # 随机100次
        t = 0
        for k in range(len(MG)):
            output = Output[k][1:]
            output1 = output[1:]  # 只包含follow
            index_list = [a for a, b in enumerate(MG[k][0]) if b == 3]
            output1 = [n for i, n in enumerate(output1) if i not in index_list]
            output1.insert(0, output[0])
            output = output1
            a = random.choice(output)
            if a == Output[k][0]:
                t += 1
        low.append(t)

    if len(allFlist) == 0:
        allFlist.append(0)

    for t in range(len(Flag[0])):
        true = 0
        nnn = 0
        ffsf = []
        ffst = []
        for i in range(len(Flag)):
            true += Flag[i][t]
            nnn += aa[i][t]
            if Flag[i][t] == 0:
                ffsf.append(FFSlist[i])
            else:
                ffst.append(FFSlist[i])
        if len(ffsf) == 0:
            ffsf = 0
        if len(ffst) == 0:
            ffst = 0
        FFSF.append(ffsf)
        FFST.append(ffst)
        percent_identify = true / len(Flag)
        if nn == 0:
            pp = 0
        else:
            pp = nnn / nn
        percent_allF = allF / len(Flag)
        percent_failed = failedcase / len(Flag)
        percent_failed2 = failedcase2 / (len(Flag) * tt)
        value = [round(float(np.mean(low)), 2), round(percent_failed * 100, 2), round(np.mean(VMGlist) * 100, 2),
                 round(np.mean(allFlist) * 100, 2), round(np.mean(SMGlist) * 100, 2), round(np.mean(FFSlist) * 100, 2),
                 round(np.mean(ffsf) * 100, 2), round(np.mean(ffst) * 100, 2), round(percent_allF * 100, 2),
                 round(percent_failed2 * 100, 2), round(float(np.mean(count)), 2), round(float(np.mean(var)), 2),
                 round(pp * 100, 2)]
        data = {
            Formula[t]: value
        }
        datadist.update(data)
        if failedcase == 0:
            return row

    for i, j in datadist.items():  # i--公式名称, j--指标值
        ws.cell(row, 1).value = i  # 添加第 1 列的数据
        for col in range(2, len(j) + 2):  # values列表中索引
            ws.cell(row, col).value = j[col - 2]
        row += 1  # 行数
    row += 2  # 行数
    return row


# def getMetrics_v9(row, ws, mu, MG, pf, FR, Output):
#     """
#     根据TY改, 全部MR
#     """
#     Formula = ['Naish1', 'Naish2', 'Wong1', 'Russel&Rao', 'Binary', 'Jaccard', 'Anderberg', 'Sørensen-Dice', 'Dice',
#                'Goodman', 'Tarantula', 'qe', 'CBI Inc.', 'Wong2', 'Hamann', 'Simple Matching', 'Sokal',
#                'Rogers&Tanimoto', 'Hamming etc.', 'Euclid', 'Scott', 'Rogot1', 'Kulczynski2', 'Ochiai',
#                'M2', 'AMPLE2', 'Wong3', 'Arithmetic Mean', 'Cohen', 'Fleiss']
#
#     # 统计指标
#     datadist = {}
#     tablelist = {"Mutant" + str(mu): ['identify(%)', 'failed case(%)', 'VMG(%)', 'SMG(%)',
#                                       'false satisfied MG(%)', 'FFSF', 'FFST', 'All Failed(%)',
#                                       'failed case2(%)', 'a1', 'a11', 'a12', 'a2', 'a21', 'a22', 'a3',
#                                       'a31', 'a32', 'a4', 'a41', 'a42', 'a5', 'a51', 'a52', 'a6', 'a61', 'a62']}
#     datadist.update(tablelist)
#
#     FFSlist = []
#     FFSF = []
#     FFST = []
#     failedcase = 0
#     failedcase2 = 0
#     VMGlist = []
#     SMGlist = []
#     allF = 0
#     tt = len(MG[0][0]) + 1
#     Flag = []
#     AA = []
#     BB = []
#     for k in range(len(MG)):
#         S_MG = 0
#         V_MG = 0
#         FFS = 0
#         for i in range(len(MG[k])):
#             S_MG += (MG[k][i].count(0) + MG[k][i].count(3))
#             FFS += MG[k][i].count(3)
#             V_MG += MG[k][i].count(1)
#         if S_MG == 0:
#             percent_FFS = 0
#         else:
#             percent_FFS = FFS / S_MG
#
#         # if percent_FFS > 0.3:
#         #     continue
#
#         # 如果全部failed，则标记
#         num = pf[k][0:tt].count(1)
#         if tt == num:
#             allF += 1
#
#         if pf[k][0] == 1:
#             failedcase += 1
#         failedcase2 += pf[k][0:tt].count(1)
#
#         FFSlist.append(percent_FFS)
#         VMGlist.append(V_MG / (S_MG + V_MG))
#         SMGlist.append(S_MG / (S_MG + V_MG))
#
#         flag, aa, bb = FaultTolerance7(MG[k], pf[k], Output[k])
#         # flag = FaultTolerance3(MG[k], pf[k], Output[k])
#         Flag.append(flag)
#         AA.append(aa)
#         BB.append(bb)
#     for t in range(len(Flag[0])):
#         true = 0
#         ffsf = []
#         ffst = []
#         a = []
#         for i in range(len(Flag)):
#             true += Flag[i][t]
#             a.append(AA[i][t])
#             a.append(BB[i][t])
#             if Flag[i][t] == 0:
#                 ffsf.append(FFSlist[i])
#             else:
#                 ffst.append(FFSlist[i])
#         a1 = a.count(1)
#         a2 = a.count(2)  #
#         a3 = a.count(3)  # AA[t].count(3)
#         a4 = a.count(4)
#         a5 = a.count(5)  #
#         a6 = a.count(6)
#         a11 = a.count(11)
#         a12 = a.count(12)
#         a21 = a.count(21)
#         a22 = a.count(22)
#         a31 = a.count(31)
#         a32 = a.count(32)
#         a41 = a.count(41)
#         a42 = a.count(42)
#         a51 = a.count(51)
#         a52 = a.count(52)
#         a61 = a.count(61)
#         a62 = a.count(62)
#         if len(ffsf) == 0:
#             ffsf = 0
#         if len(ffst) == 0:
#             ffst = 0
#         FFSF.append(ffsf)
#         FFST.append(ffst)
#         percent_identify = true / len(Flag)
#         percent_allF = allF / len(Flag)
#         percent_failed = failedcase / len(Flag)
#         percent_failed2 = failedcase2 / (len(Flag) * tt)
#         value = [round(percent_identify * 100, 2), round(percent_failed * 100, 2), round(np.mean(VMGlist) * 100, 2),
#                  round(np.mean(SMGlist) * 100, 2), round(np.mean(FFSlist) * 100, 2), round(np.mean(ffsf) * 100, 2),
#                  round(np.mean(ffst) * 100, 2), round(percent_allF * 100, 2), round(percent_failed2 * 100, 2),
#                  a1, a11, a12, a2, a21, a22, a3,
#                  a31, a32, a4, a41, a42, a5, a51, a52, a6, a61, a62]
#         data = {
#             Formula[t]: value
#         }
#         datadist.update(data)
#         if failedcase == 0:
#             return row
#
#     for i, j in datadist.items():  # i--公式名称, j--指标值
#         ws.cell(row, 1).value = i  # 添加第 1 列的数据
#         for col in range(2, len(j) + 2):  # values列表中索引
#             ws.cell(row, col).value = j[col - 2]
#         row += 1  # 行数
#     row += 2  # 行数
#     return row
#
#
# def getMetrics_v10(row, ws, mu, MG, pf, FR, Output, EMR):
#     """
#     去掉不等的MR
#     """
#     Formula = ['Naish1', 'Naish2', 'Wong1', 'Russel&Rao', 'Binary', 'Jaccard', 'Anderberg', 'Sørensen-Dice', 'Dice',
#                'Goodman', 'Tarantula', 'qe', 'CBI Inc.', 'Wong2', 'Hamann', 'Simple Matching', 'Sokal',
#                'Rogers&Tanimoto', 'Hamming etc.', 'Euclid', 'Scott', 'Rogot1', 'Kulczynski2', 'Ochiai',
#                'M2', 'AMPLE2', 'Wong3', 'Arithmetic Mean', 'Cohen', 'Fleiss']
#
#     # 统计指标
#     datadist = {}
#     tablelist = {"Mutant" + str(mu): ['identify(%)', 'failed case(%)', 'VMG(%)', 'SMG(%)',
#                                       'false satisfied MG(%)', 'FFSF', 'FFST', 'All Failed(%)',
#                                       'failed case2(%)', 'a1', 'a11', 'a12', 'a2', 'a21', 'a22', 'a3',
#                                       'a31', 'a32', 'a4', 'a41', 'a42', 'a5', 'a51', 'a52', 'a6', 'a61', 'a62']}
#     datadist.update(tablelist)
#     FFSlist = []
#     FFSF = []
#     FFST = []
#     failedcase = 0
#     failedcase2 = 0
#     VMGlist = []
#     SMGlist = []
#     allF = 0
#     tt1 = len(EMR) + 1
#     Flag = []
#     AA = []
#     BB = []
#     for k in range(len(MG)):
#         S_MG = 0
#         V_MG = 0
#         FFS = 0
#         for i in range(len(MG[k])):
#             S_MG += (MG[k][i].count(0) + MG[k][i].count(3))
#             FFS += MG[k][i].count(3)
#             V_MG += MG[k][i].count(1)
#         if S_MG == 0:
#             percent_FFS = 0
#         else:
#             percent_FFS = FFS / S_MG
#
#         # if percent_FFS > 0.3:
#         #     continue
#
#         # 如果全部failed，则标记
#         if pf[k][0] == 1:
#             num = 1
#             failedcase += 1
#             failedcase2 += 1
#         else:
#             num = 0
#         for i in EMR:
#             if pf[k][i + 1] == 1:
#                 num += 1
#                 failedcase2 += 1
#         if tt1 == num:
#             allF += 1
#
#         FFSlist.append(percent_FFS)
#         VMGlist.append(V_MG / (S_MG + V_MG))
#         SMGlist.append(S_MG / (S_MG + V_MG))
#
#         flag, aa, bb = FaultTolerance8(MG[k], pf[k], Output[k], EMR)
#         # flag = FaultTolerance4(MG[k], pf[k], Output[k], EMR)
#         Flag.append(flag)
#         AA.append(aa)
#         BB.append(bb)
#     for t in range(len(Flag[0])):
#         true = 0
#         ffsf = []
#         ffst = []
#         a = []
#         for i in range(len(Flag)):
#             true += Flag[i][t]
#             a.append(AA[i][t])
#             a.append(BB[i][t])
#             if Flag[i][t] == 0:
#                 ffsf.append(FFSlist[i])
#             else:
#                 ffst.append(FFSlist[i])
#         a1 = a.count(1)
#         a2 = a.count(2)  #
#         a3 = a.count(3)  # AA[t].count(3)
#         a4 = a.count(4)
#         a5 = a.count(5)  #
#         a6 = a.count(6)
#         a11 = a.count(11)
#         a12 = a.count(12)
#         a21 = a.count(21)
#         a22 = a.count(22)
#         a31 = a.count(31)
#         a32 = a.count(32)
#         a41 = a.count(41)
#         a42 = a.count(42)
#         a51 = a.count(51)
#         a52 = a.count(52)
#         a61 = a.count(61)
#         a62 = a.count(62)
#         if len(ffsf) == 0:
#             ffsf = 0
#         if len(ffst) == 0:
#             ffst = 0
#         FFSF.append(ffsf)
#         FFST.append(ffst)
#         percent_identify = true / len(Flag)
#         percent_allF = allF / len(Flag)
#         percent_failed = failedcase / len(Flag)
#         percent_failed2 = failedcase2 / (len(Flag) * tt1)
#         value = [round(percent_identify * 100, 2), round(percent_failed * 100, 2), round(np.mean(VMGlist) * 100, 2),
#                  round(np.mean(SMGlist) * 100, 2), round(np.mean(FFSlist) * 100, 2), round(np.mean(ffsf) * 100, 2),
#                  round(np.mean(ffst) * 100, 2), round(percent_allF * 100, 2), round(percent_failed2 * 100, 2),
#                  a1, a11, a12, a2, a21, a22, a3,
#                  a31, a32, a4, a41, a42, a5, a51, a52, a6, a61, a62]
#         data = {
#             Formula[t]: value
#         }
#         datadist.update(data)
#         if failedcase == 0:
#             return row
#
#     for i, j in datadist.items():  # i--公式名称, j--指标值
#         ws.cell(row, 1).value = i  # 添加第 1 列的数据
#         for col in range(2, len(j) + 2):  # values列表中索引
#             ws.cell(row, col).value = j[col - 2]
#         row += 1  # 行数
#     row += 2  # 行数
#     return row
#
#
# def getMetrics_v11(row, ws, mu, MG, pf, FR, Output):
#     """
#     根据TY改, 全部MR
#     """
#     Formula = ['Naish1', 'Naish2', 'Wong1', 'Russel&Rao', 'Binary', 'Jaccard', 'Anderberg', 'Sørensen-Dice', 'Dice',
#                'Goodman', 'Tarantula', 'qe', 'CBI Inc.', 'Wong2', 'Hamann', 'Simple Matching', 'Sokal',
#                'Rogers&Tanimoto', 'Hamming etc.', 'Euclid', 'Scott', 'Rogot1', 'Kulczynski2', 'Ochiai',
#                'M2', 'AMPLE2', 'Wong3', 'Arithmetic Mean', 'Cohen', 'Fleiss']
#
#     # 统计指标
#     datadist = {}
#     tablelist = {"Mutant" + str(mu): ['identify(%)', 'failed case(%)', 'VMG(%)', 'SMG(%)',
#                                       'false satisfied MG(%)', 'FFSF', 'FFST', 'All Failed(%)',
#                                       'failed case2(%)', 'a1', 'a11', 'a12', 'a2', 'a21', 'a22', 'a3',
#                                       'a31', 'a32', 'a4', 'a41', 'a42', 'a5', 'a51', 'a52', 'a6', 'a61', 'a62']}
#     datadist.update(tablelist)
#
#     FFSlist = []
#     FFSF = []
#     FFST = []
#     failedcase = 0
#     failedcase2 = 0
#     VMGlist = []
#     SMGlist = []
#     allF = 0
#     tt = len(MG[0][0]) + 1
#     Flag = []
#     AA = []
#     BB = []
#     for k in range(len(MG)):
#         S_MG = 0
#         V_MG = 0
#         FFS = 0
#         for i in range(len(MG[k])):
#             S_MG += (MG[k][i].count(0) + MG[k][i].count(3))
#             FFS += MG[k][i].count(3)
#             V_MG += MG[k][i].count(1)
#         if S_MG == 0:
#             percent_FFS = 0
#         else:
#             percent_FFS = FFS / S_MG
#
#         # if percent_FFS > 0.3:
#         #     continue
#
#         # 如果全部failed，则标记
#         num = pf[k][0:tt].count(1)
#         if tt == num:
#             allF += 1
#
#         if pf[k][0] == 1:
#             failedcase += 1
#         failedcase2 += pf[k][0:tt].count(1)
#
#         FFSlist.append(percent_FFS)
#         VMGlist.append(V_MG / (S_MG + V_MG))
#         SMGlist.append(S_MG / (S_MG + V_MG))
#
#         flag, aa, bb = FaultTolerance9(MG[k], pf[k], Output[k])
#         # flag = FaultTolerance6(MG[k], pf[k], Output[k])
#         Flag.append(flag)
#         AA.append(aa)
#         BB.append(bb)
#     for t in range(len(Flag[0])):
#         true = 0
#         ffsf = []
#         ffst = []
#         a = []
#         for i in range(len(Flag)):
#             true += Flag[i][t]
#             a.append(AA[i][t])
#             a.append(BB[i][t])
#             if Flag[i][t] == 0:
#                 ffsf.append(FFSlist[i])
#             else:
#                 ffst.append(FFSlist[i])
#         a1 = a.count(1)
#         a2 = a.count(2)  #
#         a3 = a.count(3)  # AA[t].count(3)
#         a4 = a.count(4)
#         a5 = a.count(5)  #
#         a6 = a.count(6)
#         a11 = a.count(11)
#         a12 = a.count(12)
#         a21 = a.count(21)
#         a22 = a.count(22)
#         a31 = a.count(31)
#         a32 = a.count(32)
#         a41 = a.count(41)
#         a42 = a.count(42)
#         a51 = a.count(51)
#         a52 = a.count(52)
#         a61 = a.count(61)
#         a62 = a.count(62)
#         if len(ffsf) == 0:
#             ffsf = 0
#         if len(ffst) == 0:
#             ffst = 0
#         FFSF.append(ffsf)
#         FFST.append(ffst)
#         percent_identify = true / len(Flag)
#         percent_allF = allF / len(Flag)
#         percent_failed = failedcase / len(Flag)
#         percent_failed2 = failedcase2 / (len(Flag) * tt)
#         value = [round(percent_identify * 100, 2), round(percent_failed * 100, 2), round(np.mean(VMGlist) * 100, 2),
#                  round(np.mean(SMGlist) * 100, 2), round(np.mean(FFSlist) * 100, 2), round(np.mean(ffsf) * 100, 2),
#                  round(np.mean(ffst) * 100, 2), round(percent_allF * 100, 2), round(percent_failed2 * 100, 2),
#                  a1, a11, a12, a2, a21, a22, a3,
#                  a31, a32, a4, a41, a42, a5, a51, a52, a6, a61, a62]
#         data = {
#             Formula[t]: value
#         }
#         datadist.update(data)
#         if failedcase == 0:
#             return row
#
#     for i, j in datadist.items():  # i--公式名称, j--指标值
#         ws.cell(row, 1).value = i  # 添加第 1 列的数据
#         for col in range(2, len(j) + 2):  # values列表中索引
#             ws.cell(row, col).value = j[col - 2]
#         row += 1  # 行数
#     row += 2  # 行数
#     return row


def getMetrics_v12(row, ws, mu, MG, pf, Output, EMR):
    """
    upper bound and random
    """
    Formula = ['Naish1', 'Naish2', 'Wong1', 'Russel&Rao', 'Binary', 'Jaccard', 'Anderberg', 'Sørensen-Dice', 'Dice',
               'Goodman', 'Tarantula', 'qe', 'CBI Inc.', 'Wong2', 'Hamann', 'Simple Matching', 'Sokal',
               'Rogers&Tanimoto', 'Hamming etc.', 'Euclid', 'Scott', 'Rogot1', 'Kulczynski2', 'Ochiai',
               'M2', 'AMPLE2', 'Wong3', 'Arithmetic Mean', 'Cohen', 'Fleiss']

    # 统计指标
    datadist = {}
    tablelist = {"Mutant" + str(mu): ['upper', 'RM', 'failed case(%)', 'VMG(%)', 'SMG(%)',
                                      'false satisfied MG(%)', 'All Failed(%)', 'failed case2(%)']}
    datadist.update(tablelist)
    FFSlist = []
    failedcase = 0
    failedcase2 = 0
    VMGlist = []
    SMGlist = []
    allF = 0
    tt = len(MG[0][0]) + 1
    for k in range(len(MG)):
        S_MG = 0
        V_MG = 0
        FFS = 0
        for i in range(len(MG[k])):
            S_MG += (MG[k][i].count(0) + MG[k][i].count(3))
            FFS += MG[k][i].count(3)
            V_MG += MG[k][i].count(1)
        if S_MG == 0:
            percent_FFS = 0
        else:
            percent_FFS = FFS / S_MG
        # 如果全部failed，则标记
        if pf[k][0] == 1:
            num = 1
            failedcase += 1
            failedcase2 += 1
        else:
            num = 0
        for i in EMR:
            if pf[k][i + 1] == 1:
                num += 1
                failedcase2 += 1
        if tt == num:
            allF += 1
        FFSlist.append(percent_FFS)
        VMGlist.append(V_MG / (S_MG + V_MG))
        SMGlist.append(S_MG / (S_MG + V_MG))
    low = []
    for n in range(100):  # 随机100次
        t = 0
        for k in range(len(MG)):
            a = random.choice(Output[k][1:])
            if a == Output[k][0]:
                t += 1
        low.append(t)

    for t in range(30):
        percent_allF = allF / len(MG)
        percent_upper = (100 - allF) / len(MG)
        percent_failed = failedcase / len(MG)
        percent_failed2 = failedcase2 / (len(MG) * tt)
        value = [round(percent_upper * 100, 2), round(float(np.mean(low)), 2), round(percent_failed * 100, 2),
                 round(np.mean(VMGlist) * 100, 2),
                 round(np.mean(SMGlist) * 100, 2), round(np.mean(FFSlist) * 100, 2), round(percent_allF * 100, 2),
                 round(percent_failed2 * 100, 2)]
        data = {
            Formula[t]: value
        }
        datadist.update(data)
        if failedcase == 0:
            return row

    for i, j in datadist.items():  # i--公式名称, j--指标值
        ws.cell(row, 1).value = i  # 添加第 1 列的数据
        for col in range(2, len(j) + 2):  # values列表中索引
            ws.cell(row, col).value = j[col - 2]
        row += 1  # 行数
    row += 2  # 行数
    return row


# def getMetrics_v13(row, ws, mu, MG, pf, Output, EMR):
#     """
#     final voting
#     """
#     Formula = ['Naish1', 'Naish2', 'Wong1', 'Russel&Rao', 'Binary', 'Jaccard', 'Anderberg', 'Sørensen-Dice', 'Dice',
#                'Goodman', 'Tarantula', 'qe', 'CBI Inc.', 'Wong2', 'Hamann', 'Simple Matching', 'Sokal',
#                'Rogers&Tanimoto', 'Hamming etc.', 'Euclid', 'Scott', 'Rogot1', 'Kulczynski2', 'Ochiai',
#                'M2', 'AMPLE2', 'Wong3', 'Arithmetic Mean', 'Cohen', 'Fleiss']
#
#     # 统计指标
#     datadist = {}
#     tablelist = {"Mutant" + str(mu): ['identify(%)', 'failed case(%)', 'VMG(%)', 'SMG(%)',
#                                       'false satisfied MG(%)', 'false satisfied MG2(%)', 'FSF',
#                                       'FST', 'All Failed(%)', 'All Correct(%)', 'failed case2(%)',
#                                       'num', 'var', 'random']}
#     datadist.update(tablelist)
#     FFSlist = []
#     FFSlist2 = []
#     FFSF = []
#     FFST = []
#     failedcase = 0
#     failedcase2 = 0
#     VMGlist = []
#     SMGlist = []
#     allF = 0
#     allC = 0
#     tt = len(MG[0][0]) + 1
#     Flag = []
#     var = []
#     count = []
#     aa = []
#     nn = 0
#     for k in range(len(MG)):
#         A = []
#         for _ in range(len(Formula)):
#             A.append(0)
#         S_MG = 0
#         V_MG = 0
#         FFS = 0
#         for i in range(len(MG[k])):
#             S_MG += (MG[k][i].count(0) + MG[k][i].count(3))
#             FFS += MG[k][i].count(3)
#             V_MG += MG[k][i].count(1)
#         if S_MG == 0:
#             percent_FFS = 0
#         else:
#             percent_FFS = FFS / S_MG
#
#         S_MG2 = MG[k][0].count(0) + MG[k][0].count(3)
#         FFS2 = MG[k][0].count(3)
#         if S_MG2 == 0:
#             percent_FFS2 = 0
#         else:
#             percent_FFS2 = FFS2 / S_MG2
#
#         # if percent_FFS > 0.3:
#         #     continue
#
#         # 如果全部failed，则标记
#         if pf[k][0] == 1:
#             num = 1
#             failedcase += 1
#             failedcase2 += 1
#         else:
#             num = 0
#         for i in EMR:
#             if pf[k][i + 1] == 1:
#                 num += 1
#                 failedcase2 += 1
#         if tt == num:
#             allF += 1
#         if num == 0:
#             allC += 1
#
#         FFSlist.append(percent_FFS)
#         FFSlist2.append(percent_FFS2)
#         VMGlist.append(V_MG / (S_MG + V_MG))
#         SMGlist.append(S_MG / (S_MG + V_MG))
#         c = group(Output[k])
#         if len(c) == 1:
#             c.append(0)
#         var.append(np.var(c))
#         # 策略一
#         # if np.var(c) >= len(EMR) / 2:
#         #     flag = FaultTolerance4(MG[k], pf[k], Output[k], EMR)
#         # else:
#         #     nn += 1
#         #     output = [Output[k][0], Output[k][1]]
#         #     for j in range(len(MG[0][0])):
#         #         if j not in EMR:
#         #             continue
#         #         output.append(Output[k][1:][j + 1])
#         #     flag, A = FaultTolerance13(MG[k], pf[k], output)
#         #
#         # 策略二
#         # output = [Output[k][0], Output[k][1]]
#         # for j in range(len(MG[0][0])):
#         #     if j not in EMR:
#         #         continue
#         #     output.append(Output[k][1:][j + 1])
#         # m = showmax(output[1:])
#         # if np.var(c) < 0.1 * (np.var([len(EMR)+1, 0])) or len(m) > 1:
#         nn += 1
#         flag, A = FaultTolerance3(MG[k], pf[k], Output[k])
#         # else:
#         #     flag, A = FaultTolerance3(MG[k], pf[k], output)
#         if 0 in c:
#             c.remove(0)
#         count.append(len(c))
#         Flag.append(flag)
#         aa.append(A)
#     for t in range(len(Flag[0])):
#         true = 0
#         nnn = 0
#         ffsf = []
#         ffst = []
#         for i in range(len(Flag)):
#             true += Flag[i][t]
#             nnn += aa[i][t]
#             if Flag[i][t] == 0:
#                 ffsf.append(FFSlist2[i])
#             else:
#                 ffst.append(FFSlist2[i])
#         if len(ffsf) == 0:
#             ffsf = 0
#         if len(ffst) == 0:
#             ffst = 0
#         FFSF.append(ffsf)
#         FFST.append(ffst)
#         percent_identify = true / len(Flag)
#         if nn == 0:
#             pp = 0
#         else:
#             pp = nnn / nn
#         percent_allF = allF / len(Flag)
#         percent_allC = allC / len(Flag)
#         percent_failed = failedcase / len(Flag)
#         percent_failed2 = failedcase2 / (len(Flag) * tt)
#         value = [round(percent_identify * 100, 2), round(percent_failed * 100, 2), round(np.mean(VMGlist) * 100, 2),
#                  round(np.mean(SMGlist) * 100, 2), round(np.mean(FFSlist) * 100, 2), round(np.mean(FFSlist2) * 100, 2),
#                  round(np.mean(ffsf) * 100, 2), round(np.mean(ffst) * 100, 2),
#                  round(percent_allF * 100, 2), round(percent_allC * 100, 2), round(percent_failed2 * 100, 2),
#                  round(float(np.mean(count)), 2), round(float(np.mean(var)), 2), round(pp * 100, 2)]
#         data = {
#             Formula[t]: value
#         }
#         datadist.update(data)
#         if failedcase == 0:
#             return row
#
#     for i, j in datadist.items():  # i--公式名称, j--指标值
#         ws.cell(row, 1).value = i  # 添加第 1 列的数据
#         for col in range(2, len(j) + 2):  # values列表中索引
#             ws.cell(row, col).value = j[col - 2]
#         row += 1  # 行数
#     row += 2  # 行数
#     return row


def getMetrics_v14(row, ws, mu, MG, pf, Output, EMR):
    """
    FAILTIM 取最小可疑度
    """
    Formula = ['Naish1', 'Naish2', 'Wong1', 'Russel&Rao', 'Binary', 'Jaccard', 'Anderberg', 'Sørensen-Dice', 'Dice',
               'Goodman', 'Tarantula', 'qe', 'CBI Inc.', 'Wong2', 'Hamann', 'Simple Matching', 'Sokal',
               'Rogers&Tanimoto', 'Hamming etc.', 'Euclid', 'Scott', 'Rogot1', 'Kulczynski2', 'Ochiai',
               'M2', 'AMPLE2', 'Wong3', 'Arithmetic Mean', 'Cohen', 'Fleiss', 'New']

    # 统计指标
    datadist = {}
    tablelist = {"Mutant" + str(mu): ['identify(%)', 'failed case(%)', 'VMG(%)', 'allFoV(%)', 'allFoV2(%)', 'SMG(%)',
                                      'false satisfied MG(%)', 'FFSF', 'FFST', 'All Failed(%)',
                                      'failed case2(%)', 'num', 'var', 'random']}
    datadist.update(tablelist)
    FFSlist = []
    FFSF = []
    FFST = []
    failedcase = 0
    failedcase2 = 0
    VMGlist = []
    SMGlist = []
    allFlist = []
    allFlist2 = []
    allF = 0
    tt = len(MG[0][0]) + 1
    Flag = []
    var = []
    count = []
    aa = []
    nn = 0
    for k in range(len(MG)):
        A = []
        for _ in range(len(Formula)):
            A.append(0)
        S_MG = 0
        V_MG = 0
        FFS = 0
        allFoV = 0
        allFoV2 = 0
        V_MG2 = 0
        for i in range(len(MG[k])):
            S_MG += (MG[k][i].count(0) + MG[k][i].count(3))
            FFS += MG[k][i].count(3)
            V_MG += MG[k][i].count(1)
            for j in range(len(MG[k][i])):
                if MG[k][i][j] == 1 and pf[k][i] and pf[k][i * len(MG[0][0]) + j + 1]:
                    allFoV += 1
            if i == 0:
                V_MG2 += MG[k][i].count(1)
                for j in range(len(MG[k][i])):
                    if MG[k][i][j] == 1 and pf[k][i] and pf[k][i * len(MG[0][0]) + j + 1]:
                        allFoV2 += 1

        if S_MG == 0:
            percent_FFS = 0
        else:
            percent_FFS = FFS / S_MG
        if V_MG == 0:
            pass
        else:
            percent_allFoV = allFoV / V_MG
            allFlist.append(percent_allFoV)
        if V_MG2 == 0:
            pass
        else:
            percent_allFoV2 = allFoV2 / V_MG2
            allFlist2.append(percent_allFoV2)

        # 如果全部failed，则标记
        if pf[k][0] == 1:
            num = 1
            failedcase += 1
            failedcase2 += 1
        else:
            num = 0
        for i in EMR:
            if pf[k][i + 1] == 1:
                num += 1
                failedcase2 += 1
        if tt == num:
            allF += 1
        FFSlist.append(percent_FFS)
        VMGlist.append(V_MG / (S_MG + V_MG))
        SMGlist.append(S_MG / (S_MG + V_MG))
        c = group(Output[k])
        if len(c) == 1:
            c.append(0)
        var.append(np.var(c))
        nn += 1
        flag, A = FaultTolerance17(MG[k], pf[k], Output[k])
        if 0 in c:
            c.remove(0)
        count.append(len(c))
        Flag.append(flag)
        aa.append(A)

    if len(allFlist) == 0:
        allFlist.append(0)

    for t in range(len(Flag[0])):
        true = 0
        nnn = 0
        ffsf = []
        ffst = []
        for i in range(len(Flag)):
            true += Flag[i][t]
            nnn += aa[i][t]
            if Flag[i][t] == 0:
                ffsf.append(FFSlist[i])
            else:
                ffst.append(FFSlist[i])
        if len(ffsf) == 0:
            ffsf = 0
        if len(ffst) == 0:
            ffst = 0
        FFSF.append(ffsf)
        FFST.append(ffst)
        percent_identify = true / len(Flag)
        if nn == 0:
            pp = 0
        else:
            pp = nnn / nn
        percent_allF = allF / len(Flag)
        percent_failed = failedcase / len(Flag)
        percent_failed2 = failedcase2 / (len(Flag) * tt)
        value = [round(percent_identify * 100, 2), round(percent_failed * 100, 2), round(np.mean(VMGlist) * 100, 2),
                 round(np.mean(allFlist) * 100, 2), round(np.mean(allFlist2) * 100, 2), round(np.mean(SMGlist) * 100, 2),
                 round(np.mean(FFSlist) * 100, 2),
                 round(np.mean(ffsf) * 100, 2), round(np.mean(ffst) * 100, 2), round(percent_allF * 100, 2),
                 round(percent_failed2 * 100, 2), round(float(np.mean(count)), 2), round(float(np.mean(var)), 2),
                 round(pp * 100, 2)]
        data = {
            Formula[t]: value
        }
        datadist.update(data)
        if failedcase == 0:
            return row

    for i, j in datadist.items():  # i--公式名称, j--指标值
        ws.cell(row, 1).value = i  # 添加第 1 列的数据
        for col in range(2, len(j) + 2):  # values列表中索引
            ws.cell(row, col).value = j[col - 2]
        row += 1  # 行数
    row += 2  # 行数
    return row
